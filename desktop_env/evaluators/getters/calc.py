import csv
import openpyxl
from .file import get_vm_file


# I want to write a function, reads a csv file, and get all the contents in the third column in the order of rows
def get_conference_city_in_order(env, config):
    # read the csv file
    csv_path = config['csv_path']
    print(f"Reading csv file from {csv_path}")
    with open(csv_path, 'r') as f:
        reader = csv.reader(f)
    # skip the header row
    next(reader)
    # get the third column in the order of rows
    conference_city_list = [row[2] for row in reader]
    return conference_city_list


def check_second_row_deleted(result_file_path, expected_file_path):
    """
    检查 Excel 文件中的第二行（第一个数据行）是否被删除
    
    Args:
        result_file_path: 结果文件路径
        expected_file_path: 期望的文件路径（已删除第二行的版本）
    
    Returns:
        float: 如果第二行被正确删除返回 1.0，否则返回 0.0
    """
    try:
        # 加载结果文件
        result_wb = openpyxl.load_workbook(result_file_path)
        result_ws = result_wb.active
        
        # 加载期望文件
        expected_wb = openpyxl.load_workbook(expected_file_path)
        expected_ws = expected_wb.active
        
        # 获取结果文件的所有行数据
        result_rows = list(result_ws.iter_rows(values_only=True))
        expected_rows = list(expected_ws.iter_rows(values_only=True))
        
        # 检查行数是否相等（结果应该比原始少一行）
        if len(result_rows) != len(expected_rows):
            print(f"行数不匹配: 结果有 {len(result_rows)} 行，期望有 {len(expected_rows)} 行")
            return 0.0
        
        # 逐行比较数据
        for i, (result_row, expected_row) in enumerate(zip(result_rows, expected_rows)):
            if result_row != expected_row:
                print(f"第 {i+1} 行数据不匹配:")
                print(f"  结果: {result_row}")
                print(f"  期望: {expected_row}")
                return 0.0
        
        print("✓ 第二行已成功删除，数据验证通过")
        return 1.0
        
    except Exception as e:
        print(f"评估出错: {e}")
        return 0.0


def verify_row_count_decreased(result_file_path, original_row_count):
    """
    验证 Excel 文件的行数是否减少了指定数量
    
    Args:
        result_file_path: 结果文件路径
        original_row_count: 原始文件的行数
    
    Returns:
        float: 如果行数正确减少返回 1.0，否则返回 0.0
    """
    try:
        wb = openpyxl.load_workbook(result_file_path)
        ws = wb.active
        
        # 获取当前行数
        current_row_count = ws.max_row
        
        # 检查行数是否减少了 1
        if current_row_count == original_row_count - 1:
            print(f"✓ 行数验证通过: {original_row_count} → {current_row_count}")
            return 1.0
        else:
            print(f"✗ 行数验证失败: 期望 {original_row_count - 1} 行，实际 {current_row_count} 行")
            return 0.0
            
    except Exception as e:
        print(f"评估出错: {e}")
        return 0.0


def verify_second_row_deleted_without_gold(env, config):
    """
    验证 Excel 文件的第二行是否被删除（不需要金标准文件）
    
    通过以下方式验证：
    1. 首先从VM获取结果文件到宿主机
    2. 检查结果文件的行数是否比原始文件少1
    3. 检查原始文件的第二行数据是否在结果文件中不存在
    
    Args:
        env: 环境对象
        config: 配置字典，应包含：
            - path: VM中的结果文件路径
            - dest: 保存在宿主机的文件名
            - original_file_url: 原始文件的URL（用于下载和比对）
    
    Returns:
        float: 如果验证通过返回 1.0，否则返回 0.0
    """
    try:
        import tempfile
        import urllib.request
        import os
        
        # 首先从VM获取结果文件到宿主机
        vm_file_path = config.get('path', '/home/user/SalesRep.xlsx')
        dest_filename = config.get('dest', os.path.basename(vm_file_path))
        
        print(f"开始验证删除第二行任务...")
        print(f"从VM获取文件: {vm_file_path}")
        
        # 使用get_vm_file从VM获取文件
        result_file_path = get_vm_file(env, {
            'path': vm_file_path,
            'dest': dest_filename
        })
        
        if result_file_path is None:
            print(f"✗ 无法从VM获取文件: {vm_file_path}")
            return 0.0
        
        print(f"结果文件已保存到宿主机: {result_file_path}")
        
        original_file_url = config.get('original_file_url', '')
        print(f"原始文件URL: {original_file_url}")
        
        # 下载原始文件到临时位置
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            original_file_temp = tmp_file.name
        
        try:
            print(f"正在下载原始文件到临时位置: {original_file_temp}")
            urllib.request.urlretrieve(original_file_url, original_file_temp)
        except Exception as e:
            print(f"下载原始文件失败: {e}")
            # 如果下载失败，尝试从本地缓存读取
            cache_path = config.get('original_file_cache', '')
            if cache_path and os.path.exists(cache_path):
                print(f"使用缓存文件: {cache_path}")
                original_file_temp = cache_path
            else:
                print("无法获取原始文件")
                return 0.0
        
        # 加载原始文件
        print("加载原始文件...")
        original_wb = openpyxl.load_workbook(original_file_temp)
        original_ws = original_wb.active
        
        # 获取原始文件的所有行
        original_rows = list(original_ws.iter_rows(values_only=True))
        original_row_count = len(original_rows)
        
        if original_row_count < 2:
            print(f"✗ 原始文件行数不足: {original_row_count}（需要至少2行）")
            return 0.0
        
        # 保存第二行的数据（索引为1）
        second_row_data = original_rows[1]
        print(f"原始文件行数: {original_row_count}")
        print(f"原始文件第二行数据: {second_row_data}")
        
        # 加载结果文件（现在这个路径是宿主机路径，一定存在）
        print(f"加载结果文件...")
        if not os.path.exists(result_file_path):
            print(f"✗ 结果文件不存在: {result_file_path}")
            return 0.0
        
        result_wb = openpyxl.load_workbook(result_file_path)
        result_ws = result_wb.active
        
        # 获取结果文件的所有行
        result_rows = list(result_ws.iter_rows(values_only=True))
        result_row_count = len(result_rows)
        
        print(f"结果文件行数: {result_row_count}")
        
        # 验证1: 检查行数是否减少了1
        if result_row_count != original_row_count - 1:
            print(f"✗ 行数验证失败: 期望 {original_row_count - 1} 行，实际 {result_row_count} 行")
            return 0.0
        else:
            print(f"✓ 行数验证通过: {original_row_count} → {result_row_count}")
        
        # 验证2: 检查原始第二行是否存在于结果文件中
        second_row_exists = False
        for i, row in enumerate(result_rows):
            if row == second_row_data:
                print(f"✗ 原始第二行数据仍存在于结果文件的第 {i+1} 行")
                second_row_exists = True
                break
        
        if second_row_exists:
            return 0.0
        else:
            print(f"✓ 原始第二行数据已从结果文件中删除")
        
        # 验证3: 检查其他行是否保持不变（第一行和第3行之后）
        # 结果文件的第一行应该等于原始文件的第一行
        if result_rows[0] != original_rows[0]:
            print(f"✗ 第一行数据不匹配")
            print(f"  原始: {original_rows[0]}")
            print(f"  结果: {result_rows[0]}")
            return 0.0
        
        # 结果文件的第2行及之后应该等于原始文件的第3行及之后
        for i in range(1, result_row_count):
            if result_rows[i] != original_rows[i+1]:
                print(f"✗ 第 {i+1} 行数据不匹配")
                print(f"  期望（原始第 {i+2} 行）: {original_rows[i+1]}")
                print(f"  实际: {result_rows[i]}")
                return 0.0
        
        print(f"✓ 其他行数据保持不变")
        
        # 清理临时文件
        if original_file_temp != config.get('original_file_cache', ''):
            try:
                os.unlink(original_file_temp)
            except:
                pass
        
        print("=" * 60)
        print("✓ 所有验证通过！第二行已成功删除")
        print("=" * 60)
        return 1.0
        
    except Exception as e:
        import traceback
        print(f"✗ 评估出错: {e}")
        traceback.print_exc()
        return 0.0
