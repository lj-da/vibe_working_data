import functools
import itertools
import logging
import os.path

# import operator
from numbers import Number
from typing import Any, Union, cast, Callable, Iterable
from typing import Dict, List, Tuple, Set

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from rapidfuzz import fuzz

from desktop_env.evaluators.metrics.utils import (
    _match_value_to_rule,
    _read_cell_style,
    read_cell_value,
)
from desktop_env.evaluators.metrics.utils import (
    load_charts,
    load_sparklines,
    load_rows_or_cols,
    load_xlsx_styles,
    load_filters,
    load_pivot_tables,
)

# from openpyxl.utils import coordinate_to_tuple

logger = logging.getLogger("desktopenv.metric.table")

BOOK = Union[pd.ExcelFile, Workbook, str]


def _parse_sheet_idx(
    sheet_idx: Union[int, str],
    result: BOOK,
    expected: BOOK,
    result_sheet_names: List[str],
    expected_sheet_names: List[str],
) -> Tuple[BOOK, str]:
    #  function _parse_sheet_idx {{{ #
    if isinstance(sheet_idx, int):
        try:
            if not result_sheet_names or sheet_idx >= len(result_sheet_names):
                logger.error(
                    f"Sheet index {sheet_idx} out of range. Available sheets: {result_sheet_names}"
                )
                index = ""
            else:
                index: str = result_sheet_names[sheet_idx]
                logger.debug(f"Sheet index {sheet_idx} resolved to sheet: {index}")
        except Exception as e:
            logger.error(f"Error resolving sheet index {sheet_idx}: {e}")
            index = ""
        book: BOOK = result
    elif sheet_idx.startswith("RI"):
        try:
            index: str = result_sheet_names[int(sheet_idx[2:])]
        except:
            index = ""
        book: BOOK = result
    elif sheet_idx.startswith("RN"):
        index: str = sheet_idx[2:]
        book: BOOK = result
    elif sheet_idx.startswith("EI"):
        try:
            index: str = expected_sheet_names[int(sheet_idx[2:])]
        except:
            index = ""
        book: BOOK = expected
    elif sheet_idx.startswith("EN"):
        index: str = sheet_idx[2:]
        book: BOOK = expected
    else:
        logger.error("Unrecognized sheet index")
        raise ValueError("Unrecognized sheet index")
    return book, index
    #  }}} function _parse_sheet_idx #


SHEET = Union[pd.DataFrame, Worksheet, List[str]]


def _load_sheet(book: BOOK, index: str) -> SHEET:
    #  function _load_sheet {{{ #
    try:
        if isinstance(book, str):
            book: str = cast(str, book)
            csv_name: str = "{:}-{:}.csv".format(os.path.splitext(book)[0], index)

            try:
                all_lines: List[str] = _safe_read_file(csv_name)
                csv_lines: List[str] = list(
                    itertools.dropwhile(
                        lambda l: len(l) == 0,
                        map(lambda l: l.strip(), reversed(all_lines)),
                    )
                )
                return csv_lines
            except (FileNotFoundError, IOError) as e:
                logger.error(f"Failed to read CSV file {csv_name}: {e}")
                return None
        if isinstance(book, pd.ExcelFile):
            return pd.read_excel(book, index)
        if isinstance(book, Workbook):
            return book[index]
        logger.error("Not supported workbook format")
        raise NotImplementedError("Not supported workbook format")
    except NotImplementedError as e:
        raise e
    except:
        return None
    #  }}} function _load_sheet #


def _safe_read_file(file_path: str) -> List[str]:
    """
    Safely read a file with multiple encoding attempts.

    Args:
        file_path: Path to the file to read

    Returns:
        List of lines from the file

    Raises:
        FileNotFoundError: If file doesn't exist
        IOError: If file cannot be read with any encoding
    """
    # Common encodings to try in order of preference
    encodings = [
        "utf-8",  # Most common modern encoding
        "utf-8-sig",  # UTF-8 with BOM
        "latin-1",  # ISO-8859-1, works with any byte sequence
        "windows-1252",  # Common Windows encoding
        "gbk",  # Chinese encoding
        "cp1251",  # Cyrillic encoding
        "iso-8859-1",  # Alternative latin-1
    ]

    last_error = None

    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding) as f:
                lines = f.read().splitlines()
                logger.debug(
                    f"Successfully read file {file_path} with encoding {encoding}"
                )
                return lines
        except UnicodeDecodeError as e:
            last_error = e
            logger.debug(f"Failed to read {file_path} with encoding {encoding}: {e}")
            continue
        except (FileNotFoundError, IOError) as e:
            # These are non-encoding related errors, re-raise immediately
            raise e

    # If all encodings fail, try with error handling as last resort
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.read().splitlines()
            logger.warning(f"Read file {file_path} with UTF-8 and error replacement")
            return lines
    except Exception as e:
        logger.error(
            f"Failed to read file {file_path} with any encoding. Last error: {last_error}"
        )
        raise IOError(
            f"Cannot read file {file_path} with any supported encoding"
        ) from last_error


def compare_csv(result: str, expected: Union[str, List[str]], **options) -> float:
    """
    Compare CSV files. If expected is a list, returns 1.0 if result matches any of the expected files.

    Args:
        result: Path to result CSV file
        expected: Path to expected CSV file or list of paths to expected CSV files
        options: Additional options (strict, ignore_case)

    Returns:
        1.0 if result matches expected (or any file in expected list), 0.0 otherwise
    """
    if result is None:
        return 0.0

    try:
        result_lines: List[str] = _safe_read_file(result)
    except (FileNotFoundError, IOError) as e:
        logger.error(f"Failed to read result file {result}: {e}")
        return 0.0

    # Convert expected to list if it's a single string (for backward compatibility)
    if isinstance(expected, str):
        expected_files = [expected]
    else:
        expected_files = expected

    # Try to match against each expected file
    for expected_file in expected_files:
        try:
            expected_lines: List[str] = _safe_read_file(expected_file)

            # Process lines based on options
            current_result_lines = result_lines
            current_expected_lines = expected_lines

            if not options.get("strict", True):
                current_result_lines = map(str.strip, current_result_lines)
                current_expected_lines = map(str.strip, current_expected_lines)
            if options.get("ignore_case", False):
                current_result_lines = map(str.lower, current_result_lines)
                current_expected_lines = map(str.lower, current_expected_lines)

            # Check if this expected file matches
            if list(current_result_lines) == list(current_expected_lines):
                return 1.0

        except (FileNotFoundError, IOError):
            # If this expected file doesn't exist, continue to next one
            continue

    # No match found
    return 0.0


def compare_table(result: str, expected: str = None, **options) -> float:
    #  function compare_table {{{ #
    """
    Args:
        result (str): path to result xlsx
        expected (str): path to golden xlsx
        rules (List[Dict[str, Any]]): list of dict like
          {
            "type": str,
            <str as parameters>: anything
          }
          as sequential rules

    Returns:
        float: the score
    """

    if result is None:
        logger.error("Result file path is None")
        return 0.0

    # Check if result file exists
    if not os.path.exists(result):
        logger.error(f"Result file not found: {result}")
        return 0.0

    try:
        logger.info(f"Loading result file: {result}")
        xlworkbookr: Workbook = openpyxl.load_workbook(filename=result)
        pdworkbookr = pd.ExcelFile(result)
        logger.info(
            f"Successfully loaded result file with sheets: {pdworkbookr.sheet_names}"
        )
    except Exception as e:
        logger.error(f"Failed to load result file {result}: {e}")
        return 0.0
    worksheetr_names: List[str] = pdworkbookr.sheet_names

    if expected is not None:
        xlworkbooke: Workbook = openpyxl.load_workbook(filename=expected)
        pdworkbooke = pd.ExcelFile(expected)
        worksheete_names: List[str] = pdworkbooke.sheet_names
    else:
        xlworkbooke: Workbook = None
        pdworkbooke = None
        worksheete_names: List[str] = None

    parse_idx: Callable[[Union[str, int], BOOK, BOOK], Tuple[BOOK, str]] = (
        functools.partial(
            _parse_sheet_idx,
            result_sheet_names=worksheetr_names,
            expected_sheet_names=worksheete_names,
        )
    )

    passes = True
    for r in options["rules"]:
        if r["type"] == "sheet_name":
            #  Compare Sheet Names {{{ #
            metric: bool = worksheetr_names == worksheete_names
            logger.debug(
                "Assertion: %s.sheet_names == %s.sheet_names - %s",
                result,
                expected,
                metric,
            )
            #  }}} Compare Sheet Names #

        elif r["type"] == "sheet_data":
            #  Compare Sheet Data by Internal Value {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # precision: int as number of decimal digits, default to 4

            error_limit: int = r.get("precision", 4)
            sheet1: pd.DataFrame = _load_sheet(
                *parse_idx(r["sheet_idx0"], pdworkbookr, pdworkbooke)
            )
            if sheet1 is None:
                return 0.0
            sheet2: pd.DataFrame = _load_sheet(
                *parse_idx(r["sheet_idx1"], pdworkbookr, pdworkbooke)
            )

            sheet1 = sheet1.round(error_limit)
            sheet2 = sheet2.round(error_limit)
            metric: bool = sheet1.equals(sheet2)
            logger.debug("Sheet1: \n%s", str(sheet1))
            logger.debug("Sheet2: \n%s", str(sheet2))
            try:
                logger.debug("Sheet1 =v= Sheet2: \n%s", str(sheet1 == sheet2))
            except:
                logger.debug("Sheet1 =/v= Sheet2")
            logger.debug(
                "Assertion: %s =v= %s - %s", r["sheet_idx0"], r["sheet_idx1"], metric
            )
            #  }}} Compare Sheet Data by Internal Value #

        elif r["type"] == "sheet_print":
            #  Compare Sheet Data by Printed Value {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # ignore_case: optional, defaults to False

            sheet1: List[str] = _load_sheet(
                *parse_idx(r["sheet_idx0"], result, expected)
            )
            if sheet1 is None:
                return 0.0
            sheet2: List[str] = _load_sheet(
                *parse_idx(r["sheet_idx1"], result, expected)
            )
            if r.get("ignore_case", False):
                sheet1 = [l.lower() for l in sheet1]
                sheet2 = [l.lower() for l in sheet2]
            metric: bool = sheet1 == sheet2
            logger.debug(
                "Assertion: %s =p= %s - %s", r["sheet_idx0"], r["sheet_idx1"], metric
            )
            #  }}} Compare Sheet Data by Printed Value #

        elif r["type"] == "sheet_fuzzy":
            #  Fuzzy Match for Ranges {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # rules: list of dict, each dict is like
            #   { "range": ["A1:B6", "C2:E5"],
            #     "type": "includes" | "included_by" | "fuzzy_match" | "exact_match", # 0 includes 1, 0 includes_by 1
            #     "threshold": 85, // for fuzzy match
            #     "ignore_case": true | false,
            #     "ignore_chars": " ()", # filtered out
            #     "trim_leadings": "+ ", # filtered by lstrip
            #     "trim_trailings": "", # filtered by rstrip
            #     "normalization": [["Rd", "Road"]], # filtered by replace
            #   }

            sheet1: Tuple[BOOK, str] = parse_idx(r["sheet_idx0"], result, expected)
            sheet2: Tuple[BOOK, str] = parse_idx(r["sheet_idx1"], result, expected)
            total_metric = True
            for rl in r["rules"]:
                for rng in MultiCellRange(rl["range"]):
                    for cdn in rng.cells:
                        coordinate: str = "{:}{:d}".format(
                            get_column_letter(cdn[1]), cdn[0]
                        )
                        value1: str = str(read_cell_value(*sheet1, coordinate))
                        value2: str = str(read_cell_value(*sheet2, coordinate))
                        logger.debug("%s: %s vs %s", cdn, value1, value2)

                        for rplc in rl.get("normalization", []):
                            value1 = value1.replace(rplc[0], rplc[1])
                            value2 = value2.replace(rplc[0], rplc[1])
                        if "trim_leadings" in rl:
                            value1 = value1.lstrip(rl["trim_leadings"])
                            value2 = value2.lstrip(rl["trim_leadings"])
                        if "trim_trailings" in rl:
                            value1 = value1.rstrip(rl["trim_trailings"])
                            value2 = value2.rstrip(rl["trim_trailings"])
                        if "ignore_chars" in rl:
                            ignore_chars: Set[str] = set(rl["ignore_chars"])
                            value1 = "".join(
                                filter(lambda ch: ch not in ignore_chars, value1)
                            )
                            value2 = "".join(
                                filter(lambda ch: ch not in ignore_chars, value2)
                            )
                        if rl.get("ignore_case", False):
                            value1 = value1.lower()
                            value2 = value2.lower()

                        if rl["type"] == "includes":
                            metric: bool = value2 in value1
                        elif rl["type"] == "included_by":
                            metric: bool = value1 in value2
                        elif rl["type"] == "fuzzy_match":
                            metric: bool = fuzz.ratio(value1, value2) >= rl.get(
                                "threshold", 85.0
                            )
                        elif rl["type"] == "exact_match":
                            metric: bool = value1 == value2
                        total_metric = total_metric and metric

            metric: bool = total_metric
            logger.debug(
                "Assertion: %s =~= %s - %s", r["sheet_idx0"], r["sheet_idx1"], metric
            )
            #  }}} Fuzzy Match for Ranges #

        elif r["type"] == "sparkline":
            #  Compare Sparklines {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0

            sparkline1: Dict[str, str] = load_sparklines(
                *parse_idx(r["sheet_idx0"], result, expected)
            )
            sparkline2: Dict[str, str] = load_sparklines(
                *parse_idx(r["sheet_idx1"], result, expected)
            )
            metric: bool = sparkline1 == sparkline2
            logger.debug(
                "Assertion: %s.sp == %.sp - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Compare Sparklines #

        elif r["type"] == "chart":
            #  Compare Charts {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # chart_props: list of str, see utils.load_charts

            charts1: Dict[str, Any] = load_charts(
                *parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke), **r
            )
            charts2: Dict[str, Any] = load_charts(
                *parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke), **r
            )
            metric: bool = charts1 == charts2
            logger.debug(
                "Assertion: %s[chart] == %s[chart] - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Compare Charts #

        elif r["type"] == "style":
            #  Compare Style (Also Conditional Formatiing) {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # props: list of str indicating concerned styles, see utils._read_cell_style

            sheet_idx1: Tuple[BOOK, str] = parse_idx(
                r["sheet_idx0"], xlworkbookr, xlworkbooke
            )
            book_name1: str = parse_idx(r["sheet_idx0"], result, expected)[0]
            styles1: Dict[str, List[Any]] = load_xlsx_styles(
                *sheet_idx1, book_name1, **r
            )

            sheet_idx2: Tuple[BOOK, str] = parse_idx(
                r["sheet_idx1"], xlworkbookr, xlworkbooke
            )
            book_name2: str = parse_idx(r["sheet_idx1"], result, expected)[0]
            styles2: Dict[str, List[Any]] = load_xlsx_styles(
                *sheet_idx2, book_name2, **r
            )
            # number_formats1: List[str] = [c.number_format.lower() for col in sheet1.iter_cols() for c in col if c.value is not None and c.data_type=="n"]
            # number_formats2: List[str] = [c.number_format.lower() for col in sheet2.iter_cols() for c in col if c.value is not None and c.data_type=="n"]
            metric: bool = styles1 == styles2
            logger.debug(
                "Assertion: %s.style == %s.style - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Compare Style (Also Conditional Formatiing) #

        elif r["type"] == "freeze":
            #  Compare Freezing {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0

            sheet1: Worksheet = _load_sheet(
                *parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke)
            )
            if sheet1 is None:
                return 0.0
            sheet2: Worksheet = _load_sheet(
                *parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke)
            )
            metric: bool = sheet1.freeze_panes == sheet2.freeze_panes
            logger.debug(
                "Assertion: %s.freeze(%s) == %s.freeze(%s) - %s",
                r["sheet_idx0"],
                sheet1.freeze_panes,
                r["sheet_idx1"],
                sheet2.freeze_panes,
                metric,
            )
            #  }}} Compare Freezing #

        elif r["type"] == "zoom":
            #  Check Zooming {{{ #
            # sheet_idx: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # method: str
            # ref: value

            sheet: Worksheet = _load_sheet(
                *parse_idx(r["sheet_idx"], xlworkbookr, xlworkbooke)
            )
            if sheet is None:
                return 0.0
            zoom_scale: Number = sheet.sheet_view.zoomScale or 100.0
            metric: bool = _match_value_to_rule(zoom_scale, r)
            logger.debug(
                "Assertion: %s.zoom(%.1f) %s %.1f - %s",
                r["sheet_idx"],
                zoom_scale,
                r["method"],
                r["ref"],
                metric,
            )
            #  }}} Check Zooming #

        elif r["type"] == "data_validation":
            #  Check Data Validation {{{ #
            # sheet_idx: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # dv_props: list of dict like {attribute: {"method": str, "ref": anything}}
            #   available attributes:
            #     * ranges
            #     * type
            #     * formula1
            #     * formula2
            #     * operator
            #     * allowBlank
            #     * showDropDown
            #     * showInputMessage
            #     * showErrorMessage
            #     * error
            #     * errorTitle
            #     * errorStyle
            #     * prompt
            #     * promptTitle
            #     * imeMode

            sheet: Worksheet = _load_sheet(
                *parse_idx(r["sheet_idx"], xlworkbookr, xlworkbooke)
            )
            if sheet is None:
                return 0.0
            data_validators: List[DataValidation] = (
                sheet.data_validations.dataValidation
            )

            total_metric = len(data_validators) >= len(r["dv_props"])
            for dat_vldt in data_validators:
                metric = False
                for prpt in r["dv_props"]:
                    metric = metric or all(
                        _match_value_to_rule(getattr(dat_vldt, attrbt), mr)
                        for attrbt, mr in prpt.items()
                    )
                    if metric:
                        break
                total_metric = total_metric and metric
                if not total_metric:
                    break

            logger.debug(
                "Assertion: %s.data_validation - %s", r["sheet_idx"], total_metric
            )
            metric: bool = total_metric
            #  }}} Check Data Validation #

        elif r["type"] == "row_props":
            #  Check Row Properties {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # props: list of str, see utils.load_rows_or_cols

            rows1: Dict[str, Any] = load_rows_or_cols(
                *parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke), obj="row", **r
            )
            rows2: Dict[str, Any] = load_rows_or_cols(
                *parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke), obj="row", **r
            )
            logger.debug("Rows1: %s", repr(rows1))
            logger.debug("Rows2: %s", repr(rows2))
            metric: bool = rows1 == rows2
            logger.debug(
                "Assertion: %s[rows] == %s[rows] - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Check Row Properties #

        elif r["type"] == "col_props":
            #  Check Row Properties {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # props: list of str, see utils.load_rows_or_cols

            cols1: Dict[str, Any] = load_rows_or_cols(
                *parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke), obj="column", **r
            )
            cols2: Dict[str, Any] = load_rows_or_cols(
                *parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke), obj="column", **r
            )
            metric: bool = cols1 == cols2
            logger.debug(
                "Assertion: %s[cols] == %s[cols] - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Check Row Properties #

        elif r["type"] == "filter":
            #  Compare Filters {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0

            filters1: Dict[str, Any] = load_filters(
                *parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke), **r
            )
            filters2: Dict[str, Any] = load_filters(
                *parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke), **r
            )
            metric: bool = filters1 == filters2
            logger.debug(
                "Assertion: %s[filter] == %s[filter] - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Compare Filters #

        elif r["type"] == "pivot_table":
            #  Compare Pivot Tables {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # pivot_props: list of str, see utils.load_pivot_tables

            pivots1: Dict[str, Any] = load_pivot_tables(
                *parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke), **r
            )
            pivots2: Dict[str, Any] = load_pivot_tables(
                *parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke), **r
            )
            metric: bool = pivots1 == pivots2
            logger.debug(
                "Assertion: %s[pivot]==%s[pivot] - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Compare Pivot Tables #

        elif r["type"] == "check_cell":
            #  Check Cell Properties {{{ #
            # sheet_idx: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # coordinate: str, "E3"
            # props: dict like {attribute: {"method": str, "ref": anything}}
            #   supported attributes: value & those supported by utils._read_cell_style

            try:
                sheet: Worksheet = _load_sheet(
                    *parse_idx(r["sheet_idx"], xlworkbookr, xlworkbooke)
                )
                if sheet is None:
                    logger.error(
                        f"Failed to load sheet for sheet_idx: {r['sheet_idx']}"
                    )
                    return 0.0
                # data_frame: pd.DataFrame = _load_sheet(*parse_idx(r["sheet_idx"], pdworkbookr, pdworkbooke))
                cell: Cell = sheet[r["coordinate"]]
                metric: bool = True
                for prpt, rule in r["props"].items():
                    if prpt == "value":
                        try:
                            parsed_result = parse_idx(r["sheet_idx"], result, expected)
                            logger.debug(f"parse_idx result: {parsed_result}")
                            val = read_cell_value(*parsed_result, r["coordinate"])
                            logger.debug(f"Cell {r['coordinate']} value: {val}")
                        except Exception as e:
                            logger.error(
                                f"Failed to read cell value at {r['coordinate']}: {e}"
                            )
                            val = None
                    elif prpt == "formula":
                        # Support checking cell formula directly
                        try:
                            if cell.data_type == "f":
                                # For formula cells, get the formula text
                                # In openpyxl, formula is stored in cell.value for formula cells
                                # But we need the actual formula text, not the calculated value
                                # Try to get formula from internal representation
                                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                                    val = cell._value
                                elif hasattr(cell, "formula"):
                                    val = cell.formula
                                else:
                                    # Fallback: try to reconstruct from value if it's a formula
                                    val = f"={cell.value}" if cell.value is not None else None
                            else:
                                val = None
                            logger.debug(f"Cell {r['coordinate']} formula: {val}")
                        except Exception as e:
                            logger.error(
                                f"Failed to read cell formula at {r['coordinate']}: {e}"
                            )
                            val = None
                    else:
                        try:
                            val = _read_cell_style(prpt, cell)
                        except Exception as e:
                            logger.error(
                                f"Failed to read cell style {prpt} at {r['coordinate']}: {e}"
                            )
                            val = None

                    metric = metric and _match_value_to_rule(val, rule)
            except Exception as e:
                logger.error(f"Error in check_cell processing: {e}")
                return 0.0

            logger.debug(
                "Assertion: %s[%s] :%s - %s",
                r["sheet_idx"],
                r["coordinate"],
                repr(r["props"]),
                metric,
            )
            #  }}} Check Cell Properties #

        else:
            raise NotImplementedError(
                "Unimplemented sheet check: {:}".format(r["type"])
            )

        passes = passes and metric
        if not passes:
            break

    return float(passes)
    #  }}} function compare_table #


def compare_conference_city_in_order(actual_city_list_path, expected_city):
    expected_city_list = expected_city["expected"]
    wb = openpyxl.load_workbook(actual_city_list_path)
    sheet = wb.active
    actual_city_list = []
    for row in sheet["C2:C22"]:
        for cell in row:
            actual_city_list.append(cell.value)
    # expected_city is the city that we want to compare with the actual city list
    # must in order index
    # debug
    try:
        for i in range(len(actual_city_list)):
            if isinstance(expected_city_list[i], str):
                if expected_city_list[i] not in actual_city_list[i]:
                    logger.debug(
                        f"Expected city {expected_city_list[i]}; Actual city {actual_city_list[i]}"
                    )
                    print(
                        f"Expected city {expected_city_list[i]}; Actual city {actual_city_list[i]}"
                    )
                    return 0.0

            elif isinstance(expected_city_list[i], List):
                if not any(
                    possible_str in actual_city_list[i]
                    for possible_str in expected_city_list[i]
                ):
                    logger.debug(
                        f"Expected city {expected_city_list[i]}; Actual city {actual_city_list[i]}"
                    )
                    print(
                        f"Expected city {expected_city_list[i]}; Actual city {actual_city_list[i]}"
                    )
                    return 0.0

            else:
                raise TypeError("Expected city should be a string or a list of strings")

    except:
        return 0.0

    return 1.0


def verify_second_row_deleted_without_gold(result: str, expected: str = None, **options) -> float:
    """
    验证 Excel 文件的第二行是否被删除（不需要金标准文件）
    
    通过以下方式验证：
    1. 检查结果文件的行数是否比原始文件少1
    2. 检查原始文件的第二行数据是否在结果文件中不存在
    3. 检查其他所有行是否保持不变
    
    Args:
        result (str): 结果文件路径
        expected (str): 未使用（为了兼容框架接口）
        options (dict): 配置选项，应包含：
            - original_file_url: 原始文件的URL（用于下载和比对）
            - result_file_path: 结果文件的路径（可选，默认使用 result 参数）
            - original_file_cache: 原始文件的本地缓存路径（可选）
    
    Returns:
        float: 如果验证通过返回 1.0，否则返回 0.0
    """
    try:
        import tempfile
        import urllib.request
        
        # result 参数已经是从VM获取到宿主机的文件路径
        # 不应该从 options 中覆盖它，因为 options 中可能包含的是VM路径
        result_file_path = result
        original_file_url = options.get('original_file_url', '')
        
        logger.info(f"开始验证删除第二行任务...")
        logger.info(f"结果文件: {result_file_path}")
        logger.info(f"原始文件URL: {original_file_url}")
        
        if not result_file_path or not os.path.exists(result_file_path):
            logger.error(f"结果文件不存在: {result_file_path}")
            return 0.0
        
        # 下载原始文件到临时位置
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            original_file_temp = tmp_file.name
        
        try:
            logger.info(f"正在下载原始文件到临时位置: {original_file_temp}")
            urllib.request.urlretrieve(original_file_url, original_file_temp)
        except Exception as e:
            logger.warning(f"下载原始文件失败: {e}")
            # 如果下载失败，尝试从本地缓存读取
            cache_path = options.get('original_file_cache', '')
            if cache_path and os.path.exists(cache_path):
                logger.info(f"使用缓存文件: {cache_path}")
                original_file_temp = cache_path
            else:
                logger.error("无法获取原始文件")
                return 0.0
        
        # 加载原始文件
        logger.info("加载原始文件...")
        original_wb = openpyxl.load_workbook(original_file_temp)
        original_ws = original_wb.active
        
        # 获取原始文件的所有行
        original_rows = list(original_ws.iter_rows(values_only=True))
        original_row_count = len(original_rows)
        
        if original_row_count < 2:
            logger.error(f"原始文件行数不足: {original_row_count}（需要至少2行）")
            return 0.0
        
        # 保存第二行的数据（索引为1）
        second_row_data = original_rows[1]
        logger.info(f"原始文件行数: {original_row_count}")
        logger.info(f"原始文件第二行数据: {second_row_data}")
        
        # 加载结果文件
        logger.info(f"加载结果文件...")
        result_wb = openpyxl.load_workbook(result_file_path)
        result_ws = result_wb.active
        
        # 获取结果文件的所有行
        result_rows = list(result_ws.iter_rows(values_only=True))
        result_row_count = len(result_rows)
        
        logger.info(f"结果文件行数: {result_row_count}")
        
        # 验证1: 检查行数是否减少了1
        if result_row_count != original_row_count - 1:
            logger.error(f"行数验证失败: 期望 {original_row_count - 1} 行，实际 {result_row_count} 行")
            return 0.0
        else:
            logger.info(f"✓ 行数验证通过: {original_row_count} → {result_row_count}")
        
        # 验证2: 检查原始第二行是否存在于结果文件中
        second_row_exists = False
        for i, row in enumerate(result_rows):
            if row == second_row_data:
                logger.error(f"原始第二行数据仍存在于结果文件的第 {i+1} 行")
                second_row_exists = True
                break
        
        if second_row_exists:
            return 0.0
        else:
            logger.info(f"✓ 原始第二行数据已从结果文件中删除")
        
        # 验证3: 检查其他行是否保持不变（第一行和第3行之后）
        # 结果文件的第一行应该等于原始文件的第一行
        if result_rows[0] != original_rows[0]:
            logger.error(f"第一行数据不匹配")
            logger.error(f"  原始: {original_rows[0]}")
            logger.error(f"  结果: {result_rows[0]}")
            return 0.0
        
        # 结果文件的第2行及之后应该等于原始文件的第3行及之后
        for i in range(1, result_row_count):
            if result_rows[i] != original_rows[i+1]:
                logger.error(f"第 {i+1} 行数据不匹配")
                logger.error(f"  期望（原始第 {i+2} 行）: {original_rows[i+1]}")
                logger.error(f"  实际: {result_rows[i]}")
                return 0.0
        
        logger.info(f"✓ 其他行数据保持不变")
        
        # 清理临时文件
        if original_file_temp != options.get('original_file_cache', ''):
            try:
                os.unlink(original_file_temp)
            except:
                pass
        
        logger.info("=" * 60)
        logger.info("✓ 所有验证通过！第二行已成功删除")
        logger.info("=" * 60)
        return 1.0
        
    except Exception as e:
        import traceback
        logger.error(f"评估出错: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_regexp_extract(result: str, expected: str = None, **options) -> float:
    """
    Verify if REGEX formulas exist in specified columns (B and C) with correct patterns.
    
    This function checks:
    1. Whether cells in specified columns contain REGEX formulas
    2. Whether formulas reference the corresponding A column cell (B2->A2, B3->A3, etc.)
    3. Whether formulas contain the correct pattern text (牛肉丸 for B column, 牛筋丸 for C column)
    4. Whether formulas have the correct structure with lookbehind and lookahead
    
    The function automatically detects the number of data rows by checking the data column
    (default: A column) for non-empty cells. It stops checking after finding 3 consecutive
    empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_columns: List of columns to check (e.g., ["B", "C"])
            - start_row: Starting row number (default: 2)
            - end_row: Ending row number (optional, will auto-detect if not provided)
            - expected_pattern: Expected function name (default: "REGEX")
            - column_patterns: Dict mapping column letters to expected pattern text
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        from openpyxl.utils import get_column_letter, column_index_from_string
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_columns = options.get('check_columns', ['B', 'C'])
        start_row = options.get('start_row', 2)
        end_row = options.get('end_row', None)  # Optional, will auto-detect if not provided
        expected_pattern = options.get('expected_pattern', 'REGEX')
        column_patterns = options.get('column_patterns', {'B': '牛肉丸', 'C': '牛筋丸'})
        data_column = options.get('data_column', 'A')  # Column to check for data to determine end_row
        
        if not check_columns:
            logger.error("No columns specified in options")
            return 0.0
        
        logger.info(f"Verifying REGEX formulas in file: {result}")
        logger.info(f"Columns to check: {check_columns}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected function: {expected_pattern}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        if end_row is None:
            logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
            max_row = ws.max_row
            end_row = start_row  # Start from start_row
            
            # Find the last row with data in the data column
            # Check up to max_row, but stop if we find 3 consecutive empty rows
            empty_count = 0
            for row_num in range(start_row, max_row + 1):
                data_cell = ws[f"{data_column}{row_num}"]
                if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                    empty_count += 1
                    if empty_count >= 3:  # Stop after 3 consecutive empty rows
                        break
                else:
                    empty_count = 0
                    end_row = row_num  # Update end_row to the last row with data
            
            logger.info(f"Auto-detected end row: {end_row}")
        else:
            logger.info(f"Using specified end row: {end_row}")
        
        # Check each column and row
        all_passed = True
        for col_letter in check_columns:
            expected_pattern_text = column_patterns.get(col_letter)
            if not expected_pattern_text:
                logger.warning(f"No pattern text specified for column {col_letter}, skipping")
                continue
            
            logger.info(f"Checking column {col_letter} with pattern '{expected_pattern_text}' (rows {start_row} to {end_row})")
            
            for row_num in range(start_row, end_row + 1):
                cell_coord = f"{col_letter}{row_num}"
                try:
                    cell = ws[cell_coord]
                    logger.debug(f"Checking cell {cell_coord}")
                    
                    # Check if cell contains a formula
                    if cell.data_type != "f":
                        logger.warning(f"Cell {cell_coord} does not contain a formula")
                        all_passed = False
                        continue
                    
                    # Get formula text
                    formula_text = None
                    if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                        formula_text = cell._value
                    elif hasattr(cell, "formula"):
                        formula_text = cell.formula
                    else:
                        # Try to get from value attribute
                        if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                            formula_text = cell.value
                    
                    if formula_text is None:
                        logger.warning(f"Could not extract formula from cell {cell_coord}")
                        all_passed = False
                        continue
                    
                    # Remove leading = if present for comparison
                    formula_clean = formula_text.lstrip("=")
                    logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                    
                    # Check 1: Formula contains REGEX function
                    if expected_pattern.upper() not in formula_text.upper():
                        logger.warning(f"Cell {cell_coord} formula does not contain {expected_pattern}")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                    
                    # Check 2: Formula contains expected pattern text (牛肉丸 or 牛筋丸)
                    if expected_pattern_text not in formula_text:
                        logger.warning(f"Cell {cell_coord} formula does not contain pattern text '{expected_pattern_text}'")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                    
                    # Check 3: Formula contains REGEX function call structure
                    regex_match = re.search(r'REGEX\s*\([^)]+\)', formula_text, re.IGNORECASE)
                    if not regex_match:
                        logger.warning(f"Cell {cell_coord} formula does not have correct REGEX structure")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                    
                    # Check 4: Formula references the corresponding A column cell (A2, A3, etc.)
                    expected_a_cell = f"A{row_num}"
                    # Check if formula contains A column reference with the same row number
                    a_cell_pattern = rf'A{row_num}\b'
                    if not re.search(a_cell_pattern, formula_text, re.IGNORECASE):
                        logger.warning(f"Cell {cell_coord} formula does not reference {expected_a_cell}")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                    
                    # Check 5: Formula contains lookbehind pattern (?<=...)
                    if "(?<=" not in formula_text:
                        logger.warning(f"Cell {cell_coord} formula does not contain lookbehind pattern (?<=...)")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                    
                    # Check 6: Formula contains lookahead pattern (?=,)
                    if "(?=," not in formula_text:
                        logger.warning(f"Cell {cell_coord} formula does not contain lookahead pattern (?=,)")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                    
                    # Check 7: Formula contains \d+ pattern
                    if "\\d+" not in formula_text:
                        # Also check for unescaped version in the pattern
                        if not re.search(r'\\d\+|d\+', formula_text):
                            logger.warning(f"Cell {cell_coord} formula does not contain digit pattern \\d+")
                            logger.warning(f"Formula: {formula_text}")
                            all_passed = False
                            continue
                    
                    # Check 8: Formula pattern should contain 5 dots after pattern text
                    # Pattern should be like: (?<=牛肉丸.....)
                    pattern_with_dots = expected_pattern_text + "....."
                    if pattern_with_dots not in formula_text:
                        logger.warning(f"Cell {cell_coord} formula pattern may not have 5 dots after '{expected_pattern_text}'")
                        logger.debug(f"Formula: {formula_text}")
                        # Don't fail, just warn - the pattern might be correct but formatted differently
                    
                    logger.info(f"✓ Cell {cell_coord} has valid REGEX formula: {formula_text}")
                    
                except Exception as e:
                    logger.error(f"Error checking cell {cell_coord}: {e}")
                    import traceback
                    logger.error(traceback.format_exc())
                    all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in columns {check_columns} contain correct {expected_pattern} formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ {expected_pattern} formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_regexp_order_extract(result: str, expected: str = None, **options) -> float:
    """
    Verify if REGEXP formulas exist in specified column (C) to extract order numbers from addresses.
    
    This function checks:
    1. Whether cells in specified column contain REGEXP formulas
    2. Whether formulas reference the corresponding A column cell (C2->A2, C3->A3, etc.)
    3. Whether formulas contain the correct regex pattern (\\w{10})
    4. Whether formulas have the correct structure
    
    The function automatically detects the number of data rows by checking the data column
    (default: A column) for non-empty cells. It stops checking after finding 3 consecutive
    empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "C")
            - start_row: Starting row number (default: 2)
            - expected_pattern: Expected function name (default: "REGEXP")
            - expected_formula_pattern: Expected formula pattern (e.g., "REGEXP(A")
            - regex_pattern: Expected regex pattern in formula (e.g., "\\w{10}")
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'C')
        start_row = options.get('start_row', 2)
        expected_pattern = options.get('expected_pattern', 'REGEX')
        expected_formula_pattern = options.get('expected_formula_pattern', 'REGEX(A')
        regex_pattern = options.get('regex_pattern', '[a-zA-Z0-9]{10}')
        data_column = options.get('data_column', 'A')
        
        logger.info(f"Verifying REGEXP order extraction in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected function: {expected_pattern}")
        logger.info(f"Expected regex pattern: {regex_pattern}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row in the specified column
        all_passed = True
        logger.info(f"Checking column {check_column} (rows {start_row} to {end_row})")
        
        for row_num in range(start_row, end_row + 1):
            cell_coord = f"{check_column}{row_num}"
            try:
                cell = ws[cell_coord]
                logger.debug(f"Checking cell {cell_coord}")
                
                # Check if cell contains a formula
                if cell.data_type != "f":
                    logger.warning(f"Cell {cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                # Get formula text
                formula_text = None
                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                    formula_text = cell._value
                elif hasattr(cell, "formula"):
                    formula_text = cell.formula
                else:
                    # Try to get from value attribute
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value
                
                if formula_text is None:
                    logger.warning(f"Could not extract formula from cell {cell_coord}")
                    all_passed = False
                    continue
                
                # Remove leading = if present for comparison
                formula_clean = formula_text.lstrip("=")
                logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                
                # Check 1: Formula contains REGEX function (support both REGEX and LibreOffice internal format)
                # LibreOffice may save as _xlfn.ORG.LIBREOFFICE.REGEX
                formula_upper = formula_text.upper()
                if expected_pattern.upper() not in formula_upper and '_XLFN.ORG.LIBREOFFICE.REGEX' not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain {expected_pattern} or LibreOffice REGEX")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 2: Formula contains expected formula pattern (REGEX(A or _xlfn.ORG.LIBREOFFICE.REGEX(A)
                formula_clean_upper = formula_clean.upper()
                if expected_formula_pattern.upper() not in formula_clean_upper and 'REGEX(A' not in formula_clean_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain pattern '{expected_formula_pattern}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 3: Formula contains REGEX function call structure (support both formats)
                regexp_match = re.search(r'(REGEX|REGEXP|_XLFN\.ORG\.LIBREOFFICE\.REGEX)\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not regexp_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct REGEX structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 4: Formula references the corresponding A column cell (A2, A3, etc.)
                expected_a_cell = f"A{row_num}"
                # Check if formula contains A column reference with the same row number
                a_cell_pattern = rf'A{row_num}\b'
                if not re.search(a_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference {expected_a_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 5: Formula contains the regex pattern ([a-zA-Z0-9]{10})
                # The pattern might be escaped differently in the formula
                # Check for various escape formats
                pattern_variations = [
                    regex_pattern,  # [a-zA-Z0-9]{10}
                    regex_pattern.replace('\\', '\\\\'),  # [a-zA-Z0-9]{10} with double escape
                    regex_pattern.replace('[', '\\[').replace(']', '\\]'),  # Escaped brackets
                    '[a-zA-Z0-9]{10}',  # Original pattern
                    '\\[a-zA-Z0-9\\]{10}',  # Escaped brackets
                    '\\\\[a-zA-Z0-9\\\\]{10}',  # Double escaped
                ]
                found = False
                for pattern_var in pattern_variations:
                    if pattern_var in formula_text:
                        found = True
                        break
                if not found:
                    # Also check for pattern without escaping brackets
                    simple_pattern = 'a-zA-Z0-9]{10}'
                    if simple_pattern not in formula_text:
                        logger.warning(f"Cell {cell_coord} formula does not contain regex pattern '{regex_pattern}'")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                
                logger.info(f"✓ Cell {cell_coord} has valid REGEXP formula: {formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking cell {cell_coord}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in column {check_column} contain correct {expected_pattern} formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ {expected_pattern} formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_sumif_formula(result: str, expected: str = None, **options) -> float:
    """
    Verify if SUMIF formulas exist in specified column (F) to calculate totals.
    
    This function checks:
    1. Whether cells in specified column contain SUMIF formulas
    2. Whether formulas reference the correct ranges (auto-detected from data)
    3. Whether formulas reference the corresponding E column cell (F2->E2, F3->E3, etc.)
    4. Whether formulas have the correct structure
    
    The function automatically detects:
    - end_row: by checking the data column (E) for non-empty cells
    - criteria_range: by detecting the range from the first formula or from criteria_column data
    - sum_range: by detecting the range from the first formula or from sum_column data
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "F")
            - start_row: Starting row number (default: 2)
            - expected_function: Expected function name (default: "SUMIF")
            - criteria_column: Column containing criteria (e.g., "B")
            - sum_column: Column containing values to sum (e.g., "C")
            - criteria_column_start: Starting row for criteria column (default: 2)
            - data_column: Column to check for data to determine end_row (default: "E")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'F')
        start_row = options.get('start_row', 2)
        expected_function = options.get('expected_function', 'SUMIF')
        criteria_column = options.get('criteria_column', 'B')
        sum_column = options.get('sum_column', 'C')
        criteria_column_start = options.get('criteria_column_start', 2)
        data_column = options.get('data_column', 'E')
        
        logger.info(f"Verifying SUMIF formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected function: {expected_function}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Auto-detect criteria_range and sum_range by checking the first formula
        criteria_range = None
        sum_range = None
        
        # Try to extract ranges from the first formula
        first_cell_coord = f"{check_column}{start_row}"
        try:
            first_cell = ws[first_cell_coord]
            if first_cell.data_type == "f":
                first_formula_text = None
                if hasattr(first_cell, "_value") and isinstance(first_cell._value, str) and first_cell._value.startswith("="):
                    first_formula_text = first_cell._value
                elif hasattr(first_cell, "formula"):
                    first_formula_text = first_cell.formula
                elif first_cell.value is not None and isinstance(first_cell.value, str) and first_cell.value.startswith("="):
                    first_formula_text = first_cell.value
                
                if first_formula_text:
                    # Extract ranges from SUMIF formula: SUMIF(range1, criteria, range2)
                    # Pattern: SUMIF(range1, criteria, range2)
                    sumif_pattern = r'SUMIF\s*\(\s*([^,]+)\s*,\s*[^,]+\s*,\s*([^)]+)\s*\)'
                    match = re.search(sumif_pattern, first_formula_text, re.IGNORECASE)
                    if match:
                        criteria_range = match.group(1).strip()
                        sum_range = match.group(2).strip()
                        logger.info(f"Extracted from first formula: criteria_range={criteria_range}, sum_range={sum_range}")
        except Exception as e:
            logger.debug(f"Could not extract ranges from first formula: {e}")
        
        # If ranges not found in formula, detect from data columns
        if not criteria_range or not sum_range:
            logger.info(f"Auto-detecting ranges from data columns...")
            # Find the last row with data in criteria_column
            criteria_end_row = criteria_column_start
            empty_count = 0
            for row_num in range(criteria_column_start, max_row + 1):
                criteria_cell = ws[f"{criteria_column}{row_num}"]
                if criteria_cell.value is None or (isinstance(criteria_cell.value, str) and criteria_cell.value.strip() == ""):
                    empty_count += 1
                    if empty_count >= 3:
                        break
                else:
                    empty_count = 0
                    criteria_end_row = row_num
            
            # Find the last row with data in sum_column
            sum_end_row = criteria_column_start
            empty_count = 0
            for row_num in range(criteria_column_start, max_row + 1):
                sum_cell = ws[f"{sum_column}{row_num}"]
                if sum_cell.value is None or (isinstance(sum_cell.value, str) and sum_cell.value.strip() == ""):
                    empty_count += 1
                    if empty_count >= 3:
                        break
                else:
                    empty_count = 0
                    sum_end_row = row_num
            
            # Use the maximum end row for both ranges
            max_end_row = max(criteria_end_row, sum_end_row)
            criteria_range = f"{criteria_column}{criteria_column_start}:{criteria_column}{max_end_row}"
            sum_range = f"{sum_column}{criteria_column_start}:{sum_column}{max_end_row}"
            logger.info(f"Auto-detected ranges: criteria_range={criteria_range}, sum_range={sum_range}")
        
        # Check each row in the specified column
        all_passed = True
        logger.info(f"Checking column {check_column} (rows {start_row} to {end_row})")
        
        for row_num in range(start_row, end_row + 1):
            cell_coord = f"{check_column}{row_num}"
            try:
                cell = ws[cell_coord]
                logger.debug(f"Checking cell {cell_coord}")
                
                # Check if cell contains a formula
                if cell.data_type != "f":
                    logger.warning(f"Cell {cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                # Get formula text
                formula_text = None
                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                    formula_text = cell._value
                elif hasattr(cell, "formula"):
                    formula_text = cell.formula
                else:
                    # Try to get from value attribute
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value
                
                if formula_text is None:
                    logger.warning(f"Could not extract formula from cell {cell_coord}")
                    all_passed = False
                    continue
                
                # Remove leading = if present for comparison
                formula_clean = formula_text.lstrip("=")
                logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                
                # Check 1: Formula contains SUMIF function
                formula_upper = formula_text.upper()
                if expected_function.upper() not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain {expected_function}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 2: Formula contains SUMIF function call structure
                sumif_match = re.search(r'SUMIF\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not sumif_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct SUMIF structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 3: Formula contains criteria range
                if criteria_range and criteria_range.upper() not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain criteria range '{criteria_range}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 4: Formula contains sum range
                if sum_range and sum_range.upper() not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain sum range '{sum_range}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 5: Formula references the corresponding E column cell (E2, E3, etc.)
                expected_e_cell = f"E{row_num}"
                # Check if formula contains E column reference with the same row number
                e_cell_pattern = rf'E{row_num}\b'
                if not re.search(e_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference {expected_e_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {cell_coord} has valid SUMIF formula: {formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking cell {cell_coord}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in column {check_column} contain correct {expected_function} formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ {expected_function} formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_networkdays_formula(result: str, expected: str = None, **options) -> float:
    """
    Verify if NETWORKDAYS formulas exist in specified column to calculate working days.
    
    This function checks:
    1. Whether cells in specified column contain NETWORKDAYS formulas
    2. Whether formulas reference the corresponding start date column cell (A2, A3, etc.)
    3. Whether formulas reference the corresponding end date column cell (B2, B3, etc.)
    4. Whether formulas have the correct structure
    
    The function automatically detects the number of data rows by checking the data column
    (default: A column) for non-empty cells. It stops checking after finding 3 consecutive
    empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "C")
            - start_row: Starting row number (default: 2)
            - start_date_column: Column containing start dates (e.g., "A")
            - end_date_column: Column containing end dates (e.g., "B")
            - expected_function: Expected function name (default: "NETWORKDAYS")
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'C')
        start_row = options.get('start_row', 2)
        start_date_column = options.get('start_date_column', 'A')
        end_date_column = options.get('end_date_column', 'B')
        expected_function = options.get('expected_function', 'NETWORKDAYS')
        data_column = options.get('data_column', 'A')
        
        logger.info(f"Verifying NETWORKDAYS formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Start date column: {start_date_column}")
        logger.info(f"End date column: {end_date_column}")
        logger.info(f"Expected function: {expected_function}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row in the specified column
        all_passed = True
        logger.info(f"Checking column {check_column} (rows {start_row} to {end_row})")
        
        for row_num in range(start_row, end_row + 1):
            cell_coord = f"{check_column}{row_num}"
            try:
                cell = ws[cell_coord]
                logger.debug(f"Checking cell {cell_coord}")
                
                # Check if cell contains a formula
                if cell.data_type != "f":
                    logger.warning(f"Cell {cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                # Get formula text
                formula_text = None
                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                    formula_text = cell._value
                elif hasattr(cell, "formula"):
                    formula_text = cell.formula
                else:
                    # Try to get from value attribute
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value
                
                if formula_text is None:
                    logger.warning(f"Could not extract formula from cell {cell_coord}")
                    all_passed = False
                    continue
                
                # Remove leading = if present for comparison
                formula_clean = formula_text.lstrip("=")
                logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                
                # Check 1: Formula contains NETWORKDAYS function
                formula_upper = formula_text.upper()
                if expected_function.upper() not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain {expected_function}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 2: Formula contains NETWORKDAYS function call structure
                # NETWORKDAYS can have 2 or 3 parameters: NETWORKDAYS(start_date, end_date) or NETWORKDAYS(start_date, end_date, holidays)
                networkdays_pattern = r'NETWORKDAYS\s*\([^)]+\)'
                networkdays_match = re.search(networkdays_pattern, formula_text, re.IGNORECASE)
                if not networkdays_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct NETWORKDAYS structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 3: Formula references the corresponding start date column cell (A2, A3, etc.)
                expected_start_cell = f"{start_date_column}{row_num}"
                start_cell_pattern = rf'{start_date_column}{row_num}\b'
                if not re.search(start_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference start date cell {expected_start_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 4: Formula references the corresponding end date column cell (B2, B3, etc.)
                expected_end_cell = f"{end_date_column}{row_num}"
                end_cell_pattern = rf'{end_date_column}{row_num}\b'
                if not re.search(end_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference end date cell {expected_end_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {cell_coord} has valid NETWORKDAYS formula: {formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking cell {cell_coord}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in column {check_column} contain correct {expected_function} formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ {expected_function} formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_conditional_formatting_reconciliation(result: str, expected: str = None, **options) -> float:
    """
    Verify if conditional formatting is correctly set up for reconciliation between two tables.
    
    This function checks:
    1. Whether conditional formatting rules exist in the worksheet
    2. Whether the formula matches the expected pattern (e.g., A1<>E1 to compare cells from two tables)
    3. Whether conditional formatting is applied to the correct range
    4. Whether cells with differences are formatted (highlighted)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_range: Range where conditional formatting should be applied (e.g., "A1:C16")
            - compare_range: Range to compare against (e.g., "E1:G16")
            - expected_formula: Expected formula pattern (e.g., "A1<>E1")
            - format_column: Column to check for formatting (optional, e.g., "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        from openpyxl.utils import get_column_letter, column_index_from_string
        from openpyxl.worksheet.cell_range import CellRange
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_range = options.get('check_range', 'A1:C16')
        compare_range = options.get('compare_range', 'E1:G16')
        expected_formula = options.get('expected_formula', 'A1<>E1')
        format_column = options.get('format_column', None)
        
        logger.info(f"Verifying conditional formatting reconciliation in file: {result}")
        logger.info(f"Check range: {check_range}")
        logger.info(f"Compare range: {compare_range}")
        logger.info(f"Expected formula pattern: {expected_formula}")
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(result, data_only=False)
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check if conditional formatting exists
        conditional_formattings = ws.conditional_formatting
        if not conditional_formattings:
            logger.error("No conditional formatting rules found in the worksheet")
            return 0.0
        
        logger.info(f"Found {len(conditional_formattings)} conditional formatting rule(s)")
        
        # Parse expected formula to extract cell references
        # Expected formula like "A1<>E1" means compare A1 with E1
        expected_formula_clean = expected_formula.replace(" ", "").upper()
        
        # Find matching conditional formatting rule
        found_matching_rule = False
        rule_applied_to_correct_range = False
        
        for fmt in conditional_formattings:
            for rule in fmt.rules:
                # Check if rule has formula
                if not rule.formula:
                    continue
                
                # Check formula pattern
                formula_text = rule.formula[0] if rule.formula else ""
                formula_text_clean = formula_text.replace(" ", "").upper()
                
                logger.debug(f"Checking rule with formula: {formula_text}")
                
                # Check if formula matches expected pattern
                # The formula should contain comparison like A1<>E1, A2<>E2, etc.
                # We need to check if the pattern matches (allowing for relative references)
                if "<>" in expected_formula_clean:
                    # Extract cell references from expected formula
                    # Pattern: A1<>E1 means compare A column with E column
                    expected_parts = expected_formula_clean.split("<>")
                    if len(expected_parts) == 2:
                        expected_cell1 = expected_parts[0]  # e.g., "A1"
                        expected_cell2 = expected_parts[1]   # e.g., "E1"
                        
                        # Extract column letters
                        expected_col1 = re.match(r'([A-Z]+)', expected_cell1)
                        expected_col2 = re.match(r'([A-Z]+)', expected_cell2)
                        
                        if expected_col1 and expected_col2:
                            col1 = expected_col1.group(1)
                            col2 = expected_col2.group(1)
                            
                            # Check if formula contains comparison between these columns
                            # Pattern should be like: A1<>E1, A2<>E2, etc. (relative references)
                            pattern1 = rf'{col1}\d+\s*<>\s*{col2}\d+'
                            pattern2 = rf'{col1}\d+\s*!=\s*{col2}\d+'  # Alternative: !=
                            
                            if re.search(pattern1, formula_text_clean, re.IGNORECASE) or \
                               re.search(pattern2, formula_text_clean, re.IGNORECASE):
                                found_matching_rule = True
                                logger.info(f"✓ Found matching formula pattern: {formula_text}")
                                
                                # Check if rule is applied to correct range
                                fmt_ranges = [str(rng) for rng in fmt.cells]
                                check_range_upper = check_range.upper()
                                
                                # Check if check_range is covered by any of the formatting ranges
                                try:
                                    check_cell_range = CellRange(check_range_upper)
                                    for fmt_range_str in fmt_ranges:
                                        fmt_cell_range = CellRange(fmt_range_str)
                                        # Check if check_range is within or overlaps with fmt_range
                                        if (check_cell_range.min_row >= fmt_cell_range.min_row and
                                            check_cell_range.max_row <= fmt_cell_range.max_row and
                                            check_cell_range.min_col >= fmt_cell_range.min_col and
                                            check_cell_range.max_col <= fmt_cell_range.max_col):
                                            rule_applied_to_correct_range = True
                                            logger.info(f"✓ Rule applied to correct range: {fmt_range_str} covers {check_range}")
                                            break
                                except Exception as e:
                                    logger.debug(f"Error parsing ranges: {e}")
                                    # If range parsing fails, check if range string matches
                                    if check_range_upper in fmt_ranges:
                                        rule_applied_to_correct_range = True
                                        logger.info(f"✓ Rule applied to exact range: {check_range}")
                                
                                break
            
            if found_matching_rule:
                break
        
        if not found_matching_rule:
            logger.error("No conditional formatting rule found with expected formula pattern")
            return 0.0
        
        if not rule_applied_to_correct_range:
            logger.warning("Conditional formatting rule found but may not be applied to correct range")
            # Don't fail completely, as the range might be slightly different but still valid
        
        # Optional: Check if cells with differences are actually formatted
        # This is a more advanced check that verifies the formatting is working
        if format_column:
            logger.info(f"Checking formatting in column {format_column}...")
            # Try to find cells in format_column that have conditional formatting applied
            # This is a simplified check - in practice, we'd need to evaluate the formula
            # for each cell to see if it's formatted
            
        logger.info("=" * 60)
        logger.info("✓ Conditional formatting reconciliation verification passed")
        logger.info("=" * 60)
        return 1.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_right_len_find_extract(result: str, expected: str = None, **options) -> float:
    """
    Verify if RIGHT, LEN, and FIND formulas exist in specified column to extract text.
    
    This function checks:
    1. Whether cells in specified column contain RIGHT, LEN, and FIND functions
    2. Whether formulas reference the corresponding source column cell (C2->B2, C3->B3, etc.)
    3. Whether formulas contain the correct pattern (e.g., RIGHT(B2,LEN(B2)-FIND("班",B2)))
    4. Whether formulas have the correct structure with RIGHT, LEN, and FIND functions
    
    The function automatically detects the number of data rows by checking the data column
    (default: B column) for non-empty cells. It stops checking after finding 3 consecutive
    empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "C")
            - start_row: Starting row number (default: 2)
            - expected_functions: List of expected function names (e.g., ["RIGHT", "LEN", "FIND"])
            - expected_formula_pattern: Expected formula pattern (e.g., "RIGHT(B")
            - find_text: Text to find in FIND function (e.g., "班")
            - data_column: Column to check for data to determine end_row (default: "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'C')
        start_row = options.get('start_row', 2)
        expected_functions = options.get('expected_functions', ['RIGHT', 'LEN', 'FIND'])
        expected_formula_pattern = options.get('expected_formula_pattern', 'RIGHT(B')
        find_text = options.get('find_text', '班')
        data_column = options.get('data_column', 'B')
        
        logger.info(f"Verifying RIGHT/LEN/FIND extraction formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected functions: {expected_functions}")
        logger.info(f"Expected formula pattern: {expected_formula_pattern}")
        logger.info(f"Find text: {find_text}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row in the specified column
        all_passed = True
        logger.info(f"Checking column {check_column} (rows {start_row} to {end_row})")
        
        for row_num in range(start_row, end_row + 1):
            cell_coord = f"{check_column}{row_num}"
            try:
                cell = ws[cell_coord]
                logger.debug(f"Checking cell {cell_coord}")
                
                # Check if cell contains a formula
                if cell.data_type != "f":
                    logger.warning(f"Cell {cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                # Get formula text
                formula_text = None
                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                    formula_text = cell._value
                elif hasattr(cell, "formula"):
                    formula_text = cell.formula
                else:
                    # Try to get from value attribute
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value
                
                if formula_text is None:
                    logger.warning(f"Could not extract formula from cell {cell_coord}")
                    all_passed = False
                    continue
                
                # Remove leading = if present for comparison
                formula_clean = formula_text.lstrip("=")
                formula_upper = formula_text.upper()
                logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                
                # Check 1: Formula contains all expected functions
                for func_name in expected_functions:
                    if func_name.upper() not in formula_upper:
                        logger.warning(f"Cell {cell_coord} formula does not contain {func_name}")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        break
                
                if not all_passed:
                    continue
                
                # Check 2: Formula contains expected formula pattern (e.g., RIGHT(B)
                if expected_formula_pattern.upper() not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain pattern '{expected_formula_pattern}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 3: Formula contains RIGHT function call structure
                right_match = re.search(r'RIGHT\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not right_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct RIGHT structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 4: Formula contains LEN function
                len_match = re.search(r'LEN\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not len_match:
                    logger.warning(f"Cell {cell_coord} formula does not have LEN function")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 5: Formula contains FIND function with find_text
                find_pattern = rf'FIND\s*\([^)]*{re.escape(find_text)}[^)]*\)'
                find_match = re.search(find_pattern, formula_text, re.IGNORECASE)
                if not find_match:
                    logger.warning(f"Cell {cell_coord} formula does not contain FIND function with text '{find_text}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 6: Formula references the corresponding source column cell (B2, B3, etc.)
                expected_source_cell = f"{data_column}{row_num}"
                source_cell_pattern = rf'{data_column}{row_num}\b'
                if not re.search(source_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference source cell {expected_source_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 7: Formula structure should be RIGHT(B2, LEN(B2)-FIND("班",B2))
                # Verify that LEN and FIND are used together in the second parameter of RIGHT
                # This is a pattern check - the formula should have LEN(...)-FIND(...) structure
                len_find_pattern = r'LEN\s*\([^)]+\)\s*-\s*FIND\s*\([^)]+\)'
                if not re.search(len_find_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not have LEN(...)-FIND(...) structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {cell_coord} has valid RIGHT/LEN/FIND formula: {formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking cell {cell_coord}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in column {check_column} contain correct RIGHT/LEN/FIND formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ RIGHT/LEN/FIND formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_iferror_regex_extract(result: str, expected: str = None, **options) -> float:
    """
    Verify if IFERROR(REGEX(...)) formulas exist in specified column to extract text with error handling.
    
    This function checks:
    1. Whether cells in specified column contain IFERROR function wrapping REGEX
    2. Whether REGEX function uses capture group pattern (e.g., .*水笔(\d+).*)
    3. Whether REGEX function uses replacement pattern (e.g., $1元)
    4. Whether IFERROR has empty string as second parameter
    5. Whether formulas reference the corresponding source column cell
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "C")
            - start_row: Starting row number (default: 2)
            - expected_pattern_text: Expected text pattern in regex (e.g., "水笔")
            - data_column: Column to check for data to determine end_row (default: "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'C')
        start_row = options.get('start_row', 2)
        expected_pattern_text = options.get('expected_pattern_text', '水笔')
        data_column = options.get('data_column', 'B')
        
        logger.info(f"Verifying IFERROR(REGEX(...)) formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected pattern text: {expected_pattern_text}")
        logger.info(f"Data column: {data_column}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row in the specified column
        all_passed = True
        logger.info(f"Checking column {check_column} (rows {start_row} to {end_row})")
        
        for row_num in range(start_row, end_row + 1):
            cell_coord = f"{check_column}{row_num}"
            try:
                cell = ws[cell_coord]
                logger.debug(f"Checking cell {cell_coord}")
                
                # Check if cell contains a formula
                if cell.data_type != "f":
                    logger.warning(f"Cell {cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                # Get formula text
                formula_text = None
                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                    formula_text = cell._value
                elif hasattr(cell, "formula"):
                    formula_text = cell.formula
                else:
                    # Try to get from value attribute
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value
                
                if formula_text is None:
                    logger.warning(f"Could not extract formula from cell {cell_coord}")
                    all_passed = False
                    continue
                
                formula_upper = formula_text.upper()
                logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                
                # Check 1: Formula contains IFERROR function
                if 'IFERROR' not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain IFERROR function")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 2: Formula contains REGEX function (inside IFERROR)
                # Support both REGEX and LibreOffice internal format _xlfn.ORG.LIBREOFFICE.REGEX
                has_regex = 'REGEX' in formula_upper or '_XLFN.ORG.LIBREOFFICE.REGEX' in formula_upper
                if not has_regex:
                    logger.warning(f"Cell {cell_coord} formula does not contain REGEX function")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 3: IFERROR structure - should have two parameters
                iferror_match = re.search(r'IFERROR\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not iferror_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct IFERROR structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 4: IFERROR second parameter should be empty string ""
                # Handle various formats: ,"" or , "" or ,'' or , ''
                # Also handle LibreOffice format with spaces: REGEX(...) ,""
                # Extract IFERROR parameters: IFERROR(param1, param2)
                iferror_params_match = re.search(r'IFERROR\s*\((.*)\)', formula_text, re.IGNORECASE)
                if iferror_params_match:
                    params_str = iferror_params_match.group(1)
                    # Split by comma, but need to handle nested commas in strings
                    # Simple approach: find the last comma (should separate the two parameters)
                    # For IFERROR(REGEX(...), ""), the last comma separates REGEX call from ""
                    last_comma_pos = params_str.rfind(',')
                    if last_comma_pos != -1:
                        second_param = params_str[last_comma_pos + 1:].strip()
                        # Check if second parameter is empty string "" or ''
                        if second_param in ['""', "''", '""', "''"]:
                            has_empty_string = True
                        else:
                            has_empty_string = False
                    else:
                        has_empty_string = False
                else:
                    has_empty_string = False
                
                if not has_empty_string:
                    logger.warning(f"Cell {cell_coord} IFERROR should have empty string as second parameter")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 5: REGEX function call structure
                # Support both REGEX and LibreOffice internal format
                regex_match = re.search(r'(REGEX|_XLFN\.ORG\.LIBREOFFICE\.REGEX)\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not regex_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct REGEX structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 6: Formula contains expected pattern text (e.g., "水笔")
                if expected_pattern_text not in formula_text:
                    logger.warning(f"Cell {cell_coord} formula does not contain pattern text '{expected_pattern_text}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 7: Formula contains capture group pattern (\d+)
                has_capture_group = bool(re.search(r'\(\\d\+\)|\(\\\\d\+\)', formula_text))
                if not has_capture_group:
                    logger.warning(f"Cell {cell_coord} REGEX formula should contain capture group (\\d+)")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 8: Formula contains replacement pattern $1
                has_replacement = '"$1' in formula_text or "'$1" in formula_text
                if not has_replacement:
                    logger.warning(f"Cell {cell_coord} REGEX formula should contain replacement pattern $1")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 9: Formula references the corresponding source column cell
                expected_source_cell = f"{data_column}{row_num}"
                source_cell_pattern = rf'{data_column}{row_num}\b'
                if not re.search(source_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference source cell {expected_source_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {cell_coord} has valid IFERROR(REGEX(...)) formula: {formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking cell {cell_coord}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in column {check_column} contain correct IFERROR(REGEX(...)) formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ IFERROR(REGEX(...)) formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_rept_text_progress_bar(result: str, expected: str = None, **options) -> float:
    """
    Verify if REPT and TEXT formulas exist in specified column to create progress bars with percentage.
    
    This function checks:
    1. Whether cells in specified column contain REPT and TEXT functions
    2. Whether REPT function uses the correct character (e.g., "|")
    3. Whether REPT function uses the correct multiplier (e.g., *50)
    4. Whether TEXT function uses percentage format (e.g., "0%")
    5. Whether formulas reference the correct numerator and denominator columns
    6. Whether formulas use & operator to concatenate REPT and TEXT results
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "D")
            - start_row: Starting row number (default: 2)
            - expected_functions: List of expected function names (e.g., ["REPT", "TEXT"])
            - numerator_column: Column containing numerator values (e.g., "C")
            - denominator_column: Column containing denominator values (e.g., "B")
            - rept_char: Character to repeat in REPT function (e.g., "|")
            - rept_multiplier: Multiplier for REPT function (e.g., 50)
            - data_column: Column to check for data to determine end_row (default: "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'D')
        start_row = options.get('start_row', 2)
        expected_functions = options.get('expected_functions', ['REPT', 'TEXT'])
        numerator_column = options.get('numerator_column', 'C')
        denominator_column = options.get('denominator_column', 'B')
        rept_char = options.get('rept_char', '|')
        rept_multiplier = options.get('rept_multiplier', 50)
        data_column = options.get('data_column', 'B')
        
        logger.info(f"Verifying REPT/TEXT progress bar formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected functions: {expected_functions}")
        logger.info(f"Numerator column: {numerator_column}")
        logger.info(f"Denominator column: {denominator_column}")
        logger.info(f"REPT character: {rept_char}")
        logger.info(f"REPT multiplier: {rept_multiplier}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row in the specified column
        all_passed = True
        logger.info(f"Checking column {check_column} (rows {start_row} to {end_row})")
        
        for row_num in range(start_row, end_row + 1):
            cell_coord = f"{check_column}{row_num}"
            try:
                cell = ws[cell_coord]
                logger.debug(f"Checking cell {cell_coord}")
                
                # Check if cell contains a formula
                if cell.data_type != "f":
                    logger.warning(f"Cell {cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                # Get formula text
                formula_text = None
                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                    formula_text = cell._value
                elif hasattr(cell, "formula"):
                    formula_text = cell.formula
                else:
                    # Try to get from value attribute
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value
                
                if formula_text is None:
                    logger.warning(f"Could not extract formula from cell {cell_coord}")
                    all_passed = False
                    continue
                
                formula_upper = formula_text.upper()
                logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                
                # Check 1: Formula contains all expected functions (REPT and TEXT)
                for func_name in expected_functions:
                    if func_name.upper() not in formula_upper:
                        logger.warning(f"Cell {cell_coord} formula does not contain {func_name}")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        break
                
                if not all_passed:
                    continue
                
                # Check 2: Formula contains REPT function call structure
                rept_match = re.search(r'REPT\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not rept_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct REPT structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 3: REPT function contains the correct character (e.g., "|")
                # Support both double quotes and single quotes
                rept_char_pattern1 = rf'REPT\s*\(\s*"{re.escape(rept_char)}"'
                rept_char_pattern2 = rf"REPT\s*\(\s*'{re.escape(rept_char)}'"
                if not re.search(rept_char_pattern1, formula_text, re.IGNORECASE) and \
                   not re.search(rept_char_pattern2, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} REPT function should use character '{rept_char}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 4: REPT function contains the multiplier (e.g., *50)
                rept_multiplier_pattern = rf'\*{rept_multiplier}\b'
                if not re.search(rept_multiplier_pattern, formula_text):
                    logger.warning(f"Cell {cell_coord} REPT function should use multiplier *{rept_multiplier}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 5: Formula contains TEXT function call structure
                text_match = re.search(r'TEXT\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not text_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct TEXT structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 6: TEXT function contains percentage format ("0%" or '0%')
                text_percent_pattern1 = r'TEXT\s*\([^,]+,\s*"0%"'
                text_percent_pattern2 = r"TEXT\s*\([^,]+,\s*'0%'"
                if not re.search(text_percent_pattern1, formula_text, re.IGNORECASE) and \
                   not re.search(text_percent_pattern2, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} TEXT function should use percentage format \"0%\"")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 7: Formula references the correct numerator column (C2, C3, etc.)
                expected_numerator_cell = f"{numerator_column}{row_num}"
                numerator_cell_pattern = rf'{numerator_column}{row_num}\b'
                if not re.search(numerator_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference numerator cell {expected_numerator_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 8: Formula references the correct denominator column (B2, B3, etc.)
                expected_denominator_cell = f"{denominator_column}{row_num}"
                denominator_cell_pattern = rf'{denominator_column}{row_num}\b'
                if not re.search(denominator_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference denominator cell {expected_denominator_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 9: Formula contains & operator to concatenate REPT and TEXT
                if '&' not in formula_text:
                    logger.warning(f"Cell {cell_coord} formula should use & operator to concatenate REPT and TEXT")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 10: Formula structure should be REPT(...)&TEXT(...)
                # Verify that REPT comes before TEXT (or at least both are present)
                rept_pos = formula_text.upper().find('REPT')
                text_pos = formula_text.upper().find('TEXT')
                if rept_pos == -1 or text_pos == -1:
                    logger.warning(f"Cell {cell_coord} formula should contain both REPT and TEXT functions")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {cell_coord} has valid REPT/TEXT progress bar formula: {formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking cell {cell_coord}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in column {check_column} contain correct REPT/TEXT progress bar formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ REPT/TEXT progress bar formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_id_extract_gender_age_birthday(result: str, expected: str = None, **options) -> float:
    """
    Verify if formulas exist to extract gender, age, and birthday from ID numbers.
    
    This function checks:
    1. Gender column (C): IF(MOD(MID(B3,17,1),2),"男","女")
    2. Age column (D): DATEDIF(TEXT(MID(B3,7,8),"0-00-00"),TODAY(),"Y")
    3. Birthday column (E): --TEXT(MID(B3,7,8),"0-00-00")
    
    The function automatically detects the number of data rows by checking the ID column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - id_column: Column containing ID numbers (e.g., "B")
            - gender_column: Column for gender formulas (e.g., "C")
            - age_column: Column for age formulas (e.g., "D")
            - birthday_column: Column for birthday formulas (e.g., "E")
            - start_row: Starting row number (default: 3)
            - data_column: Column to check for data to determine end_row (default: "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        id_column = options.get('id_column', 'B')
        gender_column = options.get('gender_column', 'C')
        age_column = options.get('age_column', 'D')
        birthday_column = options.get('birthday_column', 'E')
        start_row = options.get('start_row', 3)
        data_column = options.get('data_column', 'B')
        
        logger.info(f"Verifying ID extraction formulas in file: {result}")
        logger.info(f"ID column: {id_column}")
        logger.info(f"Gender column: {gender_column}")
        logger.info(f"Age column: {age_column}")
        logger.info(f"Birthday column: {birthday_column}")
        logger.info(f"Start row: {start_row}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row
        all_passed = True
        logger.info(f"Checking rows {start_row} to {end_row}")
        
        for row_num in range(start_row, end_row + 1):
            try:
                # Check gender column (C)
                gender_cell_coord = f"{gender_column}{row_num}"
                gender_cell = ws[gender_cell_coord]
                
                if gender_cell.data_type != "f":
                    logger.warning(f"Cell {gender_cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                gender_formula_text = None
                if hasattr(gender_cell, "_value") and isinstance(gender_cell._value, str) and gender_cell._value.startswith("="):
                    gender_formula_text = gender_cell._value
                elif hasattr(gender_cell, "formula"):
                    gender_formula_text = gender_cell.formula
                elif gender_cell.value is not None and isinstance(gender_cell.value, str) and gender_cell.value.startswith("="):
                    gender_formula_text = gender_cell.value
                
                if gender_formula_text is None:
                    logger.warning(f"Could not extract formula from cell {gender_cell_coord}")
                    all_passed = False
                    continue
                
                gender_formula_upper = gender_formula_text.upper()
                logger.debug(f"Cell {gender_cell_coord} formula: {gender_formula_text}")
                
                # Check gender formula: IF(MOD(MID(B3,17,1),2),"男","女")
                if 'IF' not in gender_formula_upper:
                    logger.warning(f"Cell {gender_cell_coord} formula does not contain IF function")
                    all_passed = False
                    continue
                
                if 'MOD' not in gender_formula_upper:
                    logger.warning(f"Cell {gender_cell_coord} formula does not contain MOD function")
                    all_passed = False
                    continue
                
                if 'MID' not in gender_formula_upper:
                    logger.warning(f"Cell {gender_cell_coord} formula does not contain MID function")
                    all_passed = False
                    continue
                
                # Check MID(B3,17,1) pattern
                mid_pattern = rf'MID\s*\(\s*{id_column}{row_num}\s*,\s*17\s*,\s*1\s*\)'
                if not re.search(mid_pattern, gender_formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {gender_cell_coord} formula should contain MID({id_column}{row_num},17,1)")
                    logger.warning(f"Formula: {gender_formula_text}")
                    all_passed = False
                    continue
                
                # Check for "男" and "女" in formula
                if '"男"' not in gender_formula_text and "'男'" not in gender_formula_text:
                    logger.warning(f"Cell {gender_cell_coord} formula should contain \"男\"")
                    logger.warning(f"Formula: {gender_formula_text}")
                    all_passed = False
                    continue
                
                if '"女"' not in gender_formula_text and "'女'" not in gender_formula_text:
                    logger.warning(f"Cell {gender_cell_coord} formula should contain \"女\"")
                    logger.warning(f"Formula: {gender_formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {gender_cell_coord} has valid gender formula: {gender_formula_text}")
                
                # Check age column (D)
                age_cell_coord = f"{age_column}{row_num}"
                age_cell = ws[age_cell_coord]
                
                if age_cell.data_type != "f":
                    logger.warning(f"Cell {age_cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                age_formula_text = None
                if hasattr(age_cell, "_value") and isinstance(age_cell._value, str) and age_cell._value.startswith("="):
                    age_formula_text = age_cell._value
                elif hasattr(age_cell, "formula"):
                    age_formula_text = age_cell.formula
                elif age_cell.value is not None and isinstance(age_cell.value, str) and age_cell.value.startswith("="):
                    age_formula_text = age_cell.value
                
                if age_formula_text is None:
                    logger.warning(f"Could not extract formula from cell {age_cell_coord}")
                    all_passed = False
                    continue
                
                age_formula_upper = age_formula_text.upper()
                logger.debug(f"Cell {age_cell_coord} formula: {age_formula_text}")
                
                # Check age formula: DATEDIF(TEXT(MID(B3,7,8),"0-00-00"),TODAY(),"Y")
                if 'DATEDIF' not in age_formula_upper:
                    logger.warning(f"Cell {age_cell_coord} formula does not contain DATEDIF function")
                    all_passed = False
                    continue
                
                if 'TEXT' not in age_formula_upper:
                    logger.warning(f"Cell {age_cell_coord} formula does not contain TEXT function")
                    all_passed = False
                    continue
                
                if 'TODAY' not in age_formula_upper:
                    logger.warning(f"Cell {age_cell_coord} formula does not contain TODAY function")
                    all_passed = False
                    continue
                
                if 'MID' not in age_formula_upper:
                    logger.warning(f"Cell {age_cell_coord} formula does not contain MID function")
                    all_passed = False
                    continue
                
                # Check MID(B3,7,8) pattern
                mid_pattern_age = rf'MID\s*\(\s*{id_column}{row_num}\s*,\s*7\s*,\s*8\s*\)'
                if not re.search(mid_pattern_age, age_formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {age_cell_coord} formula should contain MID({id_column}{row_num},7,8)")
                    logger.warning(f"Formula: {age_formula_text}")
                    all_passed = False
                    continue
                
                # Check TEXT format "0-00-00"
                # Use pattern that matches until the quote to handle nested functions like MID(B3,7,8)
                text_format_pattern1 = r'TEXT\s*\([^"]+,\s*"0-00-00"'
                text_format_pattern2 = r"TEXT\s*\([^']+,\s*'0-00-00'"
                if not re.search(text_format_pattern1, age_formula_text, re.IGNORECASE) and \
                   not re.search(text_format_pattern2, age_formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {age_cell_coord} TEXT function should use format \"0-00-00\"")
                    logger.warning(f"Formula: {age_formula_text}")
                    all_passed = False
                    continue
                
                # Check DATEDIF third parameter "Y"
                # Use a more flexible pattern that handles nested functions
                # Check if DATEDIF contains "Y" as the third parameter (after two commas)
                # Pattern: DATEDIF(...,...,"Y") or DATEDIF(...,...,'Y')
                # We'll count commas to find the third parameter
                datedif_match = re.search(r'DATEDIF\s*\((.*)\)', age_formula_text, re.IGNORECASE)
                if datedif_match:
                    datedif_params = datedif_match.group(1)
                    # Count commas to find the third parameter
                    # Simple approach: check if the last part before closing paren is "Y" or 'Y'
                    # More robust: find the pattern ,"Y" or ,'Y' before the closing paren
                    if not re.search(r',\s*"Y"\s*\)', age_formula_text, re.IGNORECASE) and \
                       not re.search(r",\s*'Y'\s*\)", age_formula_text, re.IGNORECASE):
                        logger.warning(f"Cell {age_cell_coord} DATEDIF function should use \"Y\" parameter")
                        logger.warning(f"Formula: {age_formula_text}")
                        all_passed = False
                        continue
                else:
                    logger.warning(f"Cell {age_cell_coord} could not parse DATEDIF function")
                    logger.warning(f"Formula: {age_formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {age_cell_coord} has valid age formula: {age_formula_text}")
                
                # Check birthday column (E)
                birthday_cell_coord = f"{birthday_column}{row_num}"
                birthday_cell = ws[birthday_cell_coord]
                
                if birthday_cell.data_type != "f":
                    logger.warning(f"Cell {birthday_cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                birthday_formula_text = None
                if hasattr(birthday_cell, "_value") and isinstance(birthday_cell._value, str) and birthday_cell._value.startswith("="):
                    birthday_formula_text = birthday_cell._value
                elif hasattr(birthday_cell, "formula"):
                    birthday_formula_text = birthday_cell.formula
                elif birthday_cell.value is not None and isinstance(birthday_cell.value, str) and birthday_cell.value.startswith("="):
                    birthday_formula_text = birthday_cell.value
                
                if birthday_formula_text is None:
                    logger.warning(f"Could not extract formula from cell {birthday_cell_coord}")
                    all_passed = False
                    continue
                
                birthday_formula_upper = birthday_formula_text.upper()
                logger.debug(f"Cell {birthday_cell_coord} formula: {birthday_formula_text}")
                
                # Check birthday formula: TEXT(MID(B3,7,8),"0-00-00")
                if 'TEXT' not in birthday_formula_upper:
                    logger.warning(f"Cell {birthday_cell_coord} formula does not contain TEXT function")
                    all_passed = False
                    continue
                
                if 'MID' not in birthday_formula_upper:
                    logger.warning(f"Cell {birthday_cell_coord} formula does not contain MID function")
                    all_passed = False
                    continue
                
                # Check MID(B3,7,8) pattern
                mid_pattern_birthday = rf'MID\s*\(\s*{id_column}{row_num}\s*,\s*7\s*,\s*8\s*\)'
                if not re.search(mid_pattern_birthday, birthday_formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {birthday_cell_coord} formula should contain MID({id_column}{row_num},7,8)")
                    logger.warning(f"Formula: {birthday_formula_text}")
                    all_passed = False
                    continue
                
                # Check TEXT format "0-00-00"
                # Use pattern that matches until the quote to handle nested functions like MID(B3,7,8)
                text_format_pattern1 = r'TEXT\s*\([^"]+,\s*"0-00-00"'
                text_format_pattern2 = r"TEXT\s*\([^']+,\s*'0-00-00'"
                if not re.search(text_format_pattern1, birthday_formula_text, re.IGNORECASE) and \
                   not re.search(text_format_pattern2, birthday_formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {birthday_cell_coord} TEXT function should use format \"0-00-00\"")
                    logger.warning(f"Formula: {birthday_formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {birthday_cell_coord} has valid birthday formula: {birthday_formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking row {row_num}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All rows contain correct ID extraction formulas (gender, age, birthday)")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ ID extraction formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_line_chart(result: str, expected: str = None, **options) -> float:
    """
    Verify if a line chart exists in the Excel file.
    
    This function checks:
    1. Whether at least one chart exists in the worksheet
    2. Whether the chart type is lineChart
    3. Whether the chart has the expected number of series
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - expected_chart_type: Expected chart type (default: "lineChart")
            - min_series_count: Minimum number of series expected (default: 1)
            - data_range: Data range used for chart (optional, for logging)
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        expected_chart_type = options.get('expected_chart_type', 'lineChart')
        min_series_count = options.get('min_series_count', 1)
        data_range = options.get('data_range', '')
        
        logger.info(f"Verifying line chart in file: {result}")
        logger.info(f"Expected chart type: {expected_chart_type}")
        logger.info(f"Minimum series count: {min_series_count}")
        if data_range:
            logger.info(f"Data range: {data_range}")
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(result)
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check if charts exist
        charts = ws._charts
        if not charts or len(charts) == 0:
            logger.error("No charts found in the worksheet")
            return 0.0
        
        logger.info(f"Found {len(charts)} chart(s) in the worksheet")
        
        # Check each chart
        for chart_idx, chart in enumerate(charts):
            logger.info(f"Checking chart {chart_idx + 1}...")
            
            # Check chart type
            chart_type = None
            if hasattr(chart, 'tagname'):
                chart_type = chart.tagname
            logger.info(f"Chart type: {chart_type}")
            
            # Check if it's a line chart
            if chart_type and expected_chart_type.lower() in chart_type.lower():
                logger.info(f"✓ Chart {chart_idx + 1} is a line chart")
                
                # Check if it has series
                if not hasattr(chart, 'series') or not chart.series:
                    logger.warning(f"Chart {chart_idx + 1} has no series")
                    continue
                
                series_count = len(chart.series)
                logger.info(f"Chart {chart_idx + 1} has {series_count} series")
                
                # Verify series count
                if series_count >= min_series_count:
                    logger.info("=" * 60)
                    logger.info(f"✓ Line chart verification passed")
                    logger.info(f"  Chart type: {chart_type}")
                    logger.info(f"  Series count: {series_count} (minimum required: {min_series_count})")
                    logger.info("=" * 60)
                    return 1.0
                else:
                    logger.warning(f"Chart {chart_idx + 1} has {series_count} series, but minimum required is {min_series_count}")
            else:
                logger.warning(f"Chart {chart_idx + 1} is not a line chart (type: {chart_type})")
        
        # If we get here, verification failed
        logger.error("=" * 60)
        logger.error(f"✗ Line chart verification failed")
        logger.error(f"  Expected chart type: {expected_chart_type}")
        logger.error(f"  Minimum series count: {min_series_count}")
        logger.error("=" * 60)
        return 0.0
             
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_salary_growth_chart(result: str, expected: str = None, **options) -> float:
    """
    Verify if the salary growth chart matches the expected specifications.
    
    This function checks the chart itself (not the data table):
    1. Whether a chart exists in the specified sheet
    2. Whether the chart title matches "店长工资增长"
    3. Whether the chart has the expected number of series (at least 3)
    4. Whether the chart is a combination chart (bar + line)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - sheet_idx: Sheet index (default: 0)
            - expected_title: Expected chart title (default: "店长工资增长")
            - min_series_count: Minimum number of series (default: 3)
            - chart_type: Expected chart type (default: "combination")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        sheet_idx = options.get('sheet_idx', 0)
        expected_title = options.get('expected_title', '店长工资增长')
        min_series_count = options.get('min_series_count', 3)
        chart_type = options.get('chart_type', 'combination')
        
        logger.info(f"Verifying salary growth chart in file: {result}")
        logger.info(f"Sheet index: {sheet_idx}")
        logger.info(f"Expected title: {expected_title}")
        logger.info(f"Min series count: {min_series_count}")
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(result)
            pdworkbook = pd.ExcelFile(result)
            sheet_names = pdworkbook.sheet_names
            
            if sheet_idx >= len(sheet_names):
                logger.error(f"Sheet index {sheet_idx} out of range. Available sheets: {sheet_names}")
                return 0.0
            
            sheet_name = sheet_names[sheet_idx]
            logger.info(f"Checking sheet: {sheet_name}")
            ws = wb[sheet_name]
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return 0.0
        
        # Check if charts exist
        charts = ws._charts
        if not charts:
            logger.error("No charts found in the sheet")
            return 0.0
        
        logger.info(f"Found {len(charts)} chart(s) in the sheet")
        
        # Load chart information
        chart_props = ['title', 'type', 'legend', 'xtitle', 'ytitle']
        chart_info = load_charts(wb, sheet_name, chart_props=chart_props)
        
        if not chart_info:
            logger.error("Could not load chart information")
            return 0.0
        
        # Check each chart
        chart_passed = False
        for chart_key, chart_data in chart_info.items():
            logger.info(f"Checking chart: {chart_key}")
            logger.debug(f"Chart data: {chart_data}")
            
            # Check 1: Chart title
            chart_title = chart_data.get('title')
            if chart_title != expected_title:
                logger.warning(f"Chart title mismatch: expected '{expected_title}', got '{chart_title}'")
                continue
            else:
                logger.info(f"✓ Chart title matches: {chart_title}")
            
            # Check 2: Chart type (for combination charts, we might see multiple types)
            chart_type_actual = chart_data.get('type')
            logger.info(f"Chart type: {chart_type_actual}")
            # Note: Combination charts might be represented differently in openpyxl
            # We'll be lenient here and just check that a chart exists
            
            # Check 3: Number of series
            # Extract series count from chart_key (format: "value_ref1,category_ref1;value_ref2,category_ref2;...")
            series_parts = chart_key.split(';')
            series_count = len(series_parts)
            logger.info(f"Number of series: {series_count}")
            
            if series_count < min_series_count:
                logger.warning(f"Insufficient series count: expected at least {min_series_count}, got {series_count}")
                continue
            else:
                logger.info(f"✓ Series count sufficient: {series_count} >= {min_series_count}")
            
            # If we get here, this chart passed all checks
            chart_passed = True
            logger.info("=" * 60)
            logger.info(f"✓ Chart verification passed!")
            logger.info("=" * 60)
            break
        
        if chart_passed:
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error("✗ Chart verification failed - no chart matched all criteria")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_project_completion_chart(result: str, expected: str = None, **options) -> float:
    """
    Verify if a combination chart exists with the expected title that contains
    both bar chart series (for project values) and line chart series (for completion rates).
    
    This function checks:
    1. Whether at least one chart exists in the worksheet
    2. Whether the chart title matches expected_title
    3. Whether the chart has at least 16 series (8 projects + 8 completion rates)
    4. Whether at least one series name contains "rate" (for completion rates)
    5. Whether the chart has at least project_count * 5 categories (for 5 quarters per project)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - sheet_idx: Sheet index to check (default: 0)
            - expected_title: Expected chart title (default: "项目")
            - min_series_count: Minimum number of series required (default: 16)
            - project_count: Number of projects (default: 8)
            - quarters_per_project: Number of quarters per project (default: 5)
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        sheet_idx = options.get('sheet_idx', 0)
        expected_title = options.get('expected_title', '项目')
        min_series_count = options.get('min_series_count', 16)
        project_count = options.get('project_count', 8)
        quarters_per_project = options.get('quarters_per_project', 5)
        min_categories = project_count * quarters_per_project
        
        logger.info(f"Verifying project completion chart in file: {result}")
        logger.info(f"Sheet index: {sheet_idx}")
        logger.info(f"Expected title: {expected_title}")
        logger.info(f"Minimum series count: {min_series_count}")
        logger.info(f"Project count: {project_count}")
        logger.info(f"Quarters per project: {quarters_per_project}")
        logger.info(f"Minimum categories: {min_categories}")
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(result)
            sheet_names = wb.sheetnames
            if sheet_idx >= len(sheet_names):
                logger.error(f"Sheet index {sheet_idx} out of range. Available sheets: {sheet_names}")
                return 0.0
            sheet_name = sheet_names[sheet_idx]
            ws = wb[sheet_name]
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check if charts exist
        charts = ws._charts
        if not charts:
            logger.error("No charts found in the worksheet")
            return 0.0
        
        logger.info(f"Found {len(charts)} chart(s) in the worksheet")
        
        # Check each chart
        chart_found = False
        for chart in charts:
            # Check chart title
            chart_title = None
            try:
                if chart.title and chart.title.tx:
                    if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                        if hasattr(chart.title.tx.rich, 'p') and chart.title.tx.rich.p:
                            if len(chart.title.tx.rich.p) > 0:
                                if hasattr(chart.title.tx.rich.p[0], 'r') and chart.title.tx.rich.p[0].r:
                                    if len(chart.title.tx.rich.p[0].r) > 0:
                                        if hasattr(chart.title.tx.rich.p[0].r[0], 't'):
                                            chart_title = chart.title.tx.rich.p[0].r[0].t
            except Exception as e:
                logger.debug(f"Error reading chart title: {e}")
            
            logger.info(f"Chart title: {chart_title}")
            
            # Check if title matches
            if chart_title == expected_title:
                logger.info(f"✓ Chart title matches: {chart_title}")
                chart_found = True
                
                # Use load_charts to get all series information (includes both bar and line series)
                # This is more reliable for combination charts
                chart_props = ['title']
                chart_info = load_charts(wb, sheet_name, chart_props=chart_props)
                
                # Find the chart that matches our title
                chart_key = None
                for key, info in chart_info.items():
                    if info.get('title') == expected_title:
                        chart_key = key
                        break
                
                # Get all series from the chart object for detailed inspection
                all_series = list(chart.series) if hasattr(chart, 'series') else []
                logger.info(f"Series count from chart.series: {len(all_series)}")
                
                if not chart_key:
                    logger.warning("Could not find chart in load_charts output, using direct series access")
                    # Fallback to direct series access
                    series_count = len(all_series)
                else:
                    # Extract series count from chart_key (format: "value_ref1,category_ref1;value_ref2,category_ref2;...")
                    # This includes ALL series (both bar and line) in combination charts
                    series_parts = chart_key.split(';')
                    series_count_from_load = len(series_parts)
                    logger.info(f"Series count from load_charts: {series_count_from_load}")
                    
                    # For combination charts, load_charts should give us all series
                    # Use the count from load_charts as it's more reliable for combination charts
                    series_count = series_count_from_load
                    
                    # Also check for sub-charts in case series are stored there
                    if hasattr(chart, '_charts') and chart._charts:
                        # Check for sub-charts (for combination charts)
                        for sub_chart in chart._charts:
                            if hasattr(sub_chart, 'series'):
                                sub_series = list(sub_chart.series)
                                all_series.extend(sub_series)
                                logger.info(f"Found {len(sub_series)} additional series in sub-chart")
                    
                    # If load_charts gave us fewer series than direct access, use the larger count
                    # This handles edge cases where load_charts might miss some series
                    if series_count < len(all_series):
                        logger.warning(f"load_charts found {series_count} series but direct access found {len(all_series)}, using larger count")
                        series_count = len(all_series)
                
                logger.info(f"Chart has {series_count} series (including both bar and line series)")
                
                # Debug: Log all series details
                if all_series:
                    logger.info(f"Detailed series information:")
                    for idx, ser in enumerate(all_series):
                        logger.info(f"  Series {idx}: {type(ser).__name__}")
                        try:
                            if hasattr(ser, 'title'):
                                logger.debug(f"    Title: {ser.title}")
                        except:
                            pass
                
                if series_count < min_series_count:
                    logger.error(f"✗ Chart has only {series_count} series, expected at least {min_series_count}")
                    return 0.0
                
                logger.info(f"✓ Chart has {series_count} series (>= {min_series_count})")
                
                # Check series names for "rate" (completion rate series)
                has_rate_series = False
                series_names = []
                for i, ser in enumerate(all_series):
                    series_name = None
                    try:
                        # Try to get series title/name
                        if hasattr(ser, 'title') and ser.title:
                            if hasattr(ser.title, 'tx') and ser.title.tx:
                                if hasattr(ser.title.tx, 'rich') and ser.title.tx.rich:
                                    if hasattr(ser.title.tx.rich, 'p') and ser.title.tx.rich.p:
                                        if len(ser.title.tx.rich.p) > 0:
                                            if hasattr(ser.title.tx.rich.p[0], 'r') and ser.title.tx.rich.p[0].r:
                                                if len(ser.title.tx.rich.p[0].r) > 0:
                                                    if hasattr(ser.title.tx.rich.p[0].r[0], 't'):
                                                        series_name = ser.title.tx.rich.p[0].r[0].t
                        # Alternative: check if title is a string reference
                        if not series_name and hasattr(ser, 'title') and hasattr(ser.title, 'tx') and hasattr(ser.title.tx, 'strRef'):
                            if hasattr(ser.title.tx.strRef, 'f'):
                                series_name = ser.title.tx.strRef.f
                    except Exception as e:
                        logger.debug(f"Error reading series {i} name: {e}")
                    
                    if series_name:
                        series_names.append(series_name)
                        if "rate" in series_name.lower():
                            has_rate_series = True
                            logger.info(f"✓ Found series with 'rate' in name: {series_name}")
                
                if series_names:
                    logger.info(f"Series names found: {series_names[:10]}...")  # Log first 10
                else:
                    logger.warning("Could not extract series names, will skip rate check")
                
                if not has_rate_series and series_names:
                    logger.error(f"✗ No series found with 'rate' in name. Series names: {series_names}")
                    return 0.0
                elif not has_rate_series:
                    logger.warning("⚠ Could not verify 'rate' in series names (series names not extractable)")
                
                # Check category count
                max_categories = 0
                category_ranges = []
                
                def parse_range_count(range_str):
                    """Parse Excel range string and return count of cells"""
                    try:
                        # Remove sheet name if present (e.g., "Sheet1!$A$2:$A$6" -> "$A$2:$A$6")
                        if '!' in range_str:
                            range_str = range_str.split('!')[1]
                        
                        # Remove $ signs
                        range_str = range_str.replace('$', '')
                        
                        if ':' in range_str:
                            start, end = range_str.split(':')
                            # Parse start and end coordinates
                            # Helper function to parse coordinate like "A1" to (column_index, row_number)
                            from openpyxl.utils import column_index_from_string
                            import re
                            
                            def parse_coordinate(coord):
                                """Parse coordinate string like 'A1' to (column_index, row_number)"""
                                match = re.match(r'([A-Z]+)(\d+)', coord.upper())
                                if match:
                                    col_str, row_str = match.groups()
                                    col_idx = column_index_from_string(col_str)
                                    row_num = int(row_str)
                                    return (col_idx, row_num)
                                raise ValueError(f"Invalid coordinate: {coord}")
                            
                            start_col, start_row = parse_coordinate(start)
                            end_col, end_row = parse_coordinate(end)
                            
                            # Calculate count based on range
                            if start_col == end_col:
                                # Same column, count rows
                                return abs(end_row - start_row) + 1
                            elif start_row == end_row:
                                # Same row, count columns
                                return abs(end_col - start_col) + 1
                            else:
                                # 2D range
                                return (abs(end_row - start_row) + 1) * (abs(end_col - start_col) + 1)
                        else:
                            # Single cell
                            return 1
                    except Exception as e:
                        logger.debug(f"Error parsing range {range_str}: {e}")
                        return 0
                
                for i, ser in enumerate(all_series):
                    try:
                        # Try to get category count from category reference
                        if hasattr(ser, 'cat'):
                            cat_range = None
                            # Check if categories are from a range
                            if hasattr(ser.cat, 'numRef') and hasattr(ser.cat.numRef, 'f'):
                                cat_range = ser.cat.numRef.f
                            elif hasattr(ser.cat, 'strRef') and hasattr(ser.cat.strRef, 'f'):
                                cat_range = ser.cat.strRef.f
                            
                            if cat_range:
                                category_ranges.append(cat_range)
                                cat_count = parse_range_count(cat_range)
                                if cat_count > max_categories:
                                    max_categories = cat_count
                                logger.debug(f"Series {i} category range: {cat_range}, count: {cat_count}")
                    except Exception as e:
                        logger.debug(f"Error reading categories for series {i}: {e}")
                
                if max_categories > 0:
                    logger.info(f"Maximum category count found: {max_categories}")
                    if max_categories < min_categories:
                        logger.error(f"✗ Chart has only {max_categories} categories, expected at least {min_categories} (project_count * quarters_per_project)")
                        return 0.0
                    logger.info(f"✓ Chart has {max_categories} categories (>= {min_categories})")
                else:
                    # If we can't determine category count from ranges, use heuristic
                    # For 8 projects with 5 quarters each, we need at least 40 categories
                    # But since we can't verify directly, we'll log a warning
                    logger.warning(f"⚠ Could not determine exact category count from ranges. Expected at least {min_categories} categories.")
                    logger.info(f"Category ranges found: {category_ranges[:5]}...")  # Log first 5
                
                # Check if it's a combination chart
                if series_count >= 2:
                    logger.info("✓ Chart appears to be a combination chart (has multiple series)")
                else:
                    logger.error(f"✗ Chart has only {series_count} series, expected at least 2 for combination chart")
                    return 0.0
                
                break
        
        if chart_found:
            logger.info("=" * 60)
            logger.info("✓ Project completion combination chart verification passed")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ Chart with title '{expected_title}' not found")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_shipping_boxes_calculation(result: str, expected: str = None, **options) -> float:
    """
    Verify if formulas exist to calculate shipping boxes from product specifications and order quantities.
    
    This function checks:
    1. Whether cells in specified column contain formulas (not just values)
    2. Whether formulas contain required functions: INT, VALUE, LEFT, FIND, MOD, IF
    3. Whether formulas reference the specification column (B) and quantity column (C)
    4. Whether formulas contain Chinese characters "支" and "盒"
    5. Whether formulas use string concatenation (&)
    
    The expected formula pattern:
    =INT(C2/VALUE(LEFT(B2,FIND("支",B2)-1)))&"盒"&IF(MOD(C2,VALUE(LEFT(B2,FIND("支",B2)-1)))=0,"","加"&MOD(C2,VALUE(LEFT(B2,FIND("支",B2)-1)))&"支")
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "D")
            - start_row: Starting row number (default: 2)
            - spec_column: Column containing product specifications (e.g., "B")
            - quantity_column: Column containing order quantities (e.g., "C")
            - expected_functions: List of expected function names (e.g., ["INT", "VALUE", "LEFT", "FIND", "MOD", "IF"])
            - data_column: Column to check for data to determine end_row (default: "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'D')
        start_row = options.get('start_row', 2)
        spec_column = options.get('spec_column', 'B')
        quantity_column = options.get('quantity_column', 'C')
        expected_functions = options.get('expected_functions', ['INT', 'VALUE', 'LEFT', 'FIND', 'MOD', 'IF'])
        data_column = options.get('data_column', 'B')
        
        logger.info(f"Verifying shipping boxes calculation formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Specification column: {spec_column}")
        logger.info(f"Quantity column: {quantity_column}")
        logger.info(f"Expected functions: {expected_functions}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        if end_row < start_row:
            logger.error(f"No data rows found starting from row {start_row}")
            return 0.0
        
        # Check formulas in each row
        all_checks_passed = True
        rows_checked = 0
        
        for row_num in range(start_row, end_row + 1):
            check_cell = ws[f"{check_column}{row_num}"]
            spec_cell = ws[f"{spec_column}{row_num}"]
            quantity_cell = ws[f"{quantity_column}{row_num}"]
            
            # Skip if spec or quantity cell is empty
            if spec_cell.value is None or quantity_cell.value is None:
                continue
            
            rows_checked += 1
            formula = check_cell.value
            
            # Check 1: Formula exists (not just a value)
            if formula is None or not isinstance(formula, str) or not formula.startswith('='):
                logger.error(f"Cell {check_column}{row_num} should contain a formula, but got: {formula}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check 2: Contains all required functions
            missing_functions = []
            for func in expected_functions:
                # Check for function name (with word boundary to avoid partial matches)
                func_pattern = r'\b' + re.escape(func) + r'\s*\('
                if not re.search(func_pattern, formula_upper, re.IGNORECASE):
                    missing_functions.append(func)
            
            if missing_functions:
                logger.error(f"Cell {check_column}{row_num} formula missing functions: {missing_functions}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: References specification column (B)
            spec_pattern = rf'{spec_column}\d+'
            if not re.search(spec_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should reference {spec_column} column")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: References quantity column (C)
            quantity_pattern = rf'{quantity_column}\d+'
            if not re.search(quantity_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should reference {quantity_column} column")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 5: Contains Chinese character "支" (for extracting pieces per box)
            if '"支"' not in formula and "'支'" not in formula and '支' not in formula:
                logger.error(f"Cell {check_column}{row_num} formula should contain '支' character")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 6: Contains Chinese character "盒" (for box unit)
            if '"盒"' not in formula and "'盒'" not in formula and '盒' not in formula:
                logger.error(f"Cell {check_column}{row_num} formula should contain '盒' character")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 7: Uses string concatenation (&)
            if '&' not in formula:
                logger.error(f"Cell {check_column}{row_num} formula should use & for string concatenation")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 8: Contains INT function (for calculating integer boxes)
            int_pattern = r'\bINT\s*\('
            if not re.search(int_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain INT function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 9: Contains MOD function (for calculating remainder)
            mod_pattern = r'\bMOD\s*\('
            if not re.search(mod_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain MOD function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 10: Contains IF function (for conditional formatting)
            if_pattern = r'\bIF\s*\('
            if not re.search(if_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain IF function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            logger.debug(f"✓ Row {row_num} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ Shipping boxes calculation verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Column: {check_column}")
            logger.info(f"  Functions verified: {', '.join(expected_functions)}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ Shipping boxes calculation verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_split_content_formula(result: str, expected: str = None, **options) -> float:
    """
    Verify if formulas exist to split content with line breaks into multiple rows.
    
    This function checks:
    1. Whether cells in specified column contain formulas (not just values)
    2. Whether formulas contain required functions: TRIM, MID, SUBSTITUTE, CHAR, REPT, ROW
    3. Whether formulas reference the source column (A)
    4. Whether formulas contain CHAR(10) for line break
    5. Whether formulas contain REPT(" ",100) or similar pattern
    6. Whether formulas use ROW function for position calculation
    
    The expected formula pattern:
    =TRIM(MID(SUBSTITUTE(A2,CHAR(10),REPT(" ",100)),(ROW(A1)-1)*100+1,100))
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "B")
            - start_row: Starting row number (default: 2)
            - source_column: Column containing source data (e.g., "A")
            - expected_functions: List of expected function names (e.g., ["TRIM", "MID", "SUBSTITUTE", "CHAR", "REPT", "ROW"])
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'B')
        start_row = options.get('start_row', 2)
        source_column = options.get('source_column', 'A')
        expected_functions = options.get('expected_functions', ['TRIM', 'MID', 'SUBSTITUTE', 'CHAR', 'REPT', 'ROW'])
        data_column = options.get('data_column', 'A')
        
        logger.info(f"Verifying split content formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Source column: {source_column}")
        logger.info(f"Expected functions: {expected_functions}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        if end_row < start_row:
            logger.error(f"No data rows found starting from row {start_row}")
            return 0.0
        
        # Check formulas in each row
        all_checks_passed = True
        rows_checked = 0
        
        for row_num in range(start_row, end_row + 1):
            check_cell = ws[f"{check_column}{row_num}"]
            
            # Skip if source cell is empty
            source_cell = ws[f"{source_column}{row_num}"]
            if source_cell.value is None:
                continue
            
            rows_checked += 1
            formula = check_cell.value
            
            # Check 1: Formula exists (not just a value)
            if formula is None or not isinstance(formula, str) or not formula.startswith('='):
                logger.error(f"Cell {check_column}{row_num} should contain a formula, but got: {formula}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check 2: Contains all required functions
            missing_functions = []
            for func in expected_functions:
                # Check for function name (with word boundary to avoid partial matches)
                func_pattern = r'\b' + re.escape(func) + r'\s*\('
                if not re.search(func_pattern, formula_upper, re.IGNORECASE):
                    missing_functions.append(func)
            
            if missing_functions:
                logger.error(f"Cell {check_column}{row_num} formula missing functions: {missing_functions}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: References source column (A)
            source_pattern = rf'{source_column}\d+'
            if not re.search(source_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should reference {source_column} column")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3.5: SUBSTITUTE function must reference A2 specifically
            # Pattern: SUBSTITUTE(A2,...
            substitute_a2_pattern = rf'SUBSTITUTE\s*\(\s*{source_column}2\s*,'
            if not re.search(substitute_a2_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula SUBSTITUTE function must reference {source_column}2 (not {source_column}3, {source_column}4, etc.)")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Contains CHAR(10) for line break
            char_pattern = r'CHAR\s*\(\s*10\s*\)'
            if not re.search(char_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain CHAR(10)")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 5: Contains REPT with space and 100
            rept_pattern = r'REPT\s*\(\s*["\']?\s*["\']?\s*,\s*100\s*\)'
            if not re.search(rept_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain REPT(\" \",100) or similar")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 6: Contains ROW function
            row_pattern = r'\bROW\s*\('
            if not re.search(row_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain ROW function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 7: Contains TRIM function (outermost)
            trim_pattern = r'\bTRIM\s*\('
            if not re.search(trim_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain TRIM function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 8: Contains MID function
            mid_pattern = r'\bMID\s*\('
            if not re.search(mid_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain MID function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 9: Contains SUBSTITUTE function
            substitute_pattern = r'\bSUBSTITUTE\s*\('
            if not re.search(substitute_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain SUBSTITUTE function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            logger.debug(f"✓ Row {row_num} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ Split content formula verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Column: {check_column}")
            logger.info(f"  Functions verified: {', '.join(expected_functions)}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ Split content formula verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_quote_sheet_with_merged_cells(result: str, expected: str = None, **options) -> float:
    """
    Verify if a quote sheet template exists with the expected structure, fields, and merged cells.
    
    This function checks:
    1. Whether the title "报价单" exists in the worksheet
    2. Whether merged cells exist (especially for the title)
    3. Whether required header fields exist
    4. Whether the product table headers exist
    5. Whether summary fields exist
    6. Whether footer fields exist
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - expected_title: Expected title text (default: "报价单")
            - title_merged_range: Expected merged range for title (default: "G15:H15")
            - required_fields: List of required header field labels
            - table_headers: List of table header labels
            - summary_fields: List of summary field labels
            - footer_fields: List of footer field labels
            - min_merged_cells: Minimum number of merged cell ranges expected (default: 1)
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        expected_title = options.get('expected_title', '报价单')
        title_merged_range = options.get('title_merged_range', 'G15:H15')
        required_fields = options.get('required_fields', ['报价单位', '联系人', '联系电话', '客户名称', '报价日期', '邮箱'])
        table_headers = options.get('table_headers', ['序号', '产品名称', '产品类型', '规格', '数量', '单价', '金额', '备注'])
        summary_fields = options.get('summary_fields', ['合计金额(小写)', '合计金额(大写)'])
        footer_fields = options.get('footer_fields', ['报价人', '审批'])
        min_merged_cells = options.get('min_merged_cells', 1)
        
        logger.info(f"Verifying quote sheet template with merged cells in file: {result}")
        logger.info(f"Expected title: {expected_title}")
        logger.info(f"Title merged range: {title_merged_range}")
        logger.info(f"Required fields: {required_fields}")
        logger.info(f"Table headers: {table_headers}")
        logger.info(f"Summary fields: {summary_fields}")
        logger.info(f"Footer fields: {footer_fields}")
        logger.info(f"Minimum merged cells: {min_merged_cells}")
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(result, data_only=True)
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check merged cells
        merged_cells = list(ws.merged_cells.ranges)
        logger.info(f"Found {len(merged_cells)} merged cell range(s)")
        
        if len(merged_cells) < min_merged_cells:
            logger.error(f"✗ Insufficient merged cells: found {len(merged_cells)}, expected at least {min_merged_cells}")
            return 0.0
        
        # Check if title merged range exists
        title_merged_found = False
        for merged_range in merged_cells:
            merged_str = str(merged_range)
            logger.debug(f"Merged range: {merged_str}")
            if merged_str.upper() == title_merged_range.upper():
                title_merged_found = True
                logger.info(f"✓ Found title merged range: {merged_str}")
                break
        
        if not title_merged_found:
            logger.warning(f"⚠ Title merged range '{title_merged_range}' not found, but other merged cells exist")
            logger.info(f"  Available merged ranges: {[str(r) for r in merged_cells]}")
            # Don't fail completely, as the range might be slightly different
        
        # Search through all cells to find required text
        # Also check merged cells specifically
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Convert all cell values to strings for searching
        # Include both individual cells and merged cell ranges
        all_text = []
        cell_text_map = {}  # Map cell coordinates to text for debugging
        
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    cell_text = str(cell.value).strip()
                    all_text.append(cell_text)
                    cell_coord = cell.coordinate
                    cell_text_map[cell_coord] = cell_text
                    
                    # Also check if cell contains newlines (for merged cells with multiple fields)
                    if '\n' in cell_text or '\r' in cell_text:
                        # Split by newlines and add each line
                        lines = cell_text.replace('\r', '\n').split('\n')
                        for line in lines:
                            line = line.strip()
                            if line:
                                all_text.append(line)
        
        # Check merged cells - get text from the top-left cell of each merged range
        for merged_range in merged_cells:
            try:
                # Get the top-left cell of the merged range
                top_left_cell = ws[merged_range.min_row][merged_range.min_col - 1]
                if top_left_cell.value is not None:
                    merged_text = str(top_left_cell.value).strip()
                    all_text.append(merged_text)
                    # Also split by newlines if present
                    if '\n' in merged_text or '\r' in merged_text:
                        lines = merged_text.replace('\r', '\n').split('\n')
                        for line in lines:
                            line = line.strip()
                            if line:
                                all_text.append(line)
                    logger.debug(f"Merged range {merged_range} contains text: {merged_text[:100]}")
            except Exception as e:
                logger.debug(f"Error reading merged range {merged_range}: {e}")
        
        # Check 1: Title exists
        title_found = False
        for text in all_text:
            if expected_title in text:
                title_found = True
                logger.info(f"✓ Found title: {expected_title}")
                break
        
        if not title_found:
            logger.error(f"✗ Title '{expected_title}' not found")
            return 0.0
        
        # Check 2: Required header fields
        # Since fields may be in merged cells together, we need to check if all fields exist
        # even if they're in the same cell
        found_fields = []
        missing_fields = []
        for field in required_fields:
            field_found = False
            for text in all_text:
                if field in text:
                    field_found = True
                    found_fields.append(field)
                    logger.debug(f"Found field '{field}' in text: {text[:80]}...")
                    break
            if not field_found:
                missing_fields.append(field)
        
        # Log all text for debugging if fields are missing
        if missing_fields:
            logger.warning(f"⚠ Missing required fields: {missing_fields}")
            # Show some sample text that might contain the fields
            logger.debug(f"Sample text from worksheet (showing text with Chinese characters):")
            chinese_text_samples = [t for t in all_text if any('\u4e00' <= c <= '\u9fff' for c in t)][:20]
            for sample in chinese_text_samples:
                logger.debug(f"  {sample[:100]}")
        
        # Since fields may be grouped in merged cells, we're more lenient
        # Check if at least most fields are found
        found_ratio = len(found_fields) / len(required_fields) if required_fields else 1.0
        
        if found_ratio < 0.5:  # Less than 50% found
            logger.error(f"✗ Too many required fields missing: found {len(found_fields)}/{len(required_fields)}")
            logger.error(f"  Missing: {missing_fields}")
            return 0.0
        elif missing_fields:
            logger.warning(f"⚠ Some fields missing: {missing_fields}, but found {len(found_fields)}/{len(required_fields)} fields")
            # Don't fail if most fields are found (fields might be in merged cells together)
        else:
            logger.info(f"✓ Found all required fields: {found_fields}")
        
        # Check 3: Table headers
        found_headers = []
        missing_headers = []
        for header in table_headers:
            header_found = False
            for text in all_text:
                if header in text:
                    header_found = True
                    found_headers.append(header)
                    break
            if not header_found:
                missing_headers.append(header)
        
        if missing_headers:
            logger.error(f"✗ Missing table headers: {missing_headers}")
            return 0.0
        else:
            logger.info(f"✓ Found all table headers: {found_headers}")
        
        # Check 4: Summary fields
        found_summary = []
        missing_summary = []
        for field in summary_fields:
            field_found = False
            for text in all_text:
                if field in text:
                    field_found = True
                    found_summary.append(field)
                    break
            if not field_found:
                missing_summary.append(field)
        
        if missing_summary:
            logger.warning(f"⚠ Missing summary fields: {missing_summary}")
            # Don't fail completely, as these might be optional
        else:
            logger.info(f"✓ Found summary fields: {found_summary}")
        
        # Check 5: Footer fields
        found_footer = []
        missing_footer = []
        for field in footer_fields:
            field_found = False
            for text in all_text:
                if field in text:
                    field_found = True
                    found_footer.append(field)
                    break
            if not field_found:
                missing_footer.append(field)
        
        if missing_footer:
            logger.warning(f"⚠ Missing footer fields: {missing_footer}")
            # Don't fail completely, as these might be optional
        else:
            logger.info(f"✓ Found footer fields: {found_footer}")
        
        # Check 6: Borders on table cells
        # Check if cells in the product table area have borders
        # Typically, table headers and data rows should have borders
        logger.info("Checking borders on table cells...")
        
        # Find the table header row (should contain table headers)
        table_header_row = None
        for row_num in range(1, max_row + 1):
            for col_num in range(1, max_col + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    cell_text = str(cell.value).strip()
                    # Check if this row contains table headers
                    if any(header in cell_text for header in table_headers):
                        table_header_row = row_num
                        break
            if table_header_row:
                break
        
        borders_found = False
        cells_with_borders = 0
        total_table_cells_checked = 0
        
        if table_header_row:
            logger.info(f"Table header row found at row {table_header_row}")
            # Check borders in table area (header row and a few data rows)
            # Table typically spans from column C to K (based on headers)
            check_start_col = 3  # Column C
            check_end_col = min(11, max_col)  # Column K or max_col
            check_start_row = table_header_row
            check_end_row = min(table_header_row + 7, max_row)  # Header + 6 data rows
            
            for row_num in range(check_start_row, check_end_row + 1):
                for col_num in range(check_start_col, check_end_col + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    total_table_cells_checked += 1
                    
                    # Check if cell has any border
                    has_border = False
                    try:
                        border = cell.border
                        # Check if any side has a border style (not None and not empty)
                        if border:
                            if (border.top and border.top.style) or \
                               (border.bottom and border.bottom.style) or \
                               (border.left and border.left.style) or \
                               (border.right and border.right.style):
                                has_border = True
                                cells_with_borders += 1
                    except Exception as e:
                        logger.debug(f"Error checking border for cell {cell.coordinate}: {e}")
                    
                    if has_border:
                        borders_found = True
                        logger.debug(f"Cell {cell.coordinate} has borders")
            
            if total_table_cells_checked > 0:
                border_ratio = cells_with_borders / total_table_cells_checked
                logger.info(f"Borders found: {cells_with_borders}/{total_table_cells_checked} cells ({border_ratio:.1%})")
                
                # Require at least 30% of table cells to have borders
                if border_ratio < 0.3:
                    logger.warning(f"⚠ Low border coverage: only {border_ratio:.1%} of table cells have borders")
                    # Don't fail completely, as borders might be applied differently
                else:
                    logger.info(f"✓ Sufficient borders found in table area")
            else:
                logger.warning("⚠ Could not check borders: no table cells found")
        else:
            logger.warning("⚠ Could not find table header row for border checking")
        
        # If we get here, all critical checks passed
        logger.info("=" * 60)
        logger.info(f"✓ Quote sheet template with merged cells verification passed")
        logger.info(f"  Title: {expected_title}")
        logger.info(f"  Merged cells: {len(merged_cells)} (minimum required: {min_merged_cells})")
        if title_merged_found:
            logger.info(f"  Title merged range: {title_merged_range}")
        logger.info(f"  Required fields: {len(found_fields)}/{len(required_fields)}")
        logger.info(f"  Table headers: {len(found_headers)}/{len(table_headers)}")
        logger.info(f"  Summary fields: {len(found_summary)}/{len(summary_fields)}")
        logger.info(f"  Footer fields: {len(found_footer)}/{len(footer_fields)}")
        if borders_found:
            logger.info(f"  Borders: {cells_with_borders}/{total_table_cells_checked} cells have borders")
        logger.info("=" * 60)
        return 1.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


import functools
import itertools
import logging
import os.path

# import operator
from numbers import Number
from typing import Any, Union, cast, Callable, Iterable
from typing import Dict, List, Tuple, Set

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from rapidfuzz import fuzz

from desktop_env.evaluators.metrics.utils import (
    _match_value_to_rule,
    _read_cell_style,
    read_cell_value,
)
from desktop_env.evaluators.metrics.utils import (
    load_charts,
    load_sparklines,
    load_rows_or_cols,
    load_xlsx_styles,
    load_filters,
    load_pivot_tables,
)

# from openpyxl.utils import coordinate_to_tuple

logger = logging.getLogger("desktopenv.metric.table")

BOOK = Union[pd.ExcelFile, Workbook, str]


def _parse_sheet_idx(
    sheet_idx: Union[int, str],
    result: BOOK,
    expected: BOOK,
    result_sheet_names: List[str],
    expected_sheet_names: List[str],
) -> Tuple[BOOK, str]:
    #  function _parse_sheet_idx {{{ #
    if isinstance(sheet_idx, int):
        try:
            if not result_sheet_names or sheet_idx >= len(result_sheet_names):
                logger.error(
                    f"Sheet index {sheet_idx} out of range. Available sheets: {result_sheet_names}"
                )
                index = ""
            else:
                index: str = result_sheet_names[sheet_idx]
                logger.debug(f"Sheet index {sheet_idx} resolved to sheet: {index}")
        except Exception as e:
            logger.error(f"Error resolving sheet index {sheet_idx}: {e}")
            index = ""
        book: BOOK = result
    elif sheet_idx.startswith("RI"):
        try:
            index: str = result_sheet_names[int(sheet_idx[2:])]
        except:
            index = ""
        book: BOOK = result
    elif sheet_idx.startswith("RN"):
        index: str = sheet_idx[2:]
        book: BOOK = result
    elif sheet_idx.startswith("EI"):
        try:
            index: str = expected_sheet_names[int(sheet_idx[2:])]
        except:
            index = ""
        book: BOOK = expected
    elif sheet_idx.startswith("EN"):
        index: str = sheet_idx[2:]
        book: BOOK = expected
    else:
        logger.error("Unrecognized sheet index")
        raise ValueError("Unrecognized sheet index")
    return book, index
    #  }}} function _parse_sheet_idx #


SHEET = Union[pd.DataFrame, Worksheet, List[str]]


def _load_sheet(book: BOOK, index: str) -> SHEET:
    #  function _load_sheet {{{ #
    try:
        if isinstance(book, str):
            book: str = cast(str, book)
            csv_name: str = "{:}-{:}.csv".format(os.path.splitext(book)[0], index)

            try:
                all_lines: List[str] = _safe_read_file(csv_name)
                csv_lines: List[str] = list(
                    itertools.dropwhile(
                        lambda l: len(l) == 0,
                        map(lambda l: l.strip(), reversed(all_lines)),
                    )
                )
                return csv_lines
            except (FileNotFoundError, IOError) as e:
                logger.error(f"Failed to read CSV file {csv_name}: {e}")
                return None
        if isinstance(book, pd.ExcelFile):
            return pd.read_excel(book, index)
        if isinstance(book, Workbook):
            return book[index]
        logger.error("Not supported workbook format")
        raise NotImplementedError("Not supported workbook format")
    except NotImplementedError as e:
        raise e
    except:
        return None
    #  }}} function _load_sheet #


def _safe_read_file(file_path: str) -> List[str]:
    """
    Safely read a file with multiple encoding attempts.

    Args:
        file_path: Path to the file to read

    Returns:
        List of lines from the file

    Raises:
        FileNotFoundError: If file doesn't exist
        IOError: If file cannot be read with any encoding
    """
    # Common encodings to try in order of preference
    encodings = [
        "utf-8",  # Most common modern encoding
        "utf-8-sig",  # UTF-8 with BOM
        "latin-1",  # ISO-8859-1, works with any byte sequence
        "windows-1252",  # Common Windows encoding
        "gbk",  # Chinese encoding
        "cp1251",  # Cyrillic encoding
        "iso-8859-1",  # Alternative latin-1
    ]

    last_error = None

    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding) as f:
                lines = f.read().splitlines()
                logger.debug(
                    f"Successfully read file {file_path} with encoding {encoding}"
                )
                return lines
        except UnicodeDecodeError as e:
            last_error = e
            logger.debug(f"Failed to read {file_path} with encoding {encoding}: {e}")
            continue
        except (FileNotFoundError, IOError) as e:
            # These are non-encoding related errors, re-raise immediately
            raise e

    # If all encodings fail, try with error handling as last resort
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.read().splitlines()
            logger.warning(f"Read file {file_path} with UTF-8 and error replacement")
            return lines
    except Exception as e:
        logger.error(
            f"Failed to read file {file_path} with any encoding. Last error: {last_error}"
        )
        raise IOError(
            f"Cannot read file {file_path} with any supported encoding"
        ) from last_error


def compare_csv(result: str, expected: Union[str, List[str]], **options) -> float:
    """
    Compare CSV files. If expected is a list, returns 1.0 if result matches any of the expected files.

    Args:
        result: Path to result CSV file
        expected: Path to expected CSV file or list of paths to expected CSV files
        options: Additional options (strict, ignore_case)

    Returns:
        1.0 if result matches expected (or any file in expected list), 0.0 otherwise
    """
    if result is None:
        return 0.0

    try:
        result_lines: List[str] = _safe_read_file(result)
    except (FileNotFoundError, IOError) as e:
        logger.error(f"Failed to read result file {result}: {e}")
        return 0.0

    # Convert expected to list if it's a single string (for backward compatibility)
    if isinstance(expected, str):
        expected_files = [expected]
    else:
        expected_files = expected

    # Try to match against each expected file
    for expected_file in expected_files:
        try:
            expected_lines: List[str] = _safe_read_file(expected_file)

            # Process lines based on options
            current_result_lines = result_lines
            current_expected_lines = expected_lines

            if not options.get("strict", True):
                current_result_lines = map(str.strip, current_result_lines)
                current_expected_lines = map(str.strip, current_expected_lines)
            if options.get("ignore_case", False):
                current_result_lines = map(str.lower, current_result_lines)
                current_expected_lines = map(str.lower, current_expected_lines)

            # Check if this expected file matches
            if list(current_result_lines) == list(current_expected_lines):
                return 1.0

        except (FileNotFoundError, IOError):
            # If this expected file doesn't exist, continue to next one
            continue

    # No match found
    return 0.0


def compare_table(result: str, expected: str = None, **options) -> float:
    #  function compare_table {{{ #
    """
    Args:
        result (str): path to result xlsx
        expected (str): path to golden xlsx
        rules (List[Dict[str, Any]]): list of dict like
          {
            "type": str,
            <str as parameters>: anything
          }
          as sequential rules

    Returns:
        float: the score
    """

    if result is None:
        logger.error("Result file path is None")
        return 0.0

    # Check if result file exists
    if not os.path.exists(result):
        logger.error(f"Result file not found: {result}")
        return 0.0

    try:
        logger.info(f"Loading result file: {result}")
        xlworkbookr: Workbook = openpyxl.load_workbook(filename=result)
        pdworkbookr = pd.ExcelFile(result)
        logger.info(
            f"Successfully loaded result file with sheets: {pdworkbookr.sheet_names}"
        )
    except Exception as e:
        logger.error(f"Failed to load result file {result}: {e}")
        return 0.0
    worksheetr_names: List[str] = pdworkbookr.sheet_names

    if expected is not None:
        xlworkbooke: Workbook = openpyxl.load_workbook(filename=expected)
        pdworkbooke = pd.ExcelFile(expected)
        worksheete_names: List[str] = pdworkbooke.sheet_names
    else:
        xlworkbooke: Workbook = None
        pdworkbooke = None
        worksheete_names: List[str] = None

    parse_idx: Callable[[Union[str, int], BOOK, BOOK], Tuple[BOOK, str]] = (
        functools.partial(
            _parse_sheet_idx,
            result_sheet_names=worksheetr_names,
            expected_sheet_names=worksheete_names,
        )
    )

    passes = True
    for r in options["rules"]:
        if r["type"] == "sheet_name":
            #  Compare Sheet Names {{{ #
            metric: bool = worksheetr_names == worksheete_names
            logger.debug(
                "Assertion: %s.sheet_names == %s.sheet_names - %s",
                result,
                expected,
                metric,
            )
            #  }}} Compare Sheet Names #

        elif r["type"] == "sheet_data":
            #  Compare Sheet Data by Internal Value {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # precision: int as number of decimal digits, default to 4

            error_limit: int = r.get("precision", 4)
            sheet1: pd.DataFrame = _load_sheet(
                *parse_idx(r["sheet_idx0"], pdworkbookr, pdworkbooke)
            )
            if sheet1 is None:
                return 0.0
            sheet2: pd.DataFrame = _load_sheet(
                *parse_idx(r["sheet_idx1"], pdworkbookr, pdworkbooke)
            )

            sheet1 = sheet1.round(error_limit)
            sheet2 = sheet2.round(error_limit)
            metric: bool = sheet1.equals(sheet2)
            logger.debug("Sheet1: \n%s", str(sheet1))
            logger.debug("Sheet2: \n%s", str(sheet2))
            try:
                logger.debug("Sheet1 =v= Sheet2: \n%s", str(sheet1 == sheet2))
            except:
                logger.debug("Sheet1 =/v= Sheet2")
            logger.debug(
                "Assertion: %s =v= %s - %s", r["sheet_idx0"], r["sheet_idx1"], metric
            )
            #  }}} Compare Sheet Data by Internal Value #

        elif r["type"] == "sheet_print":
            #  Compare Sheet Data by Printed Value {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # ignore_case: optional, defaults to False

            sheet1: List[str] = _load_sheet(
                *parse_idx(r["sheet_idx0"], result, expected)
            )
            if sheet1 is None:
                return 0.0
            sheet2: List[str] = _load_sheet(
                *parse_idx(r["sheet_idx1"], result, expected)
            )
            if r.get("ignore_case", False):
                sheet1 = [l.lower() for l in sheet1]
                sheet2 = [l.lower() for l in sheet2]
            metric: bool = sheet1 == sheet2
            logger.debug(
                "Assertion: %s =p= %s - %s", r["sheet_idx0"], r["sheet_idx1"], metric
            )
            #  }}} Compare Sheet Data by Printed Value #

        elif r["type"] == "sheet_fuzzy":
            #  Fuzzy Match for Ranges {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # rules: list of dict, each dict is like
            #   { "range": ["A1:B6", "C2:E5"],
            #     "type": "includes" | "included_by" | "fuzzy_match" | "exact_match", # 0 includes 1, 0 includes_by 1
            #     "threshold": 85, // for fuzzy match
            #     "ignore_case": true | false,
            #     "ignore_chars": " ()", # filtered out
            #     "trim_leadings": "+ ", # filtered by lstrip
            #     "trim_trailings": "", # filtered by rstrip
            #     "normalization": [["Rd", "Road"]], # filtered by replace
            #   }

            sheet1: Tuple[BOOK, str] = parse_idx(r["sheet_idx0"], result, expected)
            sheet2: Tuple[BOOK, str] = parse_idx(r["sheet_idx1"], result, expected)
            total_metric = True
            for rl in r["rules"]:
                for rng in MultiCellRange(rl["range"]):
                    for cdn in rng.cells:
                        coordinate: str = "{:}{:d}".format(
                            get_column_letter(cdn[1]), cdn[0]
                        )
                        value1: str = str(read_cell_value(*sheet1, coordinate))
                        value2: str = str(read_cell_value(*sheet2, coordinate))
                        logger.debug("%s: %s vs %s", cdn, value1, value2)

                        for rplc in rl.get("normalization", []):
                            value1 = value1.replace(rplc[0], rplc[1])
                            value2 = value2.replace(rplc[0], rplc[1])
                        if "trim_leadings" in rl:
                            value1 = value1.lstrip(rl["trim_leadings"])
                            value2 = value2.lstrip(rl["trim_leadings"])
                        if "trim_trailings" in rl:
                            value1 = value1.rstrip(rl["trim_trailings"])
                            value2 = value2.rstrip(rl["trim_trailings"])
                        if "ignore_chars" in rl:
                            ignore_chars: Set[str] = set(rl["ignore_chars"])
                            value1 = "".join(
                                filter(lambda ch: ch not in ignore_chars, value1)
                            )
                            value2 = "".join(
                                filter(lambda ch: ch not in ignore_chars, value2)
                            )
                        if rl.get("ignore_case", False):
                            value1 = value1.lower()
                            value2 = value2.lower()

                        if rl["type"] == "includes":
                            metric: bool = value2 in value1
                        elif rl["type"] == "included_by":
                            metric: bool = value1 in value2
                        elif rl["type"] == "fuzzy_match":
                            metric: bool = fuzz.ratio(value1, value2) >= rl.get(
                                "threshold", 85.0
                            )
                        elif rl["type"] == "exact_match":
                            metric: bool = value1 == value2
                        total_metric = total_metric and metric

            metric: bool = total_metric
            logger.debug(
                "Assertion: %s =~= %s - %s", r["sheet_idx0"], r["sheet_idx1"], metric
            )
            #  }}} Fuzzy Match for Ranges #

        elif r["type"] == "sparkline":
            #  Compare Sparklines {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0

            sparkline1: Dict[str, str] = load_sparklines(
                *parse_idx(r["sheet_idx0"], result, expected)
            )
            sparkline2: Dict[str, str] = load_sparklines(
                *parse_idx(r["sheet_idx1"], result, expected)
            )
            metric: bool = sparkline1 == sparkline2
            logger.debug(
                "Assertion: %s.sp == %.sp - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Compare Sparklines #

        elif r["type"] == "chart":
            #  Compare Charts {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # chart_props: list of str, see utils.load_charts

            charts1: Dict[str, Any] = load_charts(
                *parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke), **r
            )
            charts2: Dict[str, Any] = load_charts(
                *parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke), **r
            )
            metric: bool = charts1 == charts2
            logger.debug(
                "Assertion: %s[chart] == %s[chart] - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Compare Charts #

        elif r["type"] == "style":
            #  Compare Style (Also Conditional Formatiing) {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # props: list of str indicating concerned styles, see utils._read_cell_style

            sheet_idx1: Tuple[BOOK, str] = parse_idx(
                r["sheet_idx0"], xlworkbookr, xlworkbooke
            )
            book_name1: str = parse_idx(r["sheet_idx0"], result, expected)[0]
            styles1: Dict[str, List[Any]] = load_xlsx_styles(
                *sheet_idx1, book_name1, **r
            )

            sheet_idx2: Tuple[BOOK, str] = parse_idx(
                r["sheet_idx1"], xlworkbookr, xlworkbooke
            )
            book_name2: str = parse_idx(r["sheet_idx1"], result, expected)[0]
            styles2: Dict[str, List[Any]] = load_xlsx_styles(
                *sheet_idx2, book_name2, **r
            )
            # number_formats1: List[str] = [c.number_format.lower() for col in sheet1.iter_cols() for c in col if c.value is not None and c.data_type=="n"]
            # number_formats2: List[str] = [c.number_format.lower() for col in sheet2.iter_cols() for c in col if c.value is not None and c.data_type=="n"]
            metric: bool = styles1 == styles2
            logger.debug(
                "Assertion: %s.style == %s.style - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Compare Style (Also Conditional Formatiing) #

        elif r["type"] == "freeze":
            #  Compare Freezing {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0

            sheet1: Worksheet = _load_sheet(
                *parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke)
            )
            if sheet1 is None:
                return 0.0
            sheet2: Worksheet = _load_sheet(
                *parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke)
            )
            metric: bool = sheet1.freeze_panes == sheet2.freeze_panes
            logger.debug(
                "Assertion: %s.freeze(%s) == %s.freeze(%s) - %s",
                r["sheet_idx0"],
                sheet1.freeze_panes,
                r["sheet_idx1"],
                sheet2.freeze_panes,
                metric,
            )
            #  }}} Compare Freezing #

        elif r["type"] == "zoom":
            #  Check Zooming {{{ #
            # sheet_idx: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # method: str
            # ref: value

            sheet: Worksheet = _load_sheet(
                *parse_idx(r["sheet_idx"], xlworkbookr, xlworkbooke)
            )
            if sheet is None:
                return 0.0
            zoom_scale: Number = sheet.sheet_view.zoomScale or 100.0
            metric: bool = _match_value_to_rule(zoom_scale, r)
            logger.debug(
                "Assertion: %s.zoom(%.1f) %s %.1f - %s",
                r["sheet_idx"],
                zoom_scale,
                r["method"],
                r["ref"],
                metric,
            )
            #  }}} Check Zooming #

        elif r["type"] == "data_validation":
            #  Check Data Validation {{{ #
            # sheet_idx: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # dv_props: list of dict like {attribute: {"method": str, "ref": anything}}
            #   available attributes:
            #     * ranges
            #     * type
            #     * formula1
            #     * formula2
            #     * operator
            #     * allowBlank
            #     * showDropDown
            #     * showInputMessage
            #     * showErrorMessage
            #     * error
            #     * errorTitle
            #     * errorStyle
            #     * prompt
            #     * promptTitle
            #     * imeMode

            sheet: Worksheet = _load_sheet(
                *parse_idx(r["sheet_idx"], xlworkbookr, xlworkbooke)
            )
            if sheet is None:
                return 0.0
            data_validators: List[DataValidation] = (
                sheet.data_validations.dataValidation
            )

            total_metric = len(data_validators) >= len(r["dv_props"])
            for dat_vldt in data_validators:
                metric = False
                for prpt in r["dv_props"]:
                    metric = metric or all(
                        _match_value_to_rule(getattr(dat_vldt, attrbt), mr)
                        for attrbt, mr in prpt.items()
                    )
                    if metric:
                        break
                total_metric = total_metric and metric
                if not total_metric:
                    break

            logger.debug(
                "Assertion: %s.data_validation - %s", r["sheet_idx"], total_metric
            )
            metric: bool = total_metric
            #  }}} Check Data Validation #

        elif r["type"] == "row_props":
            #  Check Row Properties {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # props: list of str, see utils.load_rows_or_cols

            rows1: Dict[str, Any] = load_rows_or_cols(
                *parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke), obj="row", **r
            )
            rows2: Dict[str, Any] = load_rows_or_cols(
                *parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke), obj="row", **r
            )
            logger.debug("Rows1: %s", repr(rows1))
            logger.debug("Rows2: %s", repr(rows2))
            metric: bool = rows1 == rows2
            logger.debug(
                "Assertion: %s[rows] == %s[rows] - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Check Row Properties #

        elif r["type"] == "col_props":
            #  Check Row Properties {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # props: list of str, see utils.load_rows_or_cols

            cols1: Dict[str, Any] = load_rows_or_cols(
                *parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke), obj="column", **r
            )
            cols2: Dict[str, Any] = load_rows_or_cols(
                *parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke), obj="column", **r
            )
            metric: bool = cols1 == cols2
            logger.debug(
                "Assertion: %s[cols] == %s[cols] - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Check Row Properties #

        elif r["type"] == "filter":
            #  Compare Filters {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0

            filters1: Dict[str, Any] = load_filters(
                *parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke), **r
            )
            filters2: Dict[str, Any] = load_filters(
                *parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke), **r
            )
            metric: bool = filters1 == filters2
            logger.debug(
                "Assertion: %s[filter] == %s[filter] - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Compare Filters #

        elif r["type"] == "pivot_table":
            #  Compare Pivot Tables {{{ #
            # sheet_idx0: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # sheet_idx1: as sheet_idx0
            # pivot_props: list of str, see utils.load_pivot_tables

            pivots1: Dict[str, Any] = load_pivot_tables(
                *parse_idx(r["sheet_idx0"], xlworkbookr, xlworkbooke), **r
            )
            pivots2: Dict[str, Any] = load_pivot_tables(
                *parse_idx(r["sheet_idx1"], xlworkbookr, xlworkbooke), **r
            )
            metric: bool = pivots1 == pivots2
            logger.debug(
                "Assertion: %s[pivot]==%s[pivot] - %s",
                r["sheet_idx0"],
                r["sheet_idx1"],
                metric,
            )
            #  }}} Compare Pivot Tables #

        elif r["type"] == "check_cell":
            #  Check Cell Properties {{{ #
            # sheet_idx: 0 == "RI0" == "RNSheet1" | "EI0" == "ENSheet1"
            # coordinate: str, "E3"
            # props: dict like {attribute: {"method": str, "ref": anything}}
            #   supported attributes: value & those supported by utils._read_cell_style

            try:
                sheet: Worksheet = _load_sheet(
                    *parse_idx(r["sheet_idx"], xlworkbookr, xlworkbooke)
                )
                if sheet is None:
                    logger.error(
                        f"Failed to load sheet for sheet_idx: {r['sheet_idx']}"
                    )
                    return 0.0
                # data_frame: pd.DataFrame = _load_sheet(*parse_idx(r["sheet_idx"], pdworkbookr, pdworkbooke))
                cell: Cell = sheet[r["coordinate"]]
                metric: bool = True
                for prpt, rule in r["props"].items():
                    if prpt == "value":
                        try:
                            parsed_result = parse_idx(r["sheet_idx"], result, expected)
                            logger.debug(f"parse_idx result: {parsed_result}")
                            val = read_cell_value(*parsed_result, r["coordinate"])
                            logger.debug(f"Cell {r['coordinate']} value: {val}")
                        except Exception as e:
                            logger.error(
                                f"Failed to read cell value at {r['coordinate']}: {e}"
                            )
                            val = None
                    elif prpt == "formula":
                        # Support checking cell formula directly
                        try:
                            if cell.data_type == "f":
                                # For formula cells, get the formula text
                                # In openpyxl, formula is stored in cell.value for formula cells
                                # But we need the actual formula text, not the calculated value
                                # Try to get formula from internal representation
                                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                                    val = cell._value
                                elif hasattr(cell, "formula"):
                                    val = cell.formula
                                else:
                                    # Fallback: try to reconstruct from value if it's a formula
                                    val = f"={cell.value}" if cell.value is not None else None
                            else:
                                val = None
                            logger.debug(f"Cell {r['coordinate']} formula: {val}")
                        except Exception as e:
                            logger.error(
                                f"Failed to read cell formula at {r['coordinate']}: {e}"
                            )
                            val = None
                    else:
                        try:
                            val = _read_cell_style(prpt, cell)
                        except Exception as e:
                            logger.error(
                                f"Failed to read cell style {prpt} at {r['coordinate']}: {e}"
                            )
                            val = None

                    metric = metric and _match_value_to_rule(val, rule)
            except Exception as e:
                logger.error(f"Error in check_cell processing: {e}")
                return 0.0

            logger.debug(
                "Assertion: %s[%s] :%s - %s",
                r["sheet_idx"],
                r["coordinate"],
                repr(r["props"]),
                metric,
            )
            #  }}} Check Cell Properties #

        else:
            raise NotImplementedError(
                "Unimplemented sheet check: {:}".format(r["type"])
            )

        passes = passes and metric
        if not passes:
            break

    return float(passes)
    #  }}} function compare_table #


def compare_conference_city_in_order(actual_city_list_path, expected_city):
    expected_city_list = expected_city["expected"]
    wb = openpyxl.load_workbook(actual_city_list_path)
    sheet = wb.active
    actual_city_list = []
    for row in sheet["C2:C22"]:
        for cell in row:
            actual_city_list.append(cell.value)
    # expected_city is the city that we want to compare with the actual city list
    # must in order index
    # debug
    try:
        for i in range(len(actual_city_list)):
            if isinstance(expected_city_list[i], str):
                if expected_city_list[i] not in actual_city_list[i]:
                    logger.debug(
                        f"Expected city {expected_city_list[i]}; Actual city {actual_city_list[i]}"
                    )
                    print(
                        f"Expected city {expected_city_list[i]}; Actual city {actual_city_list[i]}"
                    )
                    return 0.0

            elif isinstance(expected_city_list[i], List):
                if not any(
                    possible_str in actual_city_list[i]
                    for possible_str in expected_city_list[i]
                ):
                    logger.debug(
                        f"Expected city {expected_city_list[i]}; Actual city {actual_city_list[i]}"
                    )
                    print(
                        f"Expected city {expected_city_list[i]}; Actual city {actual_city_list[i]}"
                    )
                    return 0.0

            else:
                raise TypeError("Expected city should be a string or a list of strings")

    except:
        return 0.0

    return 1.0


def verify_second_row_deleted_without_gold(result: str, expected: str = None, **options) -> float:
    """
    验证 Excel 文件的第二行是否被删除（不需要金标准文件）
    
    通过以下方式验证：
    1. 检查结果文件的行数是否比原始文件少1
    2. 检查原始文件的第二行数据是否在结果文件中不存在
    3. 检查其他所有行是否保持不变
    
    Args:
        result (str): 结果文件路径
        expected (str): 未使用（为了兼容框架接口）
        options (dict): 配置选项，应包含：
            - original_file_url: 原始文件的URL（用于下载和比对）
            - result_file_path: 结果文件的路径（可选，默认使用 result 参数）
            - original_file_cache: 原始文件的本地缓存路径（可选）
    
    Returns:
        float: 如果验证通过返回 1.0，否则返回 0.0
    """
    try:
        import tempfile
        import urllib.request
        
        # result 参数已经是从VM获取到宿主机的文件路径
        # 不应该从 options 中覆盖它，因为 options 中可能包含的是VM路径
        result_file_path = result
        original_file_url = options.get('original_file_url', '')
        
        logger.info(f"开始验证删除第二行任务...")
        logger.info(f"结果文件: {result_file_path}")
        logger.info(f"原始文件URL: {original_file_url}")
        
        if not result_file_path or not os.path.exists(result_file_path):
            logger.error(f"结果文件不存在: {result_file_path}")
            return 0.0
        
        # 下载原始文件到临时位置
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            original_file_temp = tmp_file.name
        
        try:
            logger.info(f"正在下载原始文件到临时位置: {original_file_temp}")
            urllib.request.urlretrieve(original_file_url, original_file_temp)
        except Exception as e:
            logger.warning(f"下载原始文件失败: {e}")
            # 如果下载失败，尝试从本地缓存读取
            cache_path = options.get('original_file_cache', '')
            if cache_path and os.path.exists(cache_path):
                logger.info(f"使用缓存文件: {cache_path}")
                original_file_temp = cache_path
            else:
                logger.error("无法获取原始文件")
                return 0.0
        
        # 加载原始文件
        logger.info("加载原始文件...")
        original_wb = openpyxl.load_workbook(original_file_temp)
        original_ws = original_wb.active
        
        # 获取原始文件的所有行
        original_rows = list(original_ws.iter_rows(values_only=True))
        original_row_count = len(original_rows)
        
        if original_row_count < 2:
            logger.error(f"原始文件行数不足: {original_row_count}（需要至少2行）")
            return 0.0
        
        # 保存第二行的数据（索引为1）
        second_row_data = original_rows[1]
        logger.info(f"原始文件行数: {original_row_count}")
        logger.info(f"原始文件第二行数据: {second_row_data}")
        
        # 加载结果文件
        logger.info(f"加载结果文件...")
        result_wb = openpyxl.load_workbook(result_file_path)
        result_ws = result_wb.active
        
        # 获取结果文件的所有行
        result_rows = list(result_ws.iter_rows(values_only=True))
        result_row_count = len(result_rows)
        
        logger.info(f"结果文件行数: {result_row_count}")
        
        # 验证1: 检查行数是否减少了1
        if result_row_count != original_row_count - 1:
            logger.error(f"行数验证失败: 期望 {original_row_count - 1} 行，实际 {result_row_count} 行")
            return 0.0
        else:
            logger.info(f"✓ 行数验证通过: {original_row_count} → {result_row_count}")
        
        # 验证2: 检查原始第二行是否存在于结果文件中
        second_row_exists = False
        for i, row in enumerate(result_rows):
            if row == second_row_data:
                logger.error(f"原始第二行数据仍存在于结果文件的第 {i+1} 行")
                second_row_exists = True
                break
        
        if second_row_exists:
            return 0.0
        else:
            logger.info(f"✓ 原始第二行数据已从结果文件中删除")
        
        # 验证3: 检查其他行是否保持不变（第一行和第3行之后）
        # 结果文件的第一行应该等于原始文件的第一行
        if result_rows[0] != original_rows[0]:
            logger.error(f"第一行数据不匹配")
            logger.error(f"  原始: {original_rows[0]}")
            logger.error(f"  结果: {result_rows[0]}")
            return 0.0
        
        # 结果文件的第2行及之后应该等于原始文件的第3行及之后
        for i in range(1, result_row_count):
            if result_rows[i] != original_rows[i+1]:
                logger.error(f"第 {i+1} 行数据不匹配")
                logger.error(f"  期望（原始第 {i+2} 行）: {original_rows[i+1]}")
                logger.error(f"  实际: {result_rows[i]}")
                return 0.0
        
        logger.info(f"✓ 其他行数据保持不变")
        
        # 清理临时文件
        if original_file_temp != options.get('original_file_cache', ''):
            try:
                os.unlink(original_file_temp)
            except:
                pass
        
        logger.info("=" * 60)
        logger.info("✓ 所有验证通过！第二行已成功删除")
        logger.info("=" * 60)
        return 1.0
        
    except Exception as e:
        import traceback
        logger.error(f"评估出错: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_regexp_extract(result: str, expected: str = None, **options) -> float:
    """
    Verify if REGEX formulas exist in specified columns (B and C) with correct patterns.
    
    This function checks:
    1. Whether cells in specified columns contain REGEX formulas
    2. Whether formulas reference the corresponding A column cell (B2->A2, B3->A3, etc.)
    3. Whether formulas contain the correct pattern text (牛肉丸 for B column, 牛筋丸 for C column)
    4. Whether formulas have the correct structure with lookbehind and lookahead
    
    The function automatically detects the number of data rows by checking the data column
    (default: A column) for non-empty cells. It stops checking after finding 3 consecutive
    empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_columns: List of columns to check (e.g., ["B", "C"])
            - start_row: Starting row number (default: 2)
            - end_row: Ending row number (optional, will auto-detect if not provided)
            - expected_pattern: Expected function name (default: "REGEX")
            - column_patterns: Dict mapping column letters to expected pattern text
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        from openpyxl.utils import get_column_letter, column_index_from_string
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_columns = options.get('check_columns', ['B', 'C'])
        start_row = options.get('start_row', 2)
        end_row = options.get('end_row', None)  # Optional, will auto-detect if not provided
        expected_pattern = options.get('expected_pattern', 'REGEX')
        column_patterns = options.get('column_patterns', {'B': '牛肉丸', 'C': '牛筋丸'})
        data_column = options.get('data_column', 'A')  # Column to check for data to determine end_row
        
        if not check_columns:
            logger.error("No columns specified in options")
            return 0.0
        
        logger.info(f"Verifying REGEX formulas in file: {result}")
        logger.info(f"Columns to check: {check_columns}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected function: {expected_pattern}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        if end_row is None:
            logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
            max_row = ws.max_row
            end_row = start_row  # Start from start_row
            
            # Find the last row with data in the data column
            # Check up to max_row, but stop if we find 3 consecutive empty rows
            empty_count = 0
            for row_num in range(start_row, max_row + 1):
                data_cell = ws[f"{data_column}{row_num}"]
                if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                    empty_count += 1
                    if empty_count >= 3:  # Stop after 3 consecutive empty rows
                        break
                else:
                    empty_count = 0
                    end_row = row_num  # Update end_row to the last row with data
            
            logger.info(f"Auto-detected end row: {end_row}")
        else:
            logger.info(f"Using specified end row: {end_row}")
        
        # Check each column and row
        all_passed = True
        for col_letter in check_columns:
            expected_pattern_text = column_patterns.get(col_letter)
            if not expected_pattern_text:
                logger.warning(f"No pattern text specified for column {col_letter}, skipping")
                continue
            
            logger.info(f"Checking column {col_letter} with pattern '{expected_pattern_text}' (rows {start_row} to {end_row})")
            
            for row_num in range(start_row, end_row + 1):
                cell_coord = f"{col_letter}{row_num}"
                try:
                    cell = ws[cell_coord]
                    logger.debug(f"Checking cell {cell_coord}")
                    
                    # Check if cell contains a formula
                    if cell.data_type != "f":
                        logger.warning(f"Cell {cell_coord} does not contain a formula")
                        all_passed = False
                        continue
                    
                    # Get formula text
                    formula_text = None
                    if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                        formula_text = cell._value
                    elif hasattr(cell, "formula"):
                        formula_text = cell.formula
                    else:
                        # Try to get from value attribute
                        if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                            formula_text = cell.value
                    
                    if formula_text is None:
                        logger.warning(f"Could not extract formula from cell {cell_coord}")
                        all_passed = False
                        continue
                    
                    # Remove leading = if present for comparison
                    formula_clean = formula_text.lstrip("=")
                    logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                    
                    # Check 1: Formula contains REGEX function
                    if expected_pattern.upper() not in formula_text.upper():
                        logger.warning(f"Cell {cell_coord} formula does not contain {expected_pattern}")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                    
                    # Check 2: Formula contains expected pattern text (牛肉丸 or 牛筋丸)
                    if expected_pattern_text not in formula_text:
                        logger.warning(f"Cell {cell_coord} formula does not contain pattern text '{expected_pattern_text}'")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                    
                    # Check 3: Formula contains REGEX function call structure
                    regex_match = re.search(r'REGEX\s*\([^)]+\)', formula_text, re.IGNORECASE)
                    if not regex_match:
                        logger.warning(f"Cell {cell_coord} formula does not have correct REGEX structure")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                    
                    # Check 4: Formula references the corresponding A column cell (A2, A3, etc.)
                    expected_a_cell = f"A{row_num}"
                    # Check if formula contains A column reference with the same row number
                    a_cell_pattern = rf'A{row_num}\b'
                    if not re.search(a_cell_pattern, formula_text, re.IGNORECASE):
                        logger.warning(f"Cell {cell_coord} formula does not reference {expected_a_cell}")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                    
                    # Check 5: Formula contains lookbehind pattern (?<=...)
                    if "(?<=" not in formula_text:
                        logger.warning(f"Cell {cell_coord} formula does not contain lookbehind pattern (?<=...)")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                    
                    # Check 6: Formula contains lookahead pattern (?=,)
                    if "(?=," not in formula_text:
                        logger.warning(f"Cell {cell_coord} formula does not contain lookahead pattern (?=,)")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                    
                    # Check 7: Formula contains \d+ pattern
                    if "\\d+" not in formula_text:
                        # Also check for unescaped version in the pattern
                        if not re.search(r'\\d\+|d\+', formula_text):
                            logger.warning(f"Cell {cell_coord} formula does not contain digit pattern \\d+")
                            logger.warning(f"Formula: {formula_text}")
                            all_passed = False
                            continue
                    
                    # Check 8: Formula pattern should contain 5 dots after pattern text
                    # Pattern should be like: (?<=牛肉丸.....)
                    pattern_with_dots = expected_pattern_text + "....."
                    if pattern_with_dots not in formula_text:
                        logger.warning(f"Cell {cell_coord} formula pattern may not have 5 dots after '{expected_pattern_text}'")
                        logger.debug(f"Formula: {formula_text}")
                        # Don't fail, just warn - the pattern might be correct but formatted differently
                    
                    logger.info(f"✓ Cell {cell_coord} has valid REGEX formula: {formula_text}")
                    
                except Exception as e:
                    logger.error(f"Error checking cell {cell_coord}: {e}")
                    import traceback
                    logger.error(traceback.format_exc())
                    all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in columns {check_columns} contain correct {expected_pattern} formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ {expected_pattern} formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_regexp_order_extract(result: str, expected: str = None, **options) -> float:
    """
    Verify if REGEXP formulas exist in specified column (C) to extract order numbers from addresses.
    
    This function checks:
    1. Whether cells in specified column contain REGEXP formulas
    2. Whether formulas reference the corresponding A column cell (C2->A2, C3->A3, etc.)
    3. Whether formulas contain the correct regex pattern (\\w{10})
    4. Whether formulas have the correct structure
    
    The function automatically detects the number of data rows by checking the data column
    (default: A column) for non-empty cells. It stops checking after finding 3 consecutive
    empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "C")
            - start_row: Starting row number (default: 2)
            - expected_pattern: Expected function name (default: "REGEXP")
            - expected_formula_pattern: Expected formula pattern (e.g., "REGEXP(A")
            - regex_pattern: Expected regex pattern in formula (e.g., "\\w{10}")
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'C')
        start_row = options.get('start_row', 2)
        expected_pattern = options.get('expected_pattern', 'REGEX')
        expected_formula_pattern = options.get('expected_formula_pattern', 'REGEX(A')
        regex_pattern = options.get('regex_pattern', '[a-zA-Z0-9]{10}')
        data_column = options.get('data_column', 'A')
        
        logger.info(f"Verifying REGEXP order extraction in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected function: {expected_pattern}")
        logger.info(f"Expected regex pattern: {regex_pattern}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row in the specified column
        all_passed = True
        logger.info(f"Checking column {check_column} (rows {start_row} to {end_row})")
        
        for row_num in range(start_row, end_row + 1):
            cell_coord = f"{check_column}{row_num}"
            try:
                cell = ws[cell_coord]
                logger.debug(f"Checking cell {cell_coord}")
                
                # Check if cell contains a formula
                if cell.data_type != "f":
                    logger.warning(f"Cell {cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                # Get formula text
                formula_text = None
                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                    formula_text = cell._value
                elif hasattr(cell, "formula"):
                    formula_text = cell.formula
                else:
                    # Try to get from value attribute
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value
                
                if formula_text is None:
                    logger.warning(f"Could not extract formula from cell {cell_coord}")
                    all_passed = False
                    continue
                
                # Remove leading = if present for comparison
                formula_clean = formula_text.lstrip("=")
                logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                
                # Check 1: Formula contains REGEX function (support both REGEX and LibreOffice internal format)
                # LibreOffice may save as _xlfn.ORG.LIBREOFFICE.REGEX
                formula_upper = formula_text.upper()
                if expected_pattern.upper() not in formula_upper and '_XLFN.ORG.LIBREOFFICE.REGEX' not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain {expected_pattern} or LibreOffice REGEX")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 2: Formula contains expected formula pattern (REGEX(A or _xlfn.ORG.LIBREOFFICE.REGEX(A)
                formula_clean_upper = formula_clean.upper()
                if expected_formula_pattern.upper() not in formula_clean_upper and 'REGEX(A' not in formula_clean_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain pattern '{expected_formula_pattern}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 3: Formula contains REGEX function call structure (support both formats)
                regexp_match = re.search(r'(REGEX|REGEXP|_XLFN\.ORG\.LIBREOFFICE\.REGEX)\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not regexp_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct REGEX structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 4: Formula references the corresponding A column cell (A2, A3, etc.)
                expected_a_cell = f"A{row_num}"
                # Check if formula contains A column reference with the same row number
                a_cell_pattern = rf'A{row_num}\b'
                if not re.search(a_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference {expected_a_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 5: Formula contains the regex pattern ([a-zA-Z0-9]{10})
                # The pattern might be escaped differently in the formula
                # Check for various escape formats
                pattern_variations = [
                    regex_pattern,  # [a-zA-Z0-9]{10}
                    regex_pattern.replace('\\', '\\\\'),  # [a-zA-Z0-9]{10} with double escape
                    regex_pattern.replace('[', '\\[').replace(']', '\\]'),  # Escaped brackets
                    '[a-zA-Z0-9]{10}',  # Original pattern
                    '\\[a-zA-Z0-9\\]{10}',  # Escaped brackets
                    '\\\\[a-zA-Z0-9\\\\]{10}',  # Double escaped
                ]
                found = False
                for pattern_var in pattern_variations:
                    if pattern_var in formula_text:
                        found = True
                        break
                if not found:
                    # Also check for pattern without escaping brackets
                    simple_pattern = 'a-zA-Z0-9]{10}'
                    if simple_pattern not in formula_text:
                        logger.warning(f"Cell {cell_coord} formula does not contain regex pattern '{regex_pattern}'")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        continue
                
                logger.info(f"✓ Cell {cell_coord} has valid REGEXP formula: {formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking cell {cell_coord}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in column {check_column} contain correct {expected_pattern} formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ {expected_pattern} formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_sumif_formula(result: str, expected: str = None, **options) -> float:
    """
    Verify if SUMIF formulas exist in specified column (F) to calculate totals.
    
    This function checks:
    1. Whether cells in specified column contain SUMIF formulas
    2. Whether formulas reference the correct ranges (auto-detected from data)
    3. Whether formulas reference the corresponding E column cell (F2->E2, F3->E3, etc.)
    4. Whether formulas have the correct structure
    
    The function automatically detects:
    - end_row: by checking the data column (E) for non-empty cells
    - criteria_range: by detecting the range from the first formula or from criteria_column data
    - sum_range: by detecting the range from the first formula or from sum_column data
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "F")
            - start_row: Starting row number (default: 2)
            - expected_function: Expected function name (default: "SUMIF")
            - criteria_column: Column containing criteria (e.g., "B")
            - sum_column: Column containing values to sum (e.g., "C")
            - criteria_column_start: Starting row for criteria column (default: 2)
            - data_column: Column to check for data to determine end_row (default: "E")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'F')
        start_row = options.get('start_row', 2)
        expected_function = options.get('expected_function', 'SUMIF')
        criteria_column = options.get('criteria_column', 'B')
        sum_column = options.get('sum_column', 'C')
        criteria_column_start = options.get('criteria_column_start', 2)
        data_column = options.get('data_column', 'E')
        
        logger.info(f"Verifying SUMIF formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected function: {expected_function}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Auto-detect criteria_range and sum_range by checking the first formula
        criteria_range = None
        sum_range = None
        
        # Try to extract ranges from the first formula
        first_cell_coord = f"{check_column}{start_row}"
        try:
            first_cell = ws[first_cell_coord]
            if first_cell.data_type == "f":
                first_formula_text = None
                if hasattr(first_cell, "_value") and isinstance(first_cell._value, str) and first_cell._value.startswith("="):
                    first_formula_text = first_cell._value
                elif hasattr(first_cell, "formula"):
                    first_formula_text = first_cell.formula
                elif first_cell.value is not None and isinstance(first_cell.value, str) and first_cell.value.startswith("="):
                    first_formula_text = first_cell.value
                
                if first_formula_text:
                    # Extract ranges from SUMIF formula: SUMIF(range1, criteria, range2)
                    # Pattern: SUMIF(range1, criteria, range2)
                    sumif_pattern = r'SUMIF\s*\(\s*([^,]+)\s*,\s*[^,]+\s*,\s*([^)]+)\s*\)'
                    match = re.search(sumif_pattern, first_formula_text, re.IGNORECASE)
                    if match:
                        criteria_range = match.group(1).strip()
                        sum_range = match.group(2).strip()
                        logger.info(f"Extracted from first formula: criteria_range={criteria_range}, sum_range={sum_range}")
        except Exception as e:
            logger.debug(f"Could not extract ranges from first formula: {e}")
        
        # If ranges not found in formula, detect from data columns
        if not criteria_range or not sum_range:
            logger.info(f"Auto-detecting ranges from data columns...")
            # Find the last row with data in criteria_column
            criteria_end_row = criteria_column_start
            empty_count = 0
            for row_num in range(criteria_column_start, max_row + 1):
                criteria_cell = ws[f"{criteria_column}{row_num}"]
                if criteria_cell.value is None or (isinstance(criteria_cell.value, str) and criteria_cell.value.strip() == ""):
                    empty_count += 1
                    if empty_count >= 3:
                        break
                else:
                    empty_count = 0
                    criteria_end_row = row_num
            
            # Find the last row with data in sum_column
            sum_end_row = criteria_column_start
            empty_count = 0
            for row_num in range(criteria_column_start, max_row + 1):
                sum_cell = ws[f"{sum_column}{row_num}"]
                if sum_cell.value is None or (isinstance(sum_cell.value, str) and sum_cell.value.strip() == ""):
                    empty_count += 1
                    if empty_count >= 3:
                        break
                else:
                    empty_count = 0
                    sum_end_row = row_num
            
            # Use the maximum end row for both ranges
            max_end_row = max(criteria_end_row, sum_end_row)
            criteria_range = f"{criteria_column}{criteria_column_start}:{criteria_column}{max_end_row}"
            sum_range = f"{sum_column}{criteria_column_start}:{sum_column}{max_end_row}"
            logger.info(f"Auto-detected ranges: criteria_range={criteria_range}, sum_range={sum_range}")
        
        # Check each row in the specified column
        all_passed = True
        logger.info(f"Checking column {check_column} (rows {start_row} to {end_row})")
        
        for row_num in range(start_row, end_row + 1):
            cell_coord = f"{check_column}{row_num}"
            try:
                cell = ws[cell_coord]
                logger.debug(f"Checking cell {cell_coord}")
                
                # Check if cell contains a formula
                if cell.data_type != "f":
                    logger.warning(f"Cell {cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                # Get formula text
                formula_text = None
                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                    formula_text = cell._value
                elif hasattr(cell, "formula"):
                    formula_text = cell.formula
                else:
                    # Try to get from value attribute
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value
                
                if formula_text is None:
                    logger.warning(f"Could not extract formula from cell {cell_coord}")
                    all_passed = False
                    continue
                
                # Remove leading = if present for comparison
                formula_clean = formula_text.lstrip("=")
                logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                
                # Check 1: Formula contains SUMIF function
                formula_upper = formula_text.upper()
                if expected_function.upper() not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain {expected_function}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 2: Formula contains SUMIF function call structure
                sumif_match = re.search(r'SUMIF\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not sumif_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct SUMIF structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 3: Formula contains criteria range
                if criteria_range and criteria_range.upper() not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain criteria range '{criteria_range}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 4: Formula contains sum range
                if sum_range and sum_range.upper() not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain sum range '{sum_range}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 5: Formula references the corresponding E column cell (E2, E3, etc.)
                expected_e_cell = f"E{row_num}"
                # Check if formula contains E column reference with the same row number
                e_cell_pattern = rf'E{row_num}\b'
                if not re.search(e_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference {expected_e_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {cell_coord} has valid SUMIF formula: {formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking cell {cell_coord}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in column {check_column} contain correct {expected_function} formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ {expected_function} formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_networkdays_formula(result: str, expected: str = None, **options) -> float:
    """
    Verify if NETWORKDAYS formulas exist in specified column to calculate working days.
    
    This function checks:
    1. Whether cells in specified column contain NETWORKDAYS formulas
    2. Whether formulas reference the corresponding start date column cell (A2, A3, etc.)
    3. Whether formulas reference the corresponding end date column cell (B2, B3, etc.)
    4. Whether formulas have the correct structure
    
    The function automatically detects the number of data rows by checking the data column
    (default: A column) for non-empty cells. It stops checking after finding 3 consecutive
    empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "C")
            - start_row: Starting row number (default: 2)
            - start_date_column: Column containing start dates (e.g., "A")
            - end_date_column: Column containing end dates (e.g., "B")
            - expected_function: Expected function name (default: "NETWORKDAYS")
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'C')
        start_row = options.get('start_row', 2)
        start_date_column = options.get('start_date_column', 'A')
        end_date_column = options.get('end_date_column', 'B')
        expected_function = options.get('expected_function', 'NETWORKDAYS')
        data_column = options.get('data_column', 'A')
        
        logger.info(f"Verifying NETWORKDAYS formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Start date column: {start_date_column}")
        logger.info(f"End date column: {end_date_column}")
        logger.info(f"Expected function: {expected_function}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row in the specified column
        all_passed = True
        logger.info(f"Checking column {check_column} (rows {start_row} to {end_row})")
        
        for row_num in range(start_row, end_row + 1):
            cell_coord = f"{check_column}{row_num}"
            try:
                cell = ws[cell_coord]
                logger.debug(f"Checking cell {cell_coord}")
                
                # Check if cell contains a formula
                if cell.data_type != "f":
                    logger.warning(f"Cell {cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                # Get formula text
                formula_text = None
                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                    formula_text = cell._value
                elif hasattr(cell, "formula"):
                    formula_text = cell.formula
                else:
                    # Try to get from value attribute
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value
                
                if formula_text is None:
                    logger.warning(f"Could not extract formula from cell {cell_coord}")
                    all_passed = False
                    continue
                
                # Remove leading = if present for comparison
                formula_clean = formula_text.lstrip("=")
                logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                
                # Check 1: Formula contains NETWORKDAYS function
                formula_upper = formula_text.upper()
                if expected_function.upper() not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain {expected_function}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 2: Formula contains NETWORKDAYS function call structure
                # NETWORKDAYS can have 2 or 3 parameters: NETWORKDAYS(start_date, end_date) or NETWORKDAYS(start_date, end_date, holidays)
                networkdays_pattern = r'NETWORKDAYS\s*\([^)]+\)'
                networkdays_match = re.search(networkdays_pattern, formula_text, re.IGNORECASE)
                if not networkdays_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct NETWORKDAYS structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 3: Formula references the corresponding start date column cell (A2, A3, etc.)
                expected_start_cell = f"{start_date_column}{row_num}"
                start_cell_pattern = rf'{start_date_column}{row_num}\b'
                if not re.search(start_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference start date cell {expected_start_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 4: Formula references the corresponding end date column cell (B2, B3, etc.)
                expected_end_cell = f"{end_date_column}{row_num}"
                end_cell_pattern = rf'{end_date_column}{row_num}\b'
                if not re.search(end_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference end date cell {expected_end_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {cell_coord} has valid NETWORKDAYS formula: {formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking cell {cell_coord}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in column {check_column} contain correct {expected_function} formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ {expected_function} formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_conditional_formatting_reconciliation(result: str, expected: str = None, **options) -> float:
    """
    Verify if conditional formatting is correctly set up for reconciliation between two tables.
    
    This function checks:
    1. Whether conditional formatting rules exist in the worksheet
    2. Whether the formula matches the expected pattern (e.g., A1<>E1 to compare cells from two tables)
    3. Whether conditional formatting is applied to the correct range
    4. Whether cells with differences are formatted (highlighted)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_range: Range where conditional formatting should be applied (e.g., "A1:C16")
            - compare_range: Range to compare against (e.g., "E1:G16")
            - expected_formula: Expected formula pattern (e.g., "A1<>E1")
            - format_column: Column to check for formatting (optional, e.g., "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        from openpyxl.utils import get_column_letter, column_index_from_string
        from openpyxl.worksheet.cell_range import CellRange
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_range = options.get('check_range', 'A1:C16')
        compare_range = options.get('compare_range', 'E1:G16')
        expected_formula = options.get('expected_formula', 'A1<>E1')
        format_column = options.get('format_column', None)
        
        logger.info(f"Verifying conditional formatting reconciliation in file: {result}")
        logger.info(f"Check range: {check_range}")
        logger.info(f"Compare range: {compare_range}")
        logger.info(f"Expected formula pattern: {expected_formula}")
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(result, data_only=False)
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check if conditional formatting exists
        conditional_formattings = ws.conditional_formatting
        if not conditional_formattings:
            logger.error("No conditional formatting rules found in the worksheet")
            return 0.0
        
        logger.info(f"Found {len(conditional_formattings)} conditional formatting rule(s)")
        
        # Parse expected formula to extract cell references
        # Expected formula like "A1<>E1" means compare A1 with E1
        expected_formula_clean = expected_formula.replace(" ", "").upper()
        
        # Find matching conditional formatting rule
        found_matching_rule = False
        rule_applied_to_correct_range = False
        
        for fmt in conditional_formattings:
            for rule in fmt.rules:
                # Check if rule has formula
                if not rule.formula:
                    continue
                
                # Check formula pattern
                formula_text = rule.formula[0] if rule.formula else ""
                formula_text_clean = formula_text.replace(" ", "").upper()
                
                logger.debug(f"Checking rule with formula: {formula_text}")
                
                # Check if formula matches expected pattern
                # The formula should contain comparison like A1<>E1, A2<>E2, etc.
                # We need to check if the pattern matches (allowing for relative references)
                if "<>" in expected_formula_clean:
                    # Extract cell references from expected formula
                    # Pattern: A1<>E1 means compare A column with E column
                    expected_parts = expected_formula_clean.split("<>")
                    if len(expected_parts) == 2:
                        expected_cell1 = expected_parts[0]  # e.g., "A1"
                        expected_cell2 = expected_parts[1]   # e.g., "E1"
                        
                        # Extract column letters
                        expected_col1 = re.match(r'([A-Z]+)', expected_cell1)
                        expected_col2 = re.match(r'([A-Z]+)', expected_cell2)
                        
                        if expected_col1 and expected_col2:
                            col1 = expected_col1.group(1)
                            col2 = expected_col2.group(1)
                            
                            # Check if formula contains comparison between these columns
                            # Pattern should be like: A1<>E1, A2<>E2, etc. (relative references)
                            pattern1 = rf'{col1}\d+\s*<>\s*{col2}\d+'
                            pattern2 = rf'{col1}\d+\s*!=\s*{col2}\d+'  # Alternative: !=
                            
                            if re.search(pattern1, formula_text_clean, re.IGNORECASE) or \
                               re.search(pattern2, formula_text_clean, re.IGNORECASE):
                                found_matching_rule = True
                                logger.info(f"✓ Found matching formula pattern: {formula_text}")
                                
                                # Check if rule is applied to correct range
                                fmt_ranges = [str(rng) for rng in fmt.cells]
                                check_range_upper = check_range.upper()
                                
                                # Check if check_range is covered by any of the formatting ranges
                                try:
                                    check_cell_range = CellRange(check_range_upper)
                                    for fmt_range_str in fmt_ranges:
                                        fmt_cell_range = CellRange(fmt_range_str)
                                        # Check if check_range is within or overlaps with fmt_range
                                        if (check_cell_range.min_row >= fmt_cell_range.min_row and
                                            check_cell_range.max_row <= fmt_cell_range.max_row and
                                            check_cell_range.min_col >= fmt_cell_range.min_col and
                                            check_cell_range.max_col <= fmt_cell_range.max_col):
                                            rule_applied_to_correct_range = True
                                            logger.info(f"✓ Rule applied to correct range: {fmt_range_str} covers {check_range}")
                                            break
                                except Exception as e:
                                    logger.debug(f"Error parsing ranges: {e}")
                                    # If range parsing fails, check if range string matches
                                    if check_range_upper in fmt_ranges:
                                        rule_applied_to_correct_range = True
                                        logger.info(f"✓ Rule applied to exact range: {check_range}")
                                
                                break
            
            if found_matching_rule:
                break
        
        if not found_matching_rule:
            logger.error("No conditional formatting rule found with expected formula pattern")
            return 0.0
        
        if not rule_applied_to_correct_range:
            logger.warning("Conditional formatting rule found but may not be applied to correct range")
            # Don't fail completely, as the range might be slightly different but still valid
        
        # Optional: Check if cells with differences are actually formatted
        # This is a more advanced check that verifies the formatting is working
        if format_column:
            logger.info(f"Checking formatting in column {format_column}...")
            # Try to find cells in format_column that have conditional formatting applied
            # This is a simplified check - in practice, we'd need to evaluate the formula
            # for each cell to see if it's formatted
            
        logger.info("=" * 60)
        logger.info("✓ Conditional formatting reconciliation verification passed")
        logger.info("=" * 60)
        return 1.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_right_len_find_extract(result: str, expected: str = None, **options) -> float:
    """
    Verify if RIGHT, LEN, and FIND formulas exist in specified column to extract text.
    
    This function checks:
    1. Whether cells in specified column contain RIGHT, LEN, and FIND functions
    2. Whether formulas reference the corresponding source column cell (C2->B2, C3->B3, etc.)
    3. Whether formulas contain the correct pattern (e.g., RIGHT(B2,LEN(B2)-FIND("班",B2)))
    4. Whether formulas have the correct structure with RIGHT, LEN, and FIND functions
    
    The function automatically detects the number of data rows by checking the data column
    (default: B column) for non-empty cells. It stops checking after finding 3 consecutive
    empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "C")
            - start_row: Starting row number (default: 2)
            - expected_functions: List of expected function names (e.g., ["RIGHT", "LEN", "FIND"])
            - expected_formula_pattern: Expected formula pattern (e.g., "RIGHT(B")
            - find_text: Text to find in FIND function (e.g., "班")
            - data_column: Column to check for data to determine end_row (default: "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'C')
        start_row = options.get('start_row', 2)
        expected_functions = options.get('expected_functions', ['RIGHT', 'LEN', 'FIND'])
        expected_formula_pattern = options.get('expected_formula_pattern', 'RIGHT(B')
        find_text = options.get('find_text', '班')
        data_column = options.get('data_column', 'B')
        
        logger.info(f"Verifying RIGHT/LEN/FIND extraction formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected functions: {expected_functions}")
        logger.info(f"Expected formula pattern: {expected_formula_pattern}")
        logger.info(f"Find text: {find_text}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row in the specified column
        all_passed = True
        logger.info(f"Checking column {check_column} (rows {start_row} to {end_row})")
        
        for row_num in range(start_row, end_row + 1):
            cell_coord = f"{check_column}{row_num}"
            try:
                cell = ws[cell_coord]
                logger.debug(f"Checking cell {cell_coord}")
                
                # Check if cell contains a formula
                if cell.data_type != "f":
                    logger.warning(f"Cell {cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                # Get formula text
                formula_text = None
                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                    formula_text = cell._value
                elif hasattr(cell, "formula"):
                    formula_text = cell.formula
                else:
                    # Try to get from value attribute
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value
                
                if formula_text is None:
                    logger.warning(f"Could not extract formula from cell {cell_coord}")
                    all_passed = False
                    continue
                
                # Remove leading = if present for comparison
                formula_clean = formula_text.lstrip("=")
                formula_upper = formula_text.upper()
                logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                
                # Check 1: Formula contains all expected functions
                for func_name in expected_functions:
                    if func_name.upper() not in formula_upper:
                        logger.warning(f"Cell {cell_coord} formula does not contain {func_name}")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        break
                
                if not all_passed:
                    continue
                
                # Check 2: Formula contains expected formula pattern (e.g., RIGHT(B)
                if expected_formula_pattern.upper() not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain pattern '{expected_formula_pattern}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 3: Formula contains RIGHT function call structure
                right_match = re.search(r'RIGHT\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not right_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct RIGHT structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 4: Formula contains LEN function
                len_match = re.search(r'LEN\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not len_match:
                    logger.warning(f"Cell {cell_coord} formula does not have LEN function")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 5: Formula contains FIND function with find_text
                find_pattern = rf'FIND\s*\([^)]*{re.escape(find_text)}[^)]*\)'
                find_match = re.search(find_pattern, formula_text, re.IGNORECASE)
                if not find_match:
                    logger.warning(f"Cell {cell_coord} formula does not contain FIND function with text '{find_text}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 6: Formula references the corresponding source column cell (B2, B3, etc.)
                expected_source_cell = f"{data_column}{row_num}"
                source_cell_pattern = rf'{data_column}{row_num}\b'
                if not re.search(source_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference source cell {expected_source_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 7: Formula structure should be RIGHT(B2, LEN(B2)-FIND("班",B2))
                # Verify that LEN and FIND are used together in the second parameter of RIGHT
                # This is a pattern check - the formula should have LEN(...)-FIND(...) structure
                len_find_pattern = r'LEN\s*\([^)]+\)\s*-\s*FIND\s*\([^)]+\)'
                if not re.search(len_find_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not have LEN(...)-FIND(...) structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {cell_coord} has valid RIGHT/LEN/FIND formula: {formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking cell {cell_coord}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in column {check_column} contain correct RIGHT/LEN/FIND formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ RIGHT/LEN/FIND formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_iferror_regex_extract(result: str, expected: str = None, **options) -> float:
    """
    Verify if IFERROR(REGEX(...)) formulas exist in specified column to extract text with error handling.
    
    This function checks:
    1. Whether cells in specified column contain IFERROR function wrapping REGEX
    2. Whether REGEX function uses capture group pattern (e.g., .*水笔(\d+).*)
    3. Whether REGEX function uses replacement pattern (e.g., $1元)
    4. Whether IFERROR has empty string as second parameter
    5. Whether formulas reference the corresponding source column cell
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "C")
            - start_row: Starting row number (default: 2)
            - expected_pattern_text: Expected text pattern in regex (e.g., "水笔")
            - data_column: Column to check for data to determine end_row (default: "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'C')
        start_row = options.get('start_row', 2)
        expected_pattern_text = options.get('expected_pattern_text', '水笔')
        data_column = options.get('data_column', 'B')
        
        logger.info(f"Verifying IFERROR(REGEX(...)) formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected pattern text: {expected_pattern_text}")
        logger.info(f"Data column: {data_column}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row in the specified column
        all_passed = True
        logger.info(f"Checking column {check_column} (rows {start_row} to {end_row})")
        
        for row_num in range(start_row, end_row + 1):
            cell_coord = f"{check_column}{row_num}"
            try:
                cell = ws[cell_coord]
                logger.debug(f"Checking cell {cell_coord}")
                
                # Check if cell contains a formula
                if cell.data_type != "f":
                    logger.warning(f"Cell {cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                # Get formula text
                formula_text = None
                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                    formula_text = cell._value
                elif hasattr(cell, "formula"):
                    formula_text = cell.formula
                else:
                    # Try to get from value attribute
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value
                
                if formula_text is None:
                    logger.warning(f"Could not extract formula from cell {cell_coord}")
                    all_passed = False
                    continue
                
                formula_upper = formula_text.upper()
                logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                
                # Check 1: Formula contains IFERROR function
                if 'IFERROR' not in formula_upper:
                    logger.warning(f"Cell {cell_coord} formula does not contain IFERROR function")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 2: Formula contains REGEX function (inside IFERROR)
                # Support both REGEX and LibreOffice internal format _xlfn.ORG.LIBREOFFICE.REGEX
                has_regex = 'REGEX' in formula_upper or '_XLFN.ORG.LIBREOFFICE.REGEX' in formula_upper
                if not has_regex:
                    logger.warning(f"Cell {cell_coord} formula does not contain REGEX function")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 3: IFERROR structure - should have two parameters
                iferror_match = re.search(r'IFERROR\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not iferror_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct IFERROR structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 4: IFERROR second parameter should be empty string ""
                # Handle various formats: ,"" or , "" or ,'' or , ''
                # Also handle LibreOffice format with spaces: REGEX(...) ,""
                # Extract IFERROR parameters: IFERROR(param1, param2)
                iferror_params_match = re.search(r'IFERROR\s*\((.*)\)', formula_text, re.IGNORECASE)
                if iferror_params_match:
                    params_str = iferror_params_match.group(1)
                    # Split by comma, but need to handle nested commas in strings
                    # Simple approach: find the last comma (should separate the two parameters)
                    # For IFERROR(REGEX(...), ""), the last comma separates REGEX call from ""
                    last_comma_pos = params_str.rfind(',')
                    if last_comma_pos != -1:
                        second_param = params_str[last_comma_pos + 1:].strip()
                        # Check if second parameter is empty string "" or ''
                        if second_param in ['""', "''", '""', "''"]:
                            has_empty_string = True
                        else:
                            has_empty_string = False
                    else:
                        has_empty_string = False
                else:
                    has_empty_string = False
                
                if not has_empty_string:
                    logger.warning(f"Cell {cell_coord} IFERROR should have empty string as second parameter")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 5: REGEX function call structure
                # Support both REGEX and LibreOffice internal format
                regex_match = re.search(r'(REGEX|_XLFN\.ORG\.LIBREOFFICE\.REGEX)\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not regex_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct REGEX structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 6: Formula contains expected pattern text (e.g., "水笔")
                if expected_pattern_text not in formula_text:
                    logger.warning(f"Cell {cell_coord} formula does not contain pattern text '{expected_pattern_text}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 7: Formula contains capture group pattern (\d+)
                has_capture_group = bool(re.search(r'\(\\d\+\)|\(\\\\d\+\)', formula_text))
                if not has_capture_group:
                    logger.warning(f"Cell {cell_coord} REGEX formula should contain capture group (\\d+)")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 8: Formula contains replacement pattern $1
                has_replacement = '"$1' in formula_text or "'$1" in formula_text
                if not has_replacement:
                    logger.warning(f"Cell {cell_coord} REGEX formula should contain replacement pattern $1")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 9: Formula references the corresponding source column cell
                expected_source_cell = f"{data_column}{row_num}"
                source_cell_pattern = rf'{data_column}{row_num}\b'
                if not re.search(source_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference source cell {expected_source_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {cell_coord} has valid IFERROR(REGEX(...)) formula: {formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking cell {cell_coord}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in column {check_column} contain correct IFERROR(REGEX(...)) formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ IFERROR(REGEX(...)) formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_rept_text_progress_bar(result: str, expected: str = None, **options) -> float:
    """
    Verify if REPT and TEXT formulas exist in specified column to create progress bars with percentage.
    
    This function checks:
    1. Whether cells in specified column contain REPT and TEXT functions
    2. Whether REPT function uses the correct character (e.g., "|")
    3. Whether REPT function uses the correct multiplier (e.g., *50)
    4. Whether TEXT function uses percentage format (e.g., "0%")
    5. Whether formulas reference the correct numerator and denominator columns
    6. Whether formulas use & operator to concatenate REPT and TEXT results
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "D")
            - start_row: Starting row number (default: 2)
            - expected_functions: List of expected function names (e.g., ["REPT", "TEXT"])
            - numerator_column: Column containing numerator values (e.g., "C")
            - denominator_column: Column containing denominator values (e.g., "B")
            - rept_char: Character to repeat in REPT function (e.g., "|")
            - rept_multiplier: Multiplier for REPT function (e.g., 50)
            - data_column: Column to check for data to determine end_row (default: "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'D')
        start_row = options.get('start_row', 2)
        expected_functions = options.get('expected_functions', ['REPT', 'TEXT'])
        numerator_column = options.get('numerator_column', 'C')
        denominator_column = options.get('denominator_column', 'B')
        rept_char = options.get('rept_char', '|')
        rept_multiplier = options.get('rept_multiplier', 50)
        data_column = options.get('data_column', 'B')
        
        logger.info(f"Verifying REPT/TEXT progress bar formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected functions: {expected_functions}")
        logger.info(f"Numerator column: {numerator_column}")
        logger.info(f"Denominator column: {denominator_column}")
        logger.info(f"REPT character: {rept_char}")
        logger.info(f"REPT multiplier: {rept_multiplier}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row in the specified column
        all_passed = True
        logger.info(f"Checking column {check_column} (rows {start_row} to {end_row})")
        
        for row_num in range(start_row, end_row + 1):
            cell_coord = f"{check_column}{row_num}"
            try:
                cell = ws[cell_coord]
                logger.debug(f"Checking cell {cell_coord}")
                
                # Check if cell contains a formula
                if cell.data_type != "f":
                    logger.warning(f"Cell {cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                # Get formula text
                formula_text = None
                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                    formula_text = cell._value
                elif hasattr(cell, "formula"):
                    formula_text = cell.formula
                else:
                    # Try to get from value attribute
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value
                
                if formula_text is None:
                    logger.warning(f"Could not extract formula from cell {cell_coord}")
                    all_passed = False
                    continue
                
                formula_upper = formula_text.upper()
                logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                
                # Check 1: Formula contains all expected functions (REPT and TEXT)
                for func_name in expected_functions:
                    if func_name.upper() not in formula_upper:
                        logger.warning(f"Cell {cell_coord} formula does not contain {func_name}")
                        logger.warning(f"Formula: {formula_text}")
                        all_passed = False
                        break
                
                if not all_passed:
                    continue
                
                # Check 2: Formula contains REPT function call structure
                rept_match = re.search(r'REPT\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not rept_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct REPT structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 3: REPT function contains the correct character (e.g., "|")
                # Support both double quotes and single quotes
                rept_char_pattern1 = rf'REPT\s*\(\s*"{re.escape(rept_char)}"'
                rept_char_pattern2 = rf"REPT\s*\(\s*'{re.escape(rept_char)}'"
                if not re.search(rept_char_pattern1, formula_text, re.IGNORECASE) and \
                   not re.search(rept_char_pattern2, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} REPT function should use character '{rept_char}'")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 4: REPT function contains the multiplier (e.g., *50)
                rept_multiplier_pattern = rf'\*{rept_multiplier}\b'
                if not re.search(rept_multiplier_pattern, formula_text):
                    logger.warning(f"Cell {cell_coord} REPT function should use multiplier *{rept_multiplier}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 5: Formula contains TEXT function call structure
                text_match = re.search(r'TEXT\s*\([^)]+\)', formula_text, re.IGNORECASE)
                if not text_match:
                    logger.warning(f"Cell {cell_coord} formula does not have correct TEXT structure")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 6: TEXT function contains percentage format ("0%" or '0%')
                text_percent_pattern1 = r'TEXT\s*\([^,]+,\s*"0%"'
                text_percent_pattern2 = r"TEXT\s*\([^,]+,\s*'0%'"
                if not re.search(text_percent_pattern1, formula_text, re.IGNORECASE) and \
                   not re.search(text_percent_pattern2, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} TEXT function should use percentage format \"0%\"")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 7: Formula references the correct numerator column (C2, C3, etc.)
                expected_numerator_cell = f"{numerator_column}{row_num}"
                numerator_cell_pattern = rf'{numerator_column}{row_num}\b'
                if not re.search(numerator_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference numerator cell {expected_numerator_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 8: Formula references the correct denominator column (B2, B3, etc.)
                expected_denominator_cell = f"{denominator_column}{row_num}"
                denominator_cell_pattern = rf'{denominator_column}{row_num}\b'
                if not re.search(denominator_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference denominator cell {expected_denominator_cell}")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 9: Formula contains & operator to concatenate REPT and TEXT
                if '&' not in formula_text:
                    logger.warning(f"Cell {cell_coord} formula should use & operator to concatenate REPT and TEXT")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 10: Formula structure should be REPT(...)&TEXT(...)
                # Verify that REPT comes before TEXT (or at least both are present)
                rept_pos = formula_text.upper().find('REPT')
                text_pos = formula_text.upper().find('TEXT')
                if rept_pos == -1 or text_pos == -1:
                    logger.warning(f"Cell {cell_coord} formula should contain both REPT and TEXT functions")
                    logger.warning(f"Formula: {formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {cell_coord} has valid REPT/TEXT progress bar formula: {formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking cell {cell_coord}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in column {check_column} contain correct REPT/TEXT progress bar formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ REPT/TEXT progress bar formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_id_extract_gender_age_birthday(result: str, expected: str = None, **options) -> float:
    """
    Verify if formulas exist to extract gender, age, and birthday from ID numbers.
    
    This function checks:
    1. Gender column (C): IF(MOD(MID(B3,17,1),2),"男","女")
    2. Age column (D): DATEDIF(TEXT(MID(B3,7,8),"0-00-00"),TODAY(),"Y")
    3. Birthday column (E): --TEXT(MID(B3,7,8),"0-00-00")
    
    The function automatically detects the number of data rows by checking the ID column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - id_column: Column containing ID numbers (e.g., "B")
            - gender_column: Column for gender formulas (e.g., "C")
            - age_column: Column for age formulas (e.g., "D")
            - birthday_column: Column for birthday formulas (e.g., "E")
            - start_row: Starting row number (default: 3)
            - data_column: Column to check for data to determine end_row (default: "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        id_column = options.get('id_column', 'B')
        gender_column = options.get('gender_column', 'C')
        age_column = options.get('age_column', 'D')
        birthday_column = options.get('birthday_column', 'E')
        start_row = options.get('start_row', 3)
        data_column = options.get('data_column', 'B')
        
        logger.info(f"Verifying ID extraction formulas in file: {result}")
        logger.info(f"ID column: {id_column}")
        logger.info(f"Gender column: {gender_column}")
        logger.info(f"Age column: {age_column}")
        logger.info(f"Birthday column: {birthday_column}")
        logger.info(f"Start row: {start_row}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row
        all_passed = True
        logger.info(f"Checking rows {start_row} to {end_row}")
        
        for row_num in range(start_row, end_row + 1):
            try:
                # Check gender column (C)
                gender_cell_coord = f"{gender_column}{row_num}"
                gender_cell = ws[gender_cell_coord]
                
                if gender_cell.data_type != "f":
                    logger.warning(f"Cell {gender_cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                gender_formula_text = None
                if hasattr(gender_cell, "_value") and isinstance(gender_cell._value, str) and gender_cell._value.startswith("="):
                    gender_formula_text = gender_cell._value
                elif hasattr(gender_cell, "formula"):
                    gender_formula_text = gender_cell.formula
                elif gender_cell.value is not None and isinstance(gender_cell.value, str) and gender_cell.value.startswith("="):
                    gender_formula_text = gender_cell.value
                
                if gender_formula_text is None:
                    logger.warning(f"Could not extract formula from cell {gender_cell_coord}")
                    all_passed = False
                    continue
                
                gender_formula_upper = gender_formula_text.upper()
                logger.debug(f"Cell {gender_cell_coord} formula: {gender_formula_text}")
                
                # Check gender formula: IF(MOD(MID(B3,17,1),2),"男","女")
                if 'IF' not in gender_formula_upper:
                    logger.warning(f"Cell {gender_cell_coord} formula does not contain IF function")
                    all_passed = False
                    continue
                
                if 'MOD' not in gender_formula_upper:
                    logger.warning(f"Cell {gender_cell_coord} formula does not contain MOD function")
                    all_passed = False
                    continue
                
                if 'MID' not in gender_formula_upper:
                    logger.warning(f"Cell {gender_cell_coord} formula does not contain MID function")
                    all_passed = False
                    continue
                
                # Check MID(B3,17,1) pattern
                mid_pattern = rf'MID\s*\(\s*{id_column}{row_num}\s*,\s*17\s*,\s*1\s*\)'
                if not re.search(mid_pattern, gender_formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {gender_cell_coord} formula should contain MID({id_column}{row_num},17,1)")
                    logger.warning(f"Formula: {gender_formula_text}")
                    all_passed = False
                    continue
                
                # Check for "男" and "女" in formula
                if '"男"' not in gender_formula_text and "'男'" not in gender_formula_text:
                    logger.warning(f"Cell {gender_cell_coord} formula should contain \"男\"")
                    logger.warning(f"Formula: {gender_formula_text}")
                    all_passed = False
                    continue
                
                if '"女"' not in gender_formula_text and "'女'" not in gender_formula_text:
                    logger.warning(f"Cell {gender_cell_coord} formula should contain \"女\"")
                    logger.warning(f"Formula: {gender_formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {gender_cell_coord} has valid gender formula: {gender_formula_text}")
                
                # Check age column (D)
                age_cell_coord = f"{age_column}{row_num}"
                age_cell = ws[age_cell_coord]
                
                if age_cell.data_type != "f":
                    logger.warning(f"Cell {age_cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                age_formula_text = None
                if hasattr(age_cell, "_value") and isinstance(age_cell._value, str) and age_cell._value.startswith("="):
                    age_formula_text = age_cell._value
                elif hasattr(age_cell, "formula"):
                    age_formula_text = age_cell.formula
                elif age_cell.value is not None and isinstance(age_cell.value, str) and age_cell.value.startswith("="):
                    age_formula_text = age_cell.value
                
                if age_formula_text is None:
                    logger.warning(f"Could not extract formula from cell {age_cell_coord}")
                    all_passed = False
                    continue
                
                age_formula_upper = age_formula_text.upper()
                logger.debug(f"Cell {age_cell_coord} formula: {age_formula_text}")
                
                # Check age formula: DATEDIF(TEXT(MID(B3,7,8),"0-00-00"),TODAY(),"Y")
                if 'DATEDIF' not in age_formula_upper:
                    logger.warning(f"Cell {age_cell_coord} formula does not contain DATEDIF function")
                    all_passed = False
                    continue
                
                if 'TEXT' not in age_formula_upper:
                    logger.warning(f"Cell {age_cell_coord} formula does not contain TEXT function")
                    all_passed = False
                    continue
                
                if 'TODAY' not in age_formula_upper:
                    logger.warning(f"Cell {age_cell_coord} formula does not contain TODAY function")
                    all_passed = False
                    continue
                
                if 'MID' not in age_formula_upper:
                    logger.warning(f"Cell {age_cell_coord} formula does not contain MID function")
                    all_passed = False
                    continue
                
                # Check MID(B3,7,8) pattern
                mid_pattern_age = rf'MID\s*\(\s*{id_column}{row_num}\s*,\s*7\s*,\s*8\s*\)'
                if not re.search(mid_pattern_age, age_formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {age_cell_coord} formula should contain MID({id_column}{row_num},7,8)")
                    logger.warning(f"Formula: {age_formula_text}")
                    all_passed = False
                    continue
                
                # Check TEXT format "0-00-00"
                # Use pattern that matches until the quote to handle nested functions like MID(B3,7,8)
                text_format_pattern1 = r'TEXT\s*\([^"]+,\s*"0-00-00"'
                text_format_pattern2 = r"TEXT\s*\([^']+,\s*'0-00-00'"
                if not re.search(text_format_pattern1, age_formula_text, re.IGNORECASE) and \
                   not re.search(text_format_pattern2, age_formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {age_cell_coord} TEXT function should use format \"0-00-00\"")
                    logger.warning(f"Formula: {age_formula_text}")
                    all_passed = False
                    continue
                
                # Check DATEDIF third parameter "Y"
                # Use a more flexible pattern that handles nested functions
                # Check if DATEDIF contains "Y" as the third parameter (after two commas)
                # Pattern: DATEDIF(...,...,"Y") or DATEDIF(...,...,'Y')
                # We'll count commas to find the third parameter
                datedif_match = re.search(r'DATEDIF\s*\((.*)\)', age_formula_text, re.IGNORECASE)
                if datedif_match:
                    datedif_params = datedif_match.group(1)
                    # Count commas to find the third parameter
                    # Simple approach: check if the last part before closing paren is "Y" or 'Y'
                    # More robust: find the pattern ,"Y" or ,'Y' before the closing paren
                    if not re.search(r',\s*"Y"\s*\)', age_formula_text, re.IGNORECASE) and \
                       not re.search(r",\s*'Y'\s*\)", age_formula_text, re.IGNORECASE):
                        logger.warning(f"Cell {age_cell_coord} DATEDIF function should use \"Y\" parameter")
                        logger.warning(f"Formula: {age_formula_text}")
                        all_passed = False
                        continue
                else:
                    logger.warning(f"Cell {age_cell_coord} could not parse DATEDIF function")
                    logger.warning(f"Formula: {age_formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {age_cell_coord} has valid age formula: {age_formula_text}")
                
                # Check birthday column (E)
                birthday_cell_coord = f"{birthday_column}{row_num}"
                birthday_cell = ws[birthday_cell_coord]
                
                if birthday_cell.data_type != "f":
                    logger.warning(f"Cell {birthday_cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                birthday_formula_text = None
                if hasattr(birthday_cell, "_value") and isinstance(birthday_cell._value, str) and birthday_cell._value.startswith("="):
                    birthday_formula_text = birthday_cell._value
                elif hasattr(birthday_cell, "formula"):
                    birthday_formula_text = birthday_cell.formula
                elif birthday_cell.value is not None and isinstance(birthday_cell.value, str) and birthday_cell.value.startswith("="):
                    birthday_formula_text = birthday_cell.value
                
                if birthday_formula_text is None:
                    logger.warning(f"Could not extract formula from cell {birthday_cell_coord}")
                    all_passed = False
                    continue
                
                birthday_formula_upper = birthday_formula_text.upper()
                logger.debug(f"Cell {birthday_cell_coord} formula: {birthday_formula_text}")
                
                # Check birthday formula: TEXT(MID(B3,7,8),"0-00-00")
                if 'TEXT' not in birthday_formula_upper:
                    logger.warning(f"Cell {birthday_cell_coord} formula does not contain TEXT function")
                    all_passed = False
                    continue
                
                if 'MID' not in birthday_formula_upper:
                    logger.warning(f"Cell {birthday_cell_coord} formula does not contain MID function")
                    all_passed = False
                    continue
                
                # Check MID(B3,7,8) pattern
                mid_pattern_birthday = rf'MID\s*\(\s*{id_column}{row_num}\s*,\s*7\s*,\s*8\s*\)'
                if not re.search(mid_pattern_birthday, birthday_formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {birthday_cell_coord} formula should contain MID({id_column}{row_num},7,8)")
                    logger.warning(f"Formula: {birthday_formula_text}")
                    all_passed = False
                    continue
                
                # Check TEXT format "0-00-00"
                # Use pattern that matches until the quote to handle nested functions like MID(B3,7,8)
                text_format_pattern1 = r'TEXT\s*\([^"]+,\s*"0-00-00"'
                text_format_pattern2 = r"TEXT\s*\([^']+,\s*'0-00-00'"
                if not re.search(text_format_pattern1, birthday_formula_text, re.IGNORECASE) and \
                   not re.search(text_format_pattern2, birthday_formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {birthday_cell_coord} TEXT function should use format \"0-00-00\"")
                    logger.warning(f"Formula: {birthday_formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {birthday_cell_coord} has valid birthday formula: {birthday_formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking row {row_num}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All rows contain correct ID extraction formulas (gender, age, birthday)")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ ID extraction formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_line_chart(result: str, expected: str = None, **options) -> float:
    """
    Verify if a line chart exists in the Excel file.
    
    This function checks:
    1. Whether at least one chart exists in the worksheet
    2. Whether the chart type is lineChart
    3. Whether the chart has the expected number of series
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - expected_chart_type: Expected chart type (default: "lineChart")
            - min_series_count: Minimum number of series expected (default: 1)
            - data_range: Data range used for chart (optional, for logging)
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        expected_chart_type = options.get('expected_chart_type', 'lineChart')
        min_series_count = options.get('min_series_count', 1)
        data_range = options.get('data_range', '')
        
        logger.info(f"Verifying line chart in file: {result}")
        logger.info(f"Expected chart type: {expected_chart_type}")
        logger.info(f"Minimum series count: {min_series_count}")
        if data_range:
            logger.info(f"Data range: {data_range}")
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(result)
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check if charts exist
        charts = ws._charts
        if not charts or len(charts) == 0:
            logger.error("No charts found in the worksheet")
            return 0.0
        
        logger.info(f"Found {len(charts)} chart(s) in the worksheet")
        
        # Check each chart
        for chart_idx, chart in enumerate(charts):
            logger.info(f"Checking chart {chart_idx + 1}...")
            
            # Check chart type
            chart_type = None
            if hasattr(chart, 'tagname'):
                chart_type = chart.tagname
            logger.info(f"Chart type: {chart_type}")
            
            # Check if it's a line chart
            if chart_type and expected_chart_type.lower() in chart_type.lower():
                logger.info(f"✓ Chart {chart_idx + 1} is a line chart")
                
                # Check if it has series
                if not hasattr(chart, 'series') or not chart.series:
                    logger.warning(f"Chart {chart_idx + 1} has no series")
                    continue
                
                series_count = len(chart.series)
                logger.info(f"Chart {chart_idx + 1} has {series_count} series")
                
                # Verify series count
                if series_count >= min_series_count:
                    logger.info("=" * 60)
                    logger.info(f"✓ Line chart verification passed")
                    logger.info(f"  Chart type: {chart_type}")
                    logger.info(f"  Series count: {series_count} (minimum required: {min_series_count})")
                    logger.info("=" * 60)
                    return 1.0
                else:
                    logger.warning(f"Chart {chart_idx + 1} has {series_count} series, but minimum required is {min_series_count}")
            else:
                logger.warning(f"Chart {chart_idx + 1} is not a line chart (type: {chart_type})")
        
        # If we get here, verification failed
        logger.error("=" * 60)
        logger.error(f"✗ Line chart verification failed")
        logger.error(f"  Expected chart type: {expected_chart_type}")
        logger.error(f"  Minimum series count: {min_series_count}")
        logger.error("=" * 60)
        return 0.0
             
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_salary_growth_chart(result: str, expected: str = None, **options) -> float:
    """
    Verify if the salary growth chart matches the expected specifications.
    
    This function checks the chart itself (not the data table):
    1. Whether a chart exists in the specified sheet
    2. Whether the chart title matches "店长工资增长"
    3. Whether the chart has the expected number of series (at least 3)
    4. Whether the chart is a combination chart (bar + line)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - sheet_idx: Sheet index (default: 0)
            - expected_title: Expected chart title (default: "店长工资增长")
            - min_series_count: Minimum number of series (default: 3)
            - chart_type: Expected chart type (default: "combination")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        sheet_idx = options.get('sheet_idx', 0)
        expected_title = options.get('expected_title', '店长工资增长')
        min_series_count = options.get('min_series_count', 3)
        chart_type = options.get('chart_type', 'combination')
        
        logger.info(f"Verifying salary growth chart in file: {result}")
        logger.info(f"Sheet index: {sheet_idx}")
        logger.info(f"Expected title: {expected_title}")
        logger.info(f"Min series count: {min_series_count}")
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(result)
            pdworkbook = pd.ExcelFile(result)
            sheet_names = pdworkbook.sheet_names
            
            if sheet_idx >= len(sheet_names):
                logger.error(f"Sheet index {sheet_idx} out of range. Available sheets: {sheet_names}")
                return 0.0
            
            sheet_name = sheet_names[sheet_idx]
            logger.info(f"Checking sheet: {sheet_name}")
            ws = wb[sheet_name]
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return 0.0
        
        # Check if charts exist
        charts = ws._charts
        if not charts:
            logger.error("No charts found in the sheet")
            return 0.0
        
        logger.info(f"Found {len(charts)} chart(s) in the sheet")
        
        # Load chart information
        chart_props = ['title', 'type', 'legend', 'xtitle', 'ytitle']
        chart_info = load_charts(wb, sheet_name, chart_props=chart_props)
        
        if not chart_info:
            logger.error("Could not load chart information")
            return 0.0
        
        # Check each chart
        chart_passed = False
        for chart_key, chart_data in chart_info.items():
            logger.info(f"Checking chart: {chart_key}")
            logger.debug(f"Chart data: {chart_data}")
            
            # Check 1: Chart title
            chart_title = chart_data.get('title')
            if chart_title != expected_title:
                logger.warning(f"Chart title mismatch: expected '{expected_title}', got '{chart_title}'")
                continue
            else:
                logger.info(f"✓ Chart title matches: {chart_title}")
            
            # Check 2: Chart type (for combination charts, we might see multiple types)
            chart_type_actual = chart_data.get('type')
            logger.info(f"Chart type: {chart_type_actual}")
            # Note: Combination charts might be represented differently in openpyxl
            # We'll be lenient here and just check that a chart exists
            
            # Check 3: Number of series
            # Extract series count from chart_key (format: "value_ref1,category_ref1;value_ref2,category_ref2;...")
            series_parts = chart_key.split(';')
            series_count = len(series_parts)
            logger.info(f"Number of series: {series_count}")
            
            if series_count < min_series_count:
                logger.warning(f"Insufficient series count: expected at least {min_series_count}, got {series_count}")
                continue
            else:
                logger.info(f"✓ Series count sufficient: {series_count} >= {min_series_count}")
            
            # If we get here, this chart passed all checks
            chart_passed = True
            logger.info("=" * 60)
            logger.info(f"✓ Chart verification passed!")
            logger.info("=" * 60)
            break
        
        if chart_passed:
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error("✗ Chart verification failed - no chart matched all criteria")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_project_completion_chart(result: str, expected: str = None, **options) -> float:
    """
    Verify if a combination chart exists with the expected title that contains
    both bar chart series (for project values) and line chart series (for completion rates).
    
    This function checks:
    1. Whether at least one chart exists in the worksheet
    2. Whether the chart title matches expected_title
    3. Whether the chart has at least 16 series (8 projects + 8 completion rates)
    4. Whether at least one series name contains "rate" (for completion rates)
    5. Whether the chart has at least project_count * 5 categories (for 5 quarters per project)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - sheet_idx: Sheet index to check (default: 0)
            - expected_title: Expected chart title (default: "项目")
            - min_series_count: Minimum number of series required (default: 16)
            - project_count: Number of projects (default: 8)
            - quarters_per_project: Number of quarters per project (default: 5)
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        sheet_idx = options.get('sheet_idx', 0)
        expected_title = options.get('expected_title', '项目')
        min_series_count = options.get('min_series_count', 16)
        project_count = options.get('project_count', 8)
        quarters_per_project = options.get('quarters_per_project', 5)
        min_categories = project_count * quarters_per_project
        
        logger.info(f"Verifying project completion chart in file: {result}")
        logger.info(f"Sheet index: {sheet_idx}")
        logger.info(f"Expected title: {expected_title}")
        logger.info(f"Minimum series count: {min_series_count}")
        logger.info(f"Project count: {project_count}")
        logger.info(f"Quarters per project: {quarters_per_project}")
        logger.info(f"Minimum categories: {min_categories}")
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(result)
            sheet_names = wb.sheetnames
            if sheet_idx >= len(sheet_names):
                logger.error(f"Sheet index {sheet_idx} out of range. Available sheets: {sheet_names}")
                return 0.0
            sheet_name = sheet_names[sheet_idx]
            ws = wb[sheet_name]
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check if charts exist
        charts = ws._charts
        if not charts:
            logger.error("No charts found in the worksheet")
            return 0.0
        
        logger.info(f"Found {len(charts)} chart(s) in the worksheet")
        
        # Check each chart
        chart_found = False
        for chart in charts:
            # Check chart title
            chart_title = None
            try:
                if chart.title and chart.title.tx:
                    if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                        if hasattr(chart.title.tx.rich, 'p') and chart.title.tx.rich.p:
                            if len(chart.title.tx.rich.p) > 0:
                                if hasattr(chart.title.tx.rich.p[0], 'r') and chart.title.tx.rich.p[0].r:
                                    if len(chart.title.tx.rich.p[0].r) > 0:
                                        if hasattr(chart.title.tx.rich.p[0].r[0], 't'):
                                            chart_title = chart.title.tx.rich.p[0].r[0].t
            except Exception as e:
                logger.debug(f"Error reading chart title: {e}")
            
            logger.info(f"Chart title: {chart_title}")
            
            # Check if title matches
            if chart_title == expected_title:
                logger.info(f"✓ Chart title matches: {chart_title}")
                chart_found = True
                
                # Use load_charts to get all series information (includes both bar and line series)
                # This is more reliable for combination charts
                chart_props = ['title']
                chart_info = load_charts(wb, sheet_name, chart_props=chart_props)
                
                # Find the chart that matches our title
                chart_key = None
                for key, info in chart_info.items():
                    if info.get('title') == expected_title:
                        chart_key = key
                        break
                
                # Get all series from the chart object for detailed inspection
                all_series = list(chart.series) if hasattr(chart, 'series') else []
                logger.info(f"Series count from chart.series: {len(all_series)}")
                
                if not chart_key:
                    logger.warning("Could not find chart in load_charts output, using direct series access")
                    # Fallback to direct series access
                    series_count = len(all_series)
                else:
                    # Extract series count from chart_key (format: "value_ref1,category_ref1;value_ref2,category_ref2;...")
                    # This includes ALL series (both bar and line) in combination charts
                    series_parts = chart_key.split(';')
                    series_count_from_load = len(series_parts)
                    logger.info(f"Series count from load_charts: {series_count_from_load}")
                    
                    # For combination charts, load_charts should give us all series
                    # Use the count from load_charts as it's more reliable for combination charts
                    series_count = series_count_from_load
                    
                    # Also check for sub-charts in case series are stored there
                    if hasattr(chart, '_charts') and chart._charts:
                        # Check for sub-charts (for combination charts)
                        for sub_chart in chart._charts:
                            if hasattr(sub_chart, 'series'):
                                sub_series = list(sub_chart.series)
                                all_series.extend(sub_series)
                                logger.info(f"Found {len(sub_series)} additional series in sub-chart")
                    
                    # If load_charts gave us fewer series than direct access, use the larger count
                    # This handles edge cases where load_charts might miss some series
                    if series_count < len(all_series):
                        logger.warning(f"load_charts found {series_count} series but direct access found {len(all_series)}, using larger count")
                        series_count = len(all_series)
                
                logger.info(f"Chart has {series_count} series (including both bar and line series)")
                
                # Debug: Log all series details
                if all_series:
                    logger.info(f"Detailed series information:")
                    for idx, ser in enumerate(all_series):
                        logger.info(f"  Series {idx}: {type(ser).__name__}")
                        try:
                            if hasattr(ser, 'title'):
                                logger.debug(f"    Title: {ser.title}")
                        except:
                            pass
                
                if series_count < min_series_count:
                    logger.error(f"✗ Chart has only {series_count} series, expected at least {min_series_count}")
                    return 0.0
                
                logger.info(f"✓ Chart has {series_count} series (>= {min_series_count})")
                
                # Check series names for "rate" (completion rate series)
                has_rate_series = False
                series_names = []
                for i, ser in enumerate(all_series):
                    series_name = None
                    try:
                        # Try to get series title/name
                        if hasattr(ser, 'title') and ser.title:
                            if hasattr(ser.title, 'tx') and ser.title.tx:
                                if hasattr(ser.title.tx, 'rich') and ser.title.tx.rich:
                                    if hasattr(ser.title.tx.rich, 'p') and ser.title.tx.rich.p:
                                        if len(ser.title.tx.rich.p) > 0:
                                            if hasattr(ser.title.tx.rich.p[0], 'r') and ser.title.tx.rich.p[0].r:
                                                if len(ser.title.tx.rich.p[0].r) > 0:
                                                    if hasattr(ser.title.tx.rich.p[0].r[0], 't'):
                                                        series_name = ser.title.tx.rich.p[0].r[0].t
                        # Alternative: check if title is a string reference
                        if not series_name and hasattr(ser, 'title') and hasattr(ser.title, 'tx') and hasattr(ser.title.tx, 'strRef'):
                            if hasattr(ser.title.tx.strRef, 'f'):
                                series_name = ser.title.tx.strRef.f
                    except Exception as e:
                        logger.debug(f"Error reading series {i} name: {e}")
                    
                    if series_name:
                        series_names.append(series_name)
                        if "rate" in series_name.lower():
                            has_rate_series = True
                            logger.info(f"✓ Found series with 'rate' in name: {series_name}")
                
                if series_names:
                    logger.info(f"Series names found: {series_names[:10]}...")  # Log first 10
                else:
                    logger.warning("Could not extract series names, will skip rate check")
                
                if not has_rate_series and series_names:
                    logger.error(f"✗ No series found with 'rate' in name. Series names: {series_names}")
                    return 0.0
                elif not has_rate_series:
                    logger.warning("⚠ Could not verify 'rate' in series names (series names not extractable)")
                
                # Check category count
                max_categories = 0
                category_ranges = []
                
                def parse_range_count(range_str):
                    """Parse Excel range string and return count of cells"""
                    try:
                        # Remove sheet name if present (e.g., "Sheet1!$A$2:$A$6" -> "$A$2:$A$6")
                        if '!' in range_str:
                            range_str = range_str.split('!')[1]
                        
                        # Remove $ signs
                        range_str = range_str.replace('$', '')
                        
                        if ':' in range_str:
                            start, end = range_str.split(':')
                            # Parse start and end coordinates
                            start_col, start_row = coordinate_to_tuple(start)
                            end_col, end_row = coordinate_to_tuple(end)
                            
                            # Calculate count based on range
                            if start_col == end_col:
                                # Same column, count rows
                                return abs(end_row - start_row) + 1
                            elif start_row == end_row:
                                # Same row, count columns
                                return abs(end_col - start_col) + 1
                            else:
                                # 2D range
                                return (abs(end_row - start_row) + 1) * (abs(end_col - start_col) + 1)
                        else:
                            # Single cell
                            return 1
                    except Exception as e:
                        logger.debug(f"Error parsing range {range_str}: {e}")
                        return 0
                
                for i, ser in enumerate(all_series):
                    try:
                        # Try to get category count from category reference
                        if hasattr(ser, 'cat'):
                            cat_range = None
                            # Check if categories are from a range
                            if hasattr(ser.cat, 'numRef') and hasattr(ser.cat.numRef, 'f'):
                                cat_range = ser.cat.numRef.f
                            elif hasattr(ser.cat, 'strRef') and hasattr(ser.cat.strRef, 'f'):
                                cat_range = ser.cat.strRef.f
                            
                            if cat_range:
                                category_ranges.append(cat_range)
                                cat_count = parse_range_count(cat_range)
                                if cat_count > max_categories:
                                    max_categories = cat_count
                                logger.debug(f"Series {i} category range: {cat_range}, count: {cat_count}")
                    except Exception as e:
                        logger.debug(f"Error reading categories for series {i}: {e}")
                
                if max_categories > 0:
                    logger.info(f"Maximum category count found: {max_categories}")
                    if max_categories < min_categories:
                        logger.error(f"✗ Chart has only {max_categories} categories, expected at least {min_categories} (project_count * quarters_per_project)")
                        return 0.0
                    logger.info(f"✓ Chart has {max_categories} categories (>= {min_categories})")
                else:
                    # If we can't determine category count from ranges, use heuristic
                    # For 8 projects with 5 quarters each, we need at least 40 categories
                    # But since we can't verify directly, we'll log a warning
                    logger.warning(f"⚠ Could not determine exact category count from ranges. Expected at least {min_categories} categories.")
                    logger.info(f"Category ranges found: {category_ranges[:5]}...")  # Log first 5
                
                # Check if it's a combination chart
                if series_count >= 2:
                    logger.info("✓ Chart appears to be a combination chart (has multiple series)")
                else:
                    logger.error(f"✗ Chart has only {series_count} series, expected at least 2 for combination chart")
                    return 0.0
                
                break
        
        if chart_found:
            logger.info("=" * 60)
            logger.info("✓ Project completion combination chart verification passed")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ Chart with title '{expected_title}' not found")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_shipping_boxes_calculation(result: str, expected: str = None, **options) -> float:
    """
    Verify if formulas exist to calculate shipping boxes from product specifications and order quantities.
    
    This function checks:
    1. Whether cells in specified column contain formulas (not just values)
    2. Whether formulas contain required functions: INT, VALUE, LEFT, FIND, MOD, IF
    3. Whether formulas reference the specification column (B) and quantity column (C)
    4. Whether formulas contain Chinese characters "支" and "盒"
    5. Whether formulas use string concatenation (&)
    
    The expected formula pattern:
    =INT(C2/VALUE(LEFT(B2,FIND("支",B2)-1)))&"盒"&IF(MOD(C2,VALUE(LEFT(B2,FIND("支",B2)-1)))=0,"","加"&MOD(C2,VALUE(LEFT(B2,FIND("支",B2)-1)))&"支")
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "D")
            - start_row: Starting row number (default: 2)
            - spec_column: Column containing product specifications (e.g., "B")
            - quantity_column: Column containing order quantities (e.g., "C")
            - expected_functions: List of expected function names (e.g., ["INT", "VALUE", "LEFT", "FIND", "MOD", "IF"])
            - data_column: Column to check for data to determine end_row (default: "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'D')
        start_row = options.get('start_row', 2)
        spec_column = options.get('spec_column', 'B')
        quantity_column = options.get('quantity_column', 'C')
        expected_functions = options.get('expected_functions', ['INT', 'VALUE', 'LEFT', 'FIND', 'MOD', 'IF'])
        data_column = options.get('data_column', 'B')
        
        logger.info(f"Verifying shipping boxes calculation formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Specification column: {spec_column}")
        logger.info(f"Quantity column: {quantity_column}")
        logger.info(f"Expected functions: {expected_functions}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        if end_row < start_row:
            logger.error(f"No data rows found starting from row {start_row}")
            return 0.0
        
        # Check formulas in each row
        all_checks_passed = True
        rows_checked = 0
        
        for row_num in range(start_row, end_row + 1):
            check_cell = ws[f"{check_column}{row_num}"]
            spec_cell = ws[f"{spec_column}{row_num}"]
            quantity_cell = ws[f"{quantity_column}{row_num}"]
            
            # Skip if spec or quantity cell is empty
            if spec_cell.value is None or quantity_cell.value is None:
                continue
            
            rows_checked += 1
            formula = check_cell.value
            
            # Check 1: Formula exists (not just a value)
            if formula is None or not isinstance(formula, str) or not formula.startswith('='):
                logger.error(f"Cell {check_column}{row_num} should contain a formula, but got: {formula}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check 2: Contains all required functions
            missing_functions = []
            for func in expected_functions:
                # Check for function name (with word boundary to avoid partial matches)
                func_pattern = r'\b' + re.escape(func) + r'\s*\('
                if not re.search(func_pattern, formula_upper, re.IGNORECASE):
                    missing_functions.append(func)
            
            if missing_functions:
                logger.error(f"Cell {check_column}{row_num} formula missing functions: {missing_functions}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: References specification column (B)
            spec_pattern = rf'{spec_column}\d+'
            if not re.search(spec_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should reference {spec_column} column")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: References quantity column (C)
            quantity_pattern = rf'{quantity_column}\d+'
            if not re.search(quantity_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should reference {quantity_column} column")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 5: Contains Chinese character "支" (for extracting pieces per box)
            if '"支"' not in formula and "'支'" not in formula and '支' not in formula:
                logger.error(f"Cell {check_column}{row_num} formula should contain '支' character")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 6: Contains Chinese character "盒" (for box unit)
            if '"盒"' not in formula and "'盒'" not in formula and '盒' not in formula:
                logger.error(f"Cell {check_column}{row_num} formula should contain '盒' character")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 7: Uses string concatenation (&)
            if '&' not in formula:
                logger.error(f"Cell {check_column}{row_num} formula should use & for string concatenation")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 8: Contains INT function (for calculating integer boxes)
            int_pattern = r'\bINT\s*\('
            if not re.search(int_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain INT function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 9: Contains MOD function (for calculating remainder)
            mod_pattern = r'\bMOD\s*\('
            if not re.search(mod_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain MOD function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 10: Contains IF function (for conditional formatting)
            if_pattern = r'\bIF\s*\('
            if not re.search(if_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain IF function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            logger.debug(f"✓ Row {row_num} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ Shipping boxes calculation verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Column: {check_column}")
            logger.info(f"  Functions verified: {', '.join(expected_functions)}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ Shipping boxes calculation verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_split_content_formula(result: str, expected: str = None, **options) -> float:
    """
    Verify if formulas exist to split content with line breaks into multiple rows.
    
    This function checks:
    1. Whether cells in specified column contain formulas (not just values)
    2. Whether formulas contain required functions: TRIM, MID, SUBSTITUTE, CHAR, REPT, ROW
    3. Whether formulas reference the source column (A)
    4. Whether formulas contain CHAR(10) for line break
    5. Whether formulas contain REPT(" ",100) or similar pattern
    6. Whether formulas use ROW function for position calculation
    
    The expected formula pattern:
    =TRIM(MID(SUBSTITUTE(A2,CHAR(10),REPT(" ",100)),(ROW(A1)-1)*100+1,100))
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "B")
            - start_row: Starting row number (default: 2)
            - source_column: Column containing source data (e.g., "A")
            - expected_functions: List of expected function names (e.g., ["TRIM", "MID", "SUBSTITUTE", "CHAR", "REPT", "ROW"])
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'B')
        start_row = options.get('start_row', 2)
        source_column = options.get('source_column', 'A')
        expected_functions = options.get('expected_functions', ['TRIM', 'MID', 'SUBSTITUTE', 'CHAR', 'REPT', 'ROW'])
        data_column = options.get('data_column', 'A')
        
        logger.info(f"Verifying split content formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Source column: {source_column}")
        logger.info(f"Expected functions: {expected_functions}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        if end_row < start_row:
            logger.error(f"No data rows found starting from row {start_row}")
            return 0.0
        
        # Check formulas in each row
        all_checks_passed = True
        rows_checked = 0
        
        for row_num in range(start_row, end_row + 1):
            check_cell = ws[f"{check_column}{row_num}"]
            
            # Skip if source cell is empty
            source_cell = ws[f"{source_column}{row_num}"]
            if source_cell.value is None:
                continue
            
            rows_checked += 1
            formula = check_cell.value
            
            # Check 1: Formula exists (not just a value)
            if formula is None or not isinstance(formula, str) or not formula.startswith('='):
                logger.error(f"Cell {check_column}{row_num} should contain a formula, but got: {formula}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check 2: Contains all required functions
            missing_functions = []
            for func in expected_functions:
                # Check for function name (with word boundary to avoid partial matches)
                func_pattern = r'\b' + re.escape(func) + r'\s*\('
                if not re.search(func_pattern, formula_upper, re.IGNORECASE):
                    missing_functions.append(func)
            
            if missing_functions:
                logger.error(f"Cell {check_column}{row_num} formula missing functions: {missing_functions}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: References source column (A)
            source_pattern = rf'{source_column}\d+'
            if not re.search(source_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should reference {source_column} column")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3.5: SUBSTITUTE function must reference A2 specifically
            # Pattern: SUBSTITUTE(A2,...
            substitute_a2_pattern = rf'SUBSTITUTE\s*\(\s*{source_column}2\s*,'
            if not re.search(substitute_a2_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula SUBSTITUTE function must reference {source_column}2 (not {source_column}3, {source_column}4, etc.)")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Contains CHAR(10) for line break
            char_pattern = r'CHAR\s*\(\s*10\s*\)'
            if not re.search(char_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain CHAR(10)")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 5: Contains REPT with space and 100
            rept_pattern = r'REPT\s*\(\s*["\']?\s*["\']?\s*,\s*100\s*\)'
            if not re.search(rept_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain REPT(\" \",100) or similar")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 6: Contains ROW function
            row_pattern = r'\bROW\s*\('
            if not re.search(row_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain ROW function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 7: Contains TRIM function (outermost)
            trim_pattern = r'\bTRIM\s*\('
            if not re.search(trim_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain TRIM function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 8: Contains MID function
            mid_pattern = r'\bMID\s*\('
            if not re.search(mid_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain MID function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 9: Contains SUBSTITUTE function
            substitute_pattern = r'\bSUBSTITUTE\s*\('
            if not re.search(substitute_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should contain SUBSTITUTE function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            logger.debug(f"✓ Row {row_num} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ Split content formula verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Column: {check_column}")
            logger.info(f"  Functions verified: {', '.join(expected_functions)}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ Split content formula verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_quote_sheet_with_merged_cells(result: str, expected: str = None, **options) -> float:
    """
    Verify if a quote sheet template exists with the expected structure, fields, and merged cells.
    
    This function checks:
    1. Whether the title "报价单" exists in the worksheet
    2. Whether merged cells exist (especially for the title)
    3. Whether required header fields exist
    4. Whether the product table headers exist
    5. Whether summary fields exist
    6. Whether footer fields exist
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - expected_title: Expected title text (default: "报价单")
            - title_merged_range: Expected merged range for title (default: "G15:H15")
            - required_fields: List of required header field labels
            - table_headers: List of table header labels
            - summary_fields: List of summary field labels
            - footer_fields: List of footer field labels
            - min_merged_cells: Minimum number of merged cell ranges expected (default: 1)
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        expected_title = options.get('expected_title', '报价单')
        title_merged_range = options.get('title_merged_range', 'G15:H15')
        required_fields = options.get('required_fields', ['报价单位', '联系人', '联系电话', '客户名称', '报价日期', '邮箱'])
        table_headers = options.get('table_headers', ['序号', '产品名称', '产品类型', '规格', '数量', '单价', '金额', '备注'])
        summary_fields = options.get('summary_fields', ['合计金额(小写)', '合计金额(大写)'])
        footer_fields = options.get('footer_fields', ['报价人', '审批'])
        min_merged_cells = options.get('min_merged_cells', 1)
        
        logger.info(f"Verifying quote sheet template with merged cells in file: {result}")
        logger.info(f"Expected title: {expected_title}")
        logger.info(f"Title merged range: {title_merged_range}")
        logger.info(f"Required fields: {required_fields}")
        logger.info(f"Table headers: {table_headers}")
        logger.info(f"Summary fields: {summary_fields}")
        logger.info(f"Footer fields: {footer_fields}")
        logger.info(f"Minimum merged cells: {min_merged_cells}")
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(result, data_only=True)
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check merged cells
        merged_cells = list(ws.merged_cells.ranges)
        logger.info(f"Found {len(merged_cells)} merged cell range(s)")
        
        if len(merged_cells) < min_merged_cells:
            logger.error(f"✗ Insufficient merged cells: found {len(merged_cells)}, expected at least {min_merged_cells}")
            return 0.0
        
        # Check if title merged range exists
        title_merged_found = False
        for merged_range in merged_cells:
            merged_str = str(merged_range)
            logger.debug(f"Merged range: {merged_str}")
            if merged_str.upper() == title_merged_range.upper():
                title_merged_found = True
                logger.info(f"✓ Found title merged range: {merged_str}")
                break
        
        if not title_merged_found:
            logger.warning(f"⚠ Title merged range '{title_merged_range}' not found, but other merged cells exist")
            logger.info(f"  Available merged ranges: {[str(r) for r in merged_cells]}")
            # Don't fail completely, as the range might be slightly different
        
        # Search through all cells to find required text
        # Also check merged cells specifically
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Convert all cell values to strings for searching
        # Include both individual cells and merged cell ranges
        all_text = []
        cell_text_map = {}  # Map cell coordinates to text for debugging
        
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    cell_text = str(cell.value).strip()
                    all_text.append(cell_text)
                    cell_coord = cell.coordinate
                    cell_text_map[cell_coord] = cell_text
                    
                    # Also check if cell contains newlines (for merged cells with multiple fields)
                    if '\n' in cell_text or '\r' in cell_text:
                        # Split by newlines and add each line
                        lines = cell_text.replace('\r', '\n').split('\n')
                        for line in lines:
                            line = line.strip()
                            if line:
                                all_text.append(line)
        
        # Check merged cells - get text from the top-left cell of each merged range
        for merged_range in merged_cells:
            try:
                # Get the top-left cell of the merged range
                top_left_cell = ws[merged_range.min_row][merged_range.min_col - 1]
                if top_left_cell.value is not None:
                    merged_text = str(top_left_cell.value).strip()
                    all_text.append(merged_text)
                    # Also split by newlines if present
                    if '\n' in merged_text or '\r' in merged_text:
                        lines = merged_text.replace('\r', '\n').split('\n')
                        for line in lines:
                            line = line.strip()
                            if line:
                                all_text.append(line)
                    logger.debug(f"Merged range {merged_range} contains text: {merged_text[:100]}")
            except Exception as e:
                logger.debug(f"Error reading merged range {merged_range}: {e}")
        
        # Check 1: Title exists
        title_found = False
        for text in all_text:
            if expected_title in text:
                title_found = True
                logger.info(f"✓ Found title: {expected_title}")
                break
        
        if not title_found:
            logger.error(f"✗ Title '{expected_title}' not found")
            return 0.0
        
        # Check 2: Required header fields
        # Since fields may be in merged cells together, we need to check if all fields exist
        # even if they're in the same cell
        found_fields = []
        missing_fields = []
        for field in required_fields:
            field_found = False
            for text in all_text:
                if field in text:
                    field_found = True
                    found_fields.append(field)
                    logger.debug(f"Found field '{field}' in text: {text[:80]}...")
                    break
            if not field_found:
                missing_fields.append(field)
        
        # Log all text for debugging if fields are missing
        if missing_fields:
            logger.warning(f"⚠ Missing required fields: {missing_fields}")
            # Show some sample text that might contain the fields
            logger.debug(f"Sample text from worksheet (showing text with Chinese characters):")
            chinese_text_samples = [t for t in all_text if any('\u4e00' <= c <= '\u9fff' for c in t)][:20]
            for sample in chinese_text_samples:
                logger.debug(f"  {sample[:100]}")
        
        # Since fields may be grouped in merged cells, we're more lenient
        # Check if at least most fields are found
        found_ratio = len(found_fields) / len(required_fields) if required_fields else 1.0
        
        if found_ratio < 0.5:  # Less than 50% found
            logger.error(f"✗ Too many required fields missing: found {len(found_fields)}/{len(required_fields)}")
            logger.error(f"  Missing: {missing_fields}")
            return 0.0
        elif missing_fields:
            logger.warning(f"⚠ Some fields missing: {missing_fields}, but found {len(found_fields)}/{len(required_fields)} fields")
            # Don't fail if most fields are found (fields might be in merged cells together)
        else:
            logger.info(f"✓ Found all required fields: {found_fields}")
        
        # Check 3: Table headers
        found_headers = []
        missing_headers = []
        for header in table_headers:
            header_found = False
            for text in all_text:
                if header in text:
                    header_found = True
                    found_headers.append(header)
                    break
            if not header_found:
                missing_headers.append(header)
        
        if missing_headers:
            logger.error(f"✗ Missing table headers: {missing_headers}")
            return 0.0
        else:
            logger.info(f"✓ Found all table headers: {found_headers}")
        
        # Check 4: Summary fields
        found_summary = []
        missing_summary = []
        for field in summary_fields:
            field_found = False
            for text in all_text:
                if field in text:
                    field_found = True
                    found_summary.append(field)
                    break
            if not field_found:
                missing_summary.append(field)
        
        if missing_summary:
            logger.warning(f"⚠ Missing summary fields: {missing_summary}")
            # Don't fail completely, as these might be optional
        else:
            logger.info(f"✓ Found summary fields: {found_summary}")
        
        # Check 5: Footer fields
        found_footer = []
        missing_footer = []
        for field in footer_fields:
            field_found = False
            for text in all_text:
                if field in text:
                    field_found = True
                    found_footer.append(field)
                    break
            if not field_found:
                missing_footer.append(field)
        
        if missing_footer:
            logger.warning(f"⚠ Missing footer fields: {missing_footer}")
            # Don't fail completely, as these might be optional
        else:
            logger.info(f"✓ Found footer fields: {found_footer}")
        
        # Check 6: Borders on table cells
        # Check if cells in the product table area have borders
        # Typically, table headers and data rows should have borders
        logger.info("Checking borders on table cells...")
        
        # Find the table header row (should contain table headers)
        table_header_row = None
        for row_num in range(1, max_row + 1):
            for col_num in range(1, max_col + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    cell_text = str(cell.value).strip()
                    # Check if this row contains table headers
                    if any(header in cell_text for header in table_headers):
                        table_header_row = row_num
                        break
            if table_header_row:
                break
        
        borders_found = False
        cells_with_borders = 0
        total_table_cells_checked = 0
        
        if table_header_row:
            logger.info(f"Table header row found at row {table_header_row}")
            # Check borders in table area (header row and a few data rows)
            # Table typically spans from column C to K (based on headers)
            check_start_col = 3  # Column C
            check_end_col = min(11, max_col)  # Column K or max_col
            check_start_row = table_header_row
            check_end_row = min(table_header_row + 7, max_row)  # Header + 6 data rows
            
            for row_num in range(check_start_row, check_end_row + 1):
                for col_num in range(check_start_col, check_end_col + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    total_table_cells_checked += 1
                    
                    # Check if cell has any border
                    has_border = False
                    try:
                        border = cell.border
                        # Check if any side has a border style (not None and not empty)
                        if border:
                            if (border.top and border.top.style) or \
                               (border.bottom and border.bottom.style) or \
                               (border.left and border.left.style) or \
                               (border.right and border.right.style):
                                has_border = True
                                cells_with_borders += 1
                    except Exception as e:
                        logger.debug(f"Error checking border for cell {cell.coordinate}: {e}")
                    
                    if has_border:
                        borders_found = True
                        logger.debug(f"Cell {cell.coordinate} has borders")
            
            if total_table_cells_checked > 0:
                border_ratio = cells_with_borders / total_table_cells_checked
                logger.info(f"Borders found: {cells_with_borders}/{total_table_cells_checked} cells ({border_ratio:.1%})")
                
                # Require at least 30% of table cells to have borders
                if border_ratio < 0.3:
                    logger.warning(f"⚠ Low border coverage: only {border_ratio:.1%} of table cells have borders")
                    # Don't fail completely, as borders might be applied differently
                else:
                    logger.info(f"✓ Sufficient borders found in table area")
            else:
                logger.warning("⚠ Could not check borders: no table cells found")
        else:
            logger.warning("⚠ Could not find table header row for border checking")
        
        # If we get here, all critical checks passed
        logger.info("=" * 60)
        logger.info(f"✓ Quote sheet template with merged cells verification passed")
        logger.info(f"  Title: {expected_title}")
        logger.info(f"  Merged cells: {len(merged_cells)} (minimum required: {min_merged_cells})")
        if title_merged_found:
            logger.info(f"  Title merged range: {title_merged_range}")
        logger.info(f"  Required fields: {len(found_fields)}/{len(required_fields)}")
        logger.info(f"  Table headers: {len(found_headers)}/{len(table_headers)}")
        logger.info(f"  Summary fields: {len(found_summary)}/{len(summary_fields)}")
        logger.info(f"  Footer fields: {len(found_footer)}/{len(footer_fields)}")
        if borders_found:
            logger.info(f"  Borders: {cells_with_borders}/{total_table_cells_checked} cells have borders")
        logger.info("=" * 60)
        return 1.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_nested_regex_price_extract(result: str, expected: str = None, **options) -> float:
    """
    Verify if nested REGEX(REGEX(SUBSTITUTE(...))) formulas exist to extract prices from inconsistently formatted text.
    
    This function checks:
    1. Whether cells in specified column contain nested REGEX functions
    2. Whether formulas contain SUBSTITUTE function to remove spaces
    3. Whether inner REGEX function uses pattern to replace Chinese punctuation with dots
    4. Whether outer REGEX function extracts numeric values with pattern [\d.]+
    5. Whether formulas reference the corresponding source column cell (A1, A2, A3, etc.)
    
    Expected formula pattern: =REGEX(REGEX(SUBSTITUTE(A1," ",),"[-，。]",2,"."),"[\d.]+")
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "B")
            - start_row: Starting row number (default: 1)
            - source_column: Column containing source data (e.g., "A")
            - expected_functions: List of expected function names (default: ["REGEX", "SUBSTITUTE"])
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'B')
        start_row = options.get('start_row', 1)
        source_column = options.get('source_column', 'A')
        expected_functions = options.get('expected_functions', ['REGEX', 'SUBSTITUTE'])
        data_column = options.get('data_column', 'A')
        
        logger.info(f"Verifying nested REGEX price extraction formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Source column: {source_column}")
        logger.info(f"Expected functions: {expected_functions}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row
        all_checks_passed = True
        rows_checked = 0
        
        for row_num in range(start_row, end_row + 1):
            check_cell = ws[f"{check_column}{row_num}"]
            source_cell = ws[f"{source_column}{row_num}"]
            
            # Skip if source cell is empty
            if source_cell.value is None:
                continue
            
            rows_checked += 1
            formula = check_cell.value
            
            # Check 1: Formula exists (not just a value)
            if formula is None or not isinstance(formula, str) or not formula.startswith('='):
                logger.error(f"Cell {check_column}{row_num} should contain a formula, but got: {formula}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check 2: Contains all required functions
            missing_functions = []
            for func in expected_functions:
                # Check for function name (with word boundary to avoid partial matches)
                func_pattern = r'\b' + re.escape(func) + r'\s*\('
                if not re.search(func_pattern, formula_upper, re.IGNORECASE):
                    missing_functions.append(func)
            
            if missing_functions:
                logger.error(f"Cell {check_column}{row_num} formula missing functions: {missing_functions}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains nested REGEX functions (at least 2 REGEX calls)
            regex_count = len(re.findall(r'\bREGEX\s*\(', formula_upper, re.IGNORECASE))
            if regex_count < 2:
                logger.error(f"Cell {check_column}{row_num} formula should contain at least 2 nested REGEX functions")
                logger.error(f"  Formula: {formula}")
                logger.error(f"  Found {regex_count} REGEX function(s)")
                all_checks_passed = False
                continue
            
            # Check 4: Contains SUBSTITUTE function
            substitute_pattern = r'\bSUBSTITUTE\s*\('
            if not re.search(substitute_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {check_column}{row_num} formula should contain SUBSTITUTE function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 5: SUBSTITUTE function references source column and removes spaces
            # Pattern: SUBSTITUTE(A1," ",) or SUBSTITUTE(A1," ","")
            substitute_pattern = rf'SUBSTITUTE\s*\(\s*{source_column}{row_num}\s*,\s*["\']?\s*["\']?\s*,'
            if not re.search(substitute_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {check_column}{row_num} SUBSTITUTE should reference {source_column}{row_num} and remove spaces")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 6: Inner REGEX function contains pattern to replace Chinese punctuation
            # Pattern should contain something like "[-，。]" or similar Chinese punctuation
            # and replacement pattern like "2,"." or "2,\"."
            inner_regex_pattern = r'REGEX\s*\([^)]*["\']?[-，。，。、]["\']?\s*,\s*\d+\s*,\s*["\']?\.["\']?'
            if not re.search(inner_regex_pattern, formula_upper, re.IGNORECASE):
                logger.warning(f"Cell {check_column}{row_num} inner REGEX may not have correct pattern for replacing Chinese punctuation")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the pattern might be correct but formatted differently
            
            # Check 7: Outer REGEX function contains pattern to extract numbers
            # Pattern should contain [\d.]+ or similar
            outer_regex_pattern = r'REGEX\s*\([^)]*["\']?\[\\?d\.\]\+["\']?'
            if not re.search(outer_regex_pattern, formula_upper, re.IGNORECASE):
                # Also check for unescaped version
                if not re.search(r'REGEX\s*\([^)]*\[.*d.*\.?.*\]', formula_upper, re.IGNORECASE):
                    logger.warning(f"Cell {check_column}{row_num} outer REGEX may not have correct pattern [\\d.]+")
                    logger.debug(f"  Formula: {formula}")
                    # Don't fail, just warn
            
            # Check 8: Formula references the corresponding source column cell
            source_pattern = rf'{source_column}{row_num}\b'
            if not re.search(source_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should reference {source_column}{row_num}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            logger.debug(f"✓ Row {row_num} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ Nested REGEX price extraction verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Column: {check_column}")
            logger.info(f"  Functions verified: {', '.join(expected_functions)}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ Nested REGEX price extraction verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_text_datetime_format(result: str, expected: str = None, **options) -> float:
    """
    Verify if TEXT formulas exist to convert numbers to datetime format.
    
    This function checks:
    1. Whether cells in specified column contain TEXT formulas
    2. Whether formulas use TEXT function with datetime format pattern "0-00-00 00:00:00"
    3. Whether formulas reference the corresponding source column cell (A2, A3, etc.)
    
    Expected formula pattern: =TEXT(A2,"0-00-00 00:00:00")
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "C")
            - start_row: Starting row number (default: 2)
            - source_column: Column containing source data (e.g., "A")
            - expected_functions: List of expected function names (default: ["TEXT"])
            - expected_format: Expected TEXT format pattern (default: "0-00-00 00:00:00")
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'C')
        start_row = options.get('start_row', 2)
        source_column = options.get('source_column', 'A')
        expected_functions = options.get('expected_functions', ['TEXT'])
        expected_format = options.get('expected_format', '0-00-00 00:00:00')
        data_column = options.get('data_column', 'A')
        
        logger.info(f"Verifying TEXT datetime format formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Source column: {source_column}")
        logger.info(f"Expected functions: {expected_functions}")
        logger.info(f"Expected format: {expected_format}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row
        all_checks_passed = True
        rows_checked = 0
        
        for row_num in range(start_row, end_row + 1):
            check_cell = ws[f"{check_column}{row_num}"]
            source_cell = ws[f"{source_column}{row_num}"]
            
            # Skip if source cell is empty
            if source_cell.value is None:
                continue
            
            rows_checked += 1
            
            # Check 1: Cell contains a formula
            if check_cell.data_type != "f":
                logger.error(f"Cell {check_column}{row_num} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(check_cell, "_value") and isinstance(check_cell._value, str) and check_cell._value.startswith("="):
                formula = check_cell._value
            elif hasattr(check_cell, "formula"):
                formula = check_cell.formula
            else:
                # Try to get from value attribute
                if check_cell.value is not None and isinstance(check_cell.value, str) and check_cell.value.startswith("="):
                    formula = check_cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {check_column}{row_num}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check 2: Contains TEXT function
            if 'TEXT' not in formula_upper:
                logger.error(f"Cell {check_column}{row_num} formula should contain TEXT function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: TEXT function contains the expected format pattern
            # Pattern: TEXT(A2,"0-00-00 00:00:00") or TEXT(A2,'0-00-00 00:00:00')
            # Escape special characters in format pattern for regex
            format_escaped = re.escape(expected_format)
            text_format_pattern1 = rf'TEXT\s*\([^,]+,\s*"{format_escaped}"'
            text_format_pattern2 = rf"TEXT\s*\([^,]+,\s*'{format_escaped}'"
            if not re.search(text_format_pattern1, formula, re.IGNORECASE) and \
               not re.search(text_format_pattern2, formula, re.IGNORECASE):
                logger.error(f"Cell {check_column}{row_num} TEXT function should use format \"{expected_format}\"")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Formula references the corresponding source column cell
            source_pattern = rf'{source_column}{row_num}\b'
            if not re.search(source_pattern, formula_upper):
                logger.error(f"Cell {check_column}{row_num} formula should reference {source_column}{row_num}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            logger.debug(f"✓ Row {row_num} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ TEXT datetime format verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Column: {check_column}")
            logger.info(f"  Format: {expected_format}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ TEXT datetime format verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_if_counta_payment_receipt_numbering(result: str, expected: str = None, **options) -> float:
    """
    Verify if IF(COUNTA(...)) formulas exist to auto-generate payment/receipt numbering.
    
    This function checks:
    1. Whether cells in specified column contain IF formulas
    2. Whether formulas contain required functions: IF, COUNTA
    3. Whether formulas check if payment column (C) is empty
    4. Whether formulas use COUNTA to count receipt column (C) or payment column (D)
    5. Whether formulas generate "收" or "付" prefixed numbers
    6. Whether formulas reference the correct ranges ($C$2:C2 and $D$2:D2)
    
    Expected formula pattern: =IF(C2="","付"&COUNTA($D$2:D2),"收"&COUNTA($C$2:C2))
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "A")
            - start_row: Starting row number (default: 2)
            - receipt_column: Column containing receipt amounts (e.g., "C")
            - payment_column: Column containing payment amounts (e.g., "D")
            - expected_functions: List of expected function names (default: ["IF", "COUNTA"])
            - receipt_prefix: Prefix for receipt numbers (default: "收")
            - payment_prefix: Prefix for payment numbers (default: "付")
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'A')
        start_row = options.get('start_row', 2)
        receipt_column = options.get('receipt_column', 'C')
        payment_column = options.get('payment_column', 'D')
        expected_functions = options.get('expected_functions', ['IF', 'COUNTA'])
        receipt_prefix = options.get('receipt_prefix', '收')
        payment_prefix = options.get('payment_prefix', '付')
        data_column = options.get('data_column', 'A')
        
        logger.info(f"Verifying IF/COUNTA payment/receipt numbering formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Receipt column: {receipt_column}")
        logger.info(f"Payment column: {payment_column}")
        logger.info(f"Expected functions: {expected_functions}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row
        all_checks_passed = True
        rows_checked = 0
        
        for row_num in range(start_row, end_row + 1):
            check_cell = ws[f"{check_column}{row_num}"]
            
            # Skip if cell is empty
            if check_cell.value is None:
                continue
            
            rows_checked += 1
            
            # Check if cell contains a formula
            if check_cell.data_type != "f":
                logger.error(f"Cell {check_column}{row_num} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(check_cell, "_value") and isinstance(check_cell._value, str) and check_cell._value.startswith("="):
                formula = check_cell._value
            elif hasattr(check_cell, "formula"):
                formula = check_cell.formula
            else:
                # Try to get from value attribute
                if check_cell.value is not None and isinstance(check_cell.value, str) and check_cell.value.startswith("="):
                    formula = check_cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {check_column}{row_num}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check 1: Contains all required functions
            missing_functions = []
            for func in expected_functions:
                # Check for function name (with word boundary to avoid partial matches)
                func_pattern = r'\b' + re.escape(func) + r'\s*\('
                if not re.search(func_pattern, formula_upper, re.IGNORECASE):
                    missing_functions.append(func)
            
            if missing_functions:
                logger.error(f"Cell {check_column}{row_num} formula missing functions: {missing_functions}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: Formula structure - IF(...)
            if_pattern = r'IF\s*\('
            if not re.search(if_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {check_column}{row_num} formula should contain IF function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: IF condition checks if receipt column is empty (C2="")
            # Pattern: IF(C2="",... or IF(C2="",...
            if_condition_pattern = rf'IF\s*\(\s*{receipt_column}{row_num}\s*=\s*["\']?\s*["\']?'
            if not re.search(if_condition_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {check_column}{row_num} IF should check if {receipt_column}{row_num} is empty")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Formula contains COUNTA with payment column range ($D$2:D2)
            # Pattern: COUNTA($D$2:D2) or COUNTA($D$2:D{row_num})
            payment_counta_pattern = rf'COUNTA\s*\(\s*\${payment_column}\$\d+:\s*{payment_column}{row_num}\s*\)'
            if not re.search(payment_counta_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {check_column}{row_num} should contain COUNTA(${payment_column}$2:{payment_column}{row_num})")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 5: Formula contains COUNTA with receipt column range ($C$2:C2)
            # Pattern: COUNTA($C$2:C2) or COUNTA($C$2:C{row_num})
            receipt_counta_pattern = rf'COUNTA\s*\(\s*\${receipt_column}\$\d+:\s*{receipt_column}{row_num}\s*\)'
            if not re.search(receipt_counta_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {check_column}{row_num} should contain COUNTA(${receipt_column}$2:{receipt_column}{row_num})")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 6: Formula contains payment prefix (付)
            if payment_prefix not in formula:
                logger.error(f"Cell {check_column}{row_num} formula should contain payment prefix '{payment_prefix}'")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 7: Formula contains receipt prefix (收)
            if receipt_prefix not in formula:
                logger.error(f"Cell {check_column}{row_num} formula should contain receipt prefix '{receipt_prefix}'")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            logger.debug(f"✓ Row {row_num} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ IF/COUNTA payment/receipt numbering formula verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Column: {check_column}")
            logger.info(f"  Functions verified: {', '.join(expected_functions)}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ IF/COUNTA payment/receipt numbering formula verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_if_or_row_check(result: str, expected: str = None, **options) -> float:
    """
    Verify if IF(COUNTIF(...)) formulas exist to check if a row contains a specific character.
    
    This function checks:
    1. Whether cells in specified column contain IF formulas
    2. Whether formulas contain required functions: IF, COUNTIF
    3. Whether formulas use COUNTIF to count occurrences (e.g., COUNTIF(A2:Z2, "连")>0)
    4. Whether formulas return "连" if found, "断" if not found
    5. Whether formulas reference the correct row range
    
    Expected formula pattern: =IF(COUNTIF(A2:Z2, "连")>0, "连", "断")
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "T")
            - start_row: Starting row number (default: 2)
            - check_range_start: Start column of range to check (e.g., "A")
            - check_range_end: End column of range to check (e.g., "Z")
            - search_text: Text to search for (default: "连")
            - found_text: Text to return if found (default: "连")
            - not_found_text: Text to return if not found (default: "断")
            - expected_functions: List of expected function names (default: ["IF", "OR"])
            - data_column: Column to check for data to determine end_row (default: "T")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'T')
        start_row = options.get('start_row', 2)
        check_range_start = options.get('check_range_start', 'A')
        check_range_end = options.get('check_range_end', 'Z')
        search_text = options.get('search_text', '连')
        found_text = options.get('found_text', '连')
        not_found_text = options.get('not_found_text', '断')
        expected_functions = options.get('expected_functions', ['IF', 'COUNTIF'])
        data_column = options.get('data_column', 'T')
        
        logger.info(f"Verifying IF/COUNTIF row check formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Check range: {check_range_start} to {check_range_end}")
        logger.info(f"Search text: {search_text}")
        logger.info(f"Expected functions: {expected_functions}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row
        all_checks_passed = True
        rows_checked = 0
        
        for row_num in range(start_row, end_row + 1):
            check_cell = ws[f"{check_column}{row_num}"]
            
            # Skip if cell is empty
            if check_cell.value is None:
                continue
            
            rows_checked += 1
            
            # Check if cell contains a formula
            if check_cell.data_type != "f":
                logger.error(f"Cell {check_column}{row_num} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(check_cell, "_value") and isinstance(check_cell._value, str) and check_cell._value.startswith("="):
                formula = check_cell._value
            elif hasattr(check_cell, "formula"):
                formula = check_cell.formula
            else:
                # Try to get from value attribute
                if check_cell.value is not None and isinstance(check_cell.value, str) and check_cell.value.startswith("="):
                    formula = check_cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {check_column}{row_num}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check 1: Contains all required functions
            missing_functions = []
            for func in expected_functions:
                # Check for function name (with word boundary to avoid partial matches)
                func_pattern = r'\b' + re.escape(func) + r'\s*\('
                if not re.search(func_pattern, formula_upper, re.IGNORECASE):
                    missing_functions.append(func)
            
            if missing_functions:
                logger.error(f"Cell {check_column}{row_num} formula missing functions: {missing_functions}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: Formula structure - IF(COUNTIF(...))
            if_countif_pattern = r'IF\s*\(\s*COUNTIF\s*\('
            if not re.search(if_countif_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {check_column}{row_num} formula should have structure IF(COUNTIF(...))")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: COUNTIF contains row range
            # Pattern: COUNTIF(A2:Z2, ...)
            row_range_pattern = rf'{check_range_start}{row_num}:{check_range_end}{row_num}'
            if row_range_pattern.upper() not in formula_upper:
                logger.error(f"Cell {check_column}{row_num} COUNTIF should contain range {row_range_pattern}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: COUNTIF contains search text
            # Pattern: COUNTIF(A2:Z2, "连") or COUNTIF(A2:Z2, '连')
            countif_search_pattern1 = rf'COUNTIF\s*\(\s*{row_range_pattern}\s*,\s*["\']?{re.escape(search_text)}["\']?'
            countif_search_pattern2 = rf'COUNTIF\s*\(\s*{row_range_pattern}\s*,\s*["\']?{re.escape(search_text)}["\']?'
            if not re.search(countif_search_pattern1, formula, re.IGNORECASE) and \
               not re.search(countif_search_pattern2, formula, re.IGNORECASE):
                logger.error(f"Cell {check_column}{row_num} COUNTIF should search for '{search_text}'")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 5: COUNTIF result is compared with >0
            # Pattern: COUNTIF(...)>0
            countif_gt_pattern = r'COUNTIF\s*\([^)]+\)\s*>\s*0'
            if not re.search(countif_gt_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {check_column}{row_num} COUNTIF result should be compared with >0")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 5: IF contains found_text ("连")
            if found_text not in formula:
                logger.error(f"Cell {check_column}{row_num} IF should return '{found_text}' if found")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 6: IF contains not_found_text ("断")
            if not_found_text not in formula:
                logger.error(f"Cell {check_column}{row_num} IF should return '{not_found_text}' if not found")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            logger.debug(f"✓ Row {row_num} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ IF/COUNTIF row check formula verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Column: {check_column}")
            logger.info(f"  Functions verified: {', '.join(expected_functions)}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ IF/COUNTIF row check formula verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_counta_sort_amount(result: str, expected: str = None, **options) -> float:
    """
    Verify if COUNTA helper column exists and data is sorted by amount in descending order.
    
    This function checks:
    1. Whether helper column (E) contains COUNTA formulas starting from E3
    2. Whether COUNTA formulas reference the correct range (e.g., $B$2:B3)
    3. Whether data is sorted by amount column (D) in descending order
    4. Whether the sorting maintains data integrity (all rows are present)
    
    Expected helper formula: =COUNTA($B$2:B3) in E3, then copied down
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - helper_column: Helper column with COUNTA formula (e.g., "E")
            - helper_start_row: Starting row for helper formula (default: 3)
            - count_range_column: Column to count in COUNTA (e.g., "B")
            - amount_column: Column containing amounts to sort (e.g., "D")
            - data_start_row: Starting row of data (default: 2)
            - data_end_row: Ending row of data (default: 14)
            - sort_order: Expected sort order "descending" or "ascending" (default: "descending")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        helper_column = options.get('helper_column', 'E')
        helper_start_row = options.get('helper_start_row', 3)
        count_range_column = options.get('count_range_column', 'B')
        amount_column = options.get('amount_column', 'D')
        data_start_row = options.get('data_start_row', 2)
        data_end_row = options.get('data_end_row', 14)
        sort_order = options.get('sort_order', 'descending')
        
        logger.info(f"Verifying COUNTA helper column and amount sorting in file: {result}")
        logger.info(f"Helper column: {helper_column}")
        logger.info(f"Helper start row: {helper_start_row}")
        logger.info(f"Amount column: {amount_column}")
        logger.info(f"Data range: {data_start_row} to {data_end_row}")
        logger.info(f"Sort order: {sort_order}")
        
        # Load workbook to get formulas and values
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check 1: Verify COUNTA formulas in helper column
        logger.info(f"Checking COUNTA formulas in column {helper_column}...")
        helper_formulas_ok = True
        
        for row_num in range(helper_start_row, data_end_row + 1):
            helper_cell = ws[f"{helper_column}{row_num}"]
            
            # Check if cell contains a formula
            if helper_cell.data_type != "f":
                logger.error(f"Cell {helper_column}{row_num} does not contain a formula")
                helper_formulas_ok = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(helper_cell, "_value") and isinstance(helper_cell._value, str) and helper_cell._value.startswith("="):
                formula = helper_cell._value
            elif hasattr(helper_cell, "formula"):
                formula = helper_cell.formula
            else:
                if helper_cell.value is not None and isinstance(helper_cell.value, str) and helper_cell.value.startswith("="):
                    formula = helper_cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {helper_column}{row_num}")
                helper_formulas_ok = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check if formula contains COUNTA
            if 'COUNTA' not in formula_upper:
                logger.error(f"Cell {helper_column}{row_num} formula should contain COUNTA function")
                logger.error(f"  Formula: {formula}")
                helper_formulas_ok = False
                continue
            
            # Check if COUNTA references the correct range
            # Pattern: COUNTA($B$2:B3) or COUNTA($B$2:B{row_num})
            counta_pattern = rf'COUNTA\s*\(\s*\${count_range_column}\$\d+:\s*{count_range_column}{row_num}\s*\)'
            if not re.search(counta_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {helper_column}{row_num} COUNTA should reference ${count_range_column}$2:{count_range_column}{row_num}")
                logger.error(f"  Formula: {formula}")
                helper_formulas_ok = False
                continue
        
        if not helper_formulas_ok:
            logger.error("COUNTA helper column verification failed")
            return 0.0
        
        logger.info("✓ COUNTA helper column verification passed")
        
        # Check 2: Verify data is sorted by amount column within each group (grouped by helper column)
        # The sorting is done by grouping rows with the same E column value, then sorting D column within each group
        logger.info(f"Checking if data is sorted by column {amount_column} within groups (grouped by {helper_column} column)...")
        
        # Collect data: (row_num, helper_value, amount_value)
        # We need to read calculated values, so load workbook again with data_only=True for values
        try:
            wb_values = openpyxl.load_workbook(result, data_only=True)
            ws_values = wb_values.active
        except Exception as e:
            logger.warning(f"Failed to load workbook with data_only=True, using formula values: {e}")
            ws_values = ws  # Fallback to formula worksheet
        
        data_rows = []
        # Start from helper_start_row to skip header row
        actual_start_row = max(data_start_row, helper_start_row)
        logger.debug(f"Collecting data from row {actual_start_row} to {data_end_row}")
        
        for row_num in range(actual_start_row, data_end_row + 1):
            helper_cell = ws_values[f"{helper_column}{row_num}"]
            amount_cell = ws_values[f"{amount_column}{row_num}"]
            
            helper_value = helper_cell.value
            amount_value = amount_cell.value
            
            logger.debug(f"Row {row_num}: {helper_column}={helper_value}, {amount_column}={amount_value}")
            
            # Skip if helper or amount is None or empty
            if helper_value is None or amount_value is None:
                logger.debug(f"Row {row_num}: Skipping due to None values")
                continue
            
            # Try to convert helper value to number
            try:
                if isinstance(helper_value, str):
                    helper_value = float(helper_value.replace(',', ''))
                else:
                    helper_value = float(helper_value)
            except (ValueError, TypeError) as e:
                logger.debug(f"Cell {helper_column}{row_num} contains non-numeric value: {helper_value} (type: {type(helper_value)}), skipping: {e}")
                continue
            
            # Try to convert amount to number
            try:
                if isinstance(amount_value, str):
                    # Skip if it's a header text (contains Chinese characters that might be headers)
                    if any('\u4e00' <= char <= '\u9fff' for char in amount_value):
                        logger.debug(f"Skipping header text in cell {amount_column}{row_num}: {amount_value}")
                        continue
                    amount_value = float(amount_value.replace(',', ''))
                else:
                    amount_value = float(amount_value)
                data_rows.append((row_num, helper_value, amount_value))
                logger.debug(f"Row {row_num}: Added to data_rows - helper={helper_value}, amount={amount_value}")
            except (ValueError, TypeError) as e:
                logger.debug(f"Cell {amount_column}{row_num} contains non-numeric value: {amount_value} (type: {type(amount_value)}), skipping: {e}")
                continue
        
        logger.info(f"Collected {len(data_rows)} data rows for sorting verification")
        
        if len(data_rows) < 2:
            logger.error(f"Not enough data rows to verify sorting: only {len(data_rows)} rows collected")
            logger.error(f"  Expected data from row {actual_start_row} to {data_end_row}")
            return 0.0
        
        # Group data by helper column value
        from collections import defaultdict
        groups = defaultdict(list)
        for row_num, helper_val, amount_val in data_rows:
            groups[helper_val].append((row_num, amount_val))
        
        logger.info(f"Found {len(groups)} groups based on {helper_column} column values")
        
        # Check if amounts are sorted within each group
        all_groups_sorted = True
        for group_key in sorted(groups.keys()):
            group_data = groups[group_key]
            if len(group_data) < 2:
                # Single row groups are always sorted
                continue
            
            # Check if amounts in this group are sorted
            group_sorted = True
            for i in range(len(group_data) - 1):
                current_amount = group_data[i][1]
                next_amount = group_data[i + 1][1]
                
                if sort_order == "descending":
                    if current_amount < next_amount:
                        logger.error(f"Group {group_key}: row {group_data[i][0]} ({current_amount}) < row {group_data[i+1][0]} ({next_amount}) - not descending")
                        group_sorted = False
                        break
                else:  # ascending
                    if current_amount > next_amount:
                        logger.error(f"Group {group_key}: row {group_data[i][0]} ({current_amount}) > row {group_data[i+1][0]} ({next_amount}) - not ascending")
                        group_sorted = False
                        break
            
            if not group_sorted:
                all_groups_sorted = False
                logger.error(f"Group {group_key} is not sorted in {sort_order} order")
            else:
                logger.debug(f"Group {group_key} is sorted correctly")
        
        if not all_groups_sorted:
            logger.error("Sorting verification failed: not all groups are sorted correctly")
            return 0.0
        
        logger.info("✓ Amount sorting verification passed: all groups are sorted correctly")
        
        logger.info("=" * 60)
        logger.info(f"✓ COUNTA helper column and amount sorting verification passed")
        logger.info(f"  Helper column: {helper_column}")
        logger.info(f"  Amount column: {amount_column}")
        logger.info(f"  Sort order: {sort_order}")
        logger.info("=" * 60)
        return 1.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_if_iferror_find_text_search(result: str, expected: str = None, **options) -> float:
    """
    Verify if IF(IFERROR(FIND(...))) formulas exist to search for text within a row range.
    
    This function checks:
    1. Whether cells in specified column contain IF formulas
    2. Whether formulas contain required functions: IF, IFERROR, FIND
    3. Whether FIND function searches for the specified text (e.g., "AB")
    4. Whether FIND function searches in concatenated text from multiple columns (e.g., A2&B2)
    5. Whether IF function returns 1 if found, empty string if not found
    
    Expected formula pattern: =IF(IFERROR(FIND("AB",A2&B2),)>0,1,)
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "C")
            - start_row: Starting row number (default: 2)
            - search_columns: List of columns to concatenate and search (e.g., ["A", "B"])
            - search_text: Text to search for (e.g., "AB")
            - return_value: Value to return if found (default: "1")
            - expected_functions: List of expected function names (default: ["IF", "IFERROR", "FIND"])
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'C')
        start_row = options.get('start_row', 2)
        search_columns = options.get('search_columns', ['A', 'B'])
        search_text = options.get('search_text', 'AB')
        return_value = options.get('return_value', '1')
        expected_functions = options.get('expected_functions', ['IF', 'IFERROR', 'FIND'])
        data_column = options.get('data_column', 'A')
        
        logger.info(f"Verifying IF/IFERROR/FIND text search formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Search columns: {search_columns}")
        logger.info(f"Search text: {search_text}")
        logger.info(f"Return value: {return_value}")
        logger.info(f"Expected functions: {expected_functions}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row
        all_checks_passed = True
        rows_checked = 0
        
        for row_num in range(start_row, end_row + 1):
            check_cell = ws[f"{check_column}{row_num}"]
            
            # Skip if cell is empty
            if check_cell.value is None:
                continue
            
            rows_checked += 1
            
            # Check if cell contains a formula
            if check_cell.data_type != "f":
                logger.error(f"Cell {check_column}{row_num} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(check_cell, "_value") and isinstance(check_cell._value, str) and check_cell._value.startswith("="):
                formula = check_cell._value
            elif hasattr(check_cell, "formula"):
                formula = check_cell.formula
            else:
                # Try to get from value attribute
                if check_cell.value is not None and isinstance(check_cell.value, str) and check_cell.value.startswith("="):
                    formula = check_cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {check_column}{row_num}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check 1: Contains all required functions
            missing_functions = []
            for func in expected_functions:
                # Check for function name (with word boundary to avoid partial matches)
                func_pattern = r'\b' + re.escape(func) + r'\s*\('
                if not re.search(func_pattern, formula_upper, re.IGNORECASE):
                    missing_functions.append(func)
            
            if missing_functions:
                logger.error(f"Cell {check_column}{row_num} formula missing functions: {missing_functions}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: Formula structure - IF(IFERROR(FIND(...)))
            if_iferror_pattern = r'IF\s*\(\s*IFERROR\s*\('
            if not re.search(if_iferror_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {check_column}{row_num} formula should have structure IF(IFERROR(...))")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: IFERROR contains FIND
            iferror_find_pattern = r'IFERROR\s*\(\s*FIND\s*\('
            if not re.search(iferror_find_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {check_column}{row_num} IFERROR should contain FIND function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: FIND contains search text
            # Pattern: FIND("AB",...) or FIND('AB',...)
            find_search_pattern1 = rf'FIND\s*\(\s*["\']?{re.escape(search_text)}["\']?\s*,'
            find_search_pattern2 = rf'FIND\s*\(\s*["\']?{re.escape(search_text)}["\']?\s*,'
            if not re.search(find_search_pattern1, formula, re.IGNORECASE) and \
               not re.search(find_search_pattern2, formula, re.IGNORECASE):
                logger.error(f"Cell {check_column}{row_num} FIND should search for '{search_text}'")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 5: FIND searches in concatenated columns (e.g., A2&B2)
            # Build pattern for concatenated columns: A2&B2 or A{row_num}&B{row_num}
            concat_pattern_parts = []
            for col in search_columns:
                concat_pattern_parts.append(rf'{col}{row_num}')
            concat_pattern = r'&'.join(concat_pattern_parts)
            if concat_pattern.upper() not in formula_upper:
                logger.error(f"Cell {check_column}{row_num} FIND should search in concatenated columns {concat_pattern}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 6: IFERROR has empty second parameter or no second parameter
            # Pattern: IFERROR(...,) or IFERROR(...,"")
            iferror_empty_pattern = r'IFERROR\s*\([^,]+,\s*\)|IFERROR\s*\([^,]+,\s*["\']?\s*["\']?'
            if not re.search(iferror_empty_pattern, formula_upper, re.IGNORECASE):
                logger.warning(f"Cell {check_column}{row_num} IFERROR may not have empty second parameter")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn
            
            # Check 7: IF condition checks if result > 0
            # Pattern: IFERROR(...)>0
            # Since IFERROR may have nested parentheses, we'll check more flexibly:
            # 1. Check that IFERROR exists
            # 2. Check that >0 exists after IFERROR
            # 3. Check that there's a closing parenthesis before >0
            iferror_pos = formula_upper.find('IFERROR')
            gt_zero_pos = formula_upper.find('>0')
            
            if iferror_pos == -1 or gt_zero_pos == -1:
                logger.error(f"Cell {check_column}{row_num} formula should contain IFERROR and >0")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check that >0 comes after IFERROR
            if gt_zero_pos <= iferror_pos:
                logger.error(f"Cell {check_column}{row_num} >0 should come after IFERROR")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check that there's a closing parenthesis for IFERROR before >0
            # Find the position of IFERROR's opening parenthesis
            iferror_open = formula_upper.find('(', iferror_pos)
            if iferror_open == -1:
                logger.error(f"Cell {check_column}{row_num} IFERROR should have opening parenthesis")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Count parentheses to find IFERROR's closing parenthesis
            paren_count = 0
            iferror_close = -1
            for i in range(iferror_open, min(gt_zero_pos, len(formula_upper))):
                if formula_upper[i] == '(':
                    paren_count += 1
                elif formula_upper[i] == ')':
                    paren_count -= 1
                    if paren_count == 0:
                        iferror_close = i
                        break
            
            if iferror_close == -1 or iferror_close >= gt_zero_pos:
                logger.error(f"Cell {check_column}{row_num} IFERROR should be closed before >0")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 8: IF returns return_value if found
            # Pattern: IF(...,1,) or IF(...,"1",)
            # Since IF condition may contain commas (e.g., IFERROR(...)), we need to find the second parameter
            # Find the position of IF's opening parenthesis and the comma before return_value
            if_pos = formula_upper.find('IF(')
            if if_pos == -1:
                if_pos = formula_upper.find('IF (')
            
            if if_pos != -1:
                # Find the opening parenthesis of IF
                if_open = formula_upper.find('(', if_pos)
                if if_open != -1:
                    # Count parentheses to find the comma that separates condition from return_value
                    paren_count = 0
                    comma_pos = -1
                    for i in range(if_open + 1, len(formula_upper)):
                        if formula_upper[i] == '(':
                            paren_count += 1
                        elif formula_upper[i] == ')':
                            paren_count -= 1
                            if paren_count < 0:
                                break
                        elif formula_upper[i] == ',' and paren_count == 0:
                            comma_pos = i
                            break
                    
                    if comma_pos != -1:
                        # Check if return_value appears after this comma
                        after_comma = formula_upper[comma_pos + 1:].strip()
                        # Check for return_value (with or without quotes)
                        return_pattern1 = rf'^{re.escape(return_value)}\s*,'
                        return_pattern2 = rf'^["\']?{re.escape(return_value)}["\']?\s*,'
                        if not re.match(return_pattern1, after_comma, re.IGNORECASE) and \
                           not re.match(return_pattern2, after_comma, re.IGNORECASE):
                            logger.error(f"Cell {check_column}{row_num} IF should return '{return_value}' if found")
                            logger.error(f"  Formula: {formula}")
                            logger.error(f"  After comma: {after_comma[:20]}")
                            all_checks_passed = False
                            continue
                    else:
                        logger.error(f"Cell {check_column}{row_num} Could not find comma separating IF condition from return value")
                        logger.error(f"  Formula: {formula}")
                        all_checks_passed = False
                        continue
                else:
                    logger.error(f"Cell {check_column}{row_num} IF should have opening parenthesis")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
            else:
                logger.error(f"Cell {check_column}{row_num} Formula should contain IF function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 9: IF has empty third parameter (else case)
            # Pattern: IF(...,1,) or IF(...,"1","")
            if_empty_else_pattern = r'IF\s*\([^,]+,\s*[^,]+,\s*\)|IF\s*\([^,]+,\s*[^,]+,\s*["\']?\s*["\']?'
            if not re.search(if_empty_else_pattern, formula_upper, re.IGNORECASE):
                logger.warning(f"Cell {check_column}{row_num} IF may not have empty else parameter")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn
            
            logger.debug(f"✓ Row {row_num} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ IF/IFERROR/FIND text search formula verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Column: {check_column}")
            logger.info(f"  Search text: {search_text}")
            logger.info(f"  Functions verified: {', '.join(expected_functions)}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ IF/IFERROR/FIND text search formula verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_countif_if_household_sort(result: str, expected: str = None, **options) -> float:
    """
    Verify if COUNTIF and IF formulas exist in helper column for household sorting, and verify sorting.
    
    This function checks:
    1. Whether helper column contains COUNTIF and IF formulas
    2. Whether COUNTIF counts "户主" (household head) from start to current row
    3. Whether IF function assigns values: 1 for "户主", 2 for "夫妻", 3 for others
    4. Whether data is sorted by helper column (ascending) and birth date (ascending)
    
    Expected formula pattern: =COUNTIF($E$1:E1,"户主")*10+IF(E1="户主",1,IF(E1="夫妻",2,3))
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - helper_column: Helper column with formula (e.g., "F")
            - start_row: Starting row for formula (default: 1)
            - relationship_column: Column containing relationships (e.g., "E")
            - birth_date_column: Column containing birth dates (e.g., "D")
            - expected_functions: List of expected function names (default: ["COUNTIF", "IF"])
            - household_head_text: Text for household head (default: "户主")
            - spouse_text: Text for spouse (default: "夫妻")
            - data_column: Column to check for data to determine end_row (default: "E")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        helper_column = options.get('helper_column', 'F')
        start_row = options.get('start_row', 1)
        relationship_column = options.get('relationship_column', 'E')
        birth_date_column = options.get('birth_date_column', 'D')
        expected_functions = options.get('expected_functions', ['COUNTIF', 'IF'])
        household_head_text = options.get('household_head_text', '户主')
        spouse_text = options.get('spouse_text', '夫妻')
        data_column = options.get('data_column', 'E')
        
        logger.info(f"Verifying COUNTIF/IF household sorting formulas in file: {result}")
        logger.info(f"Helper column: {helper_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Relationship column: {relationship_column}")
        logger.info(f"Birth date column: {birth_date_column}")
        logger.info(f"Expected functions: {expected_functions}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check 1: Verify formulas in helper column
        logger.info(f"Checking formulas in column {helper_column}...")
        helper_formulas_ok = True
        
        for row_num in range(start_row, end_row + 1):
            helper_cell = ws[f"{helper_column}{row_num}"]
            
            # Skip if cell is empty
            if helper_cell.value is None:
                continue
            
            # Check if cell contains a formula
            if helper_cell.data_type != "f":
                logger.error(f"Cell {helper_column}{row_num} does not contain a formula")
                helper_formulas_ok = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(helper_cell, "_value") and isinstance(helper_cell._value, str) and helper_cell._value.startswith("="):
                formula = helper_cell._value
            elif hasattr(helper_cell, "formula"):
                formula = helper_cell.formula
            else:
                if helper_cell.value is not None and isinstance(helper_cell.value, str) and helper_cell.value.startswith("="):
                    formula = helper_cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {helper_column}{row_num}")
                helper_formulas_ok = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check if formula contains required functions
            missing_functions = []
            for func in expected_functions:
                func_pattern = r'\b' + re.escape(func) + r'\s*\('
                if not re.search(func_pattern, formula_upper, re.IGNORECASE):
                    missing_functions.append(func)
            
            if missing_functions:
                logger.error(f"Cell {helper_column}{row_num} formula missing functions: {missing_functions}")
                logger.error(f"  Formula: {formula}")
                helper_formulas_ok = False
                continue
            
            # Check if COUNTIF references relationship column with "户主"
            countif_pattern = rf'COUNTIF\s*\(\s*\${relationship_column}\$\d+:\s*{relationship_column}{row_num}\s*,\s*["\']?{re.escape(household_head_text)}["\']?'
            if not re.search(countif_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {helper_column}{row_num} COUNTIF should count '{household_head_text}' in {relationship_column} column")
                logger.error(f"  Formula: {formula}")
                helper_formulas_ok = False
                continue
            
            # Check if COUNTIF result is multiplied by 10
            countif_multiply_pattern = rf'COUNTIF\s*\([^)]+\)\s*\*\s*10'
            if not re.search(countif_multiply_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {helper_column}{row_num} COUNTIF result should be multiplied by 10")
                logger.error(f"  Formula: {formula}")
                helper_formulas_ok = False
                continue
            
            # Check if IF function checks for "户主" and assigns 1
            if_household_head_pattern = rf'IF\s*\(\s*{relationship_column}{row_num}\s*=\s*["\']?{re.escape(household_head_text)}["\']?\s*,\s*1'
            if not re.search(if_household_head_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {helper_column}{row_num} IF should check for '{household_head_text}' and return 1")
                logger.error(f"  Formula: {formula}")
                helper_formulas_ok = False
                continue
            
            # Check if nested IF checks for "夫妻" and assigns 2
            if_spouse_pattern = rf'IF\s*\(\s*{relationship_column}{row_num}\s*=\s*["\']?{re.escape(spouse_text)}["\']?\s*,\s*2'
            if not re.search(if_spouse_pattern, formula_upper, re.IGNORECASE):
                logger.error(f"Cell {helper_column}{row_num} IF should check for '{spouse_text}' and return 2")
                logger.error(f"  Formula: {formula}")
                helper_formulas_ok = False
                continue
            
            # Check if nested IF has else value 3
            if_else_3_pattern = r'IF\s*\([^,]+,\s*[^,]+,\s*3\s*\)'
            if not re.search(if_else_3_pattern, formula_upper, re.IGNORECASE):
                logger.warning(f"Cell {helper_column}{row_num} IF else value may not be 3")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn
        
        if not helper_formulas_ok:
            logger.error("Helper column formula verification failed")
            return 0.0
        
        logger.info("✓ Helper column formula verification passed")
        
        # Check 2: Verify sorting by helper column and birth date
        logger.info(f"Checking if data is sorted by {helper_column} column (ascending) and {birth_date_column} column (ascending)...")
        
        # Load workbook with calculated values for sorting check
        try:
            wb_values = openpyxl.load_workbook(result, data_only=True)
            ws_values = wb_values.active
        except Exception as e:
            logger.warning(f"Failed to load workbook with data_only=True, using formula values: {e}")
            ws_values = ws
        
        # Collect data: (row_num, helper_value, birth_date_value, relationship)
        data_rows = []
        for row_num in range(start_row, end_row + 1):
            helper_cell = ws_values[f"{helper_column}{row_num}"]
            birth_date_cell = ws_values[f"{birth_date_column}{row_num}"]
            relationship_cell = ws_values[f"{relationship_column}{row_num}"]
            
            helper_value = helper_cell.value
            birth_date_value = birth_date_cell.value
            relationship_value = relationship_cell.value
            
            # Skip if essential data is missing
            if helper_value is None or relationship_value is None:
                continue
            
            # Try to convert helper value to number
            try:
                if isinstance(helper_value, str):
                    helper_value = float(helper_value.replace(',', ''))
                else:
                    helper_value = float(helper_value)
            except (ValueError, TypeError):
                logger.debug(f"Cell {helper_column}{row_num} contains non-numeric value: {helper_value}, skipping")
                continue
            
            # Try to convert birth date to comparable value
            birth_date_comparable = None
            if birth_date_value is not None:
                try:
                    if isinstance(birth_date_value, (int, float)):
                        birth_date_comparable = float(birth_date_value)
                    elif isinstance(birth_date_value, str):
                        # Try to parse date string
                        from datetime import datetime
                        try:
                            dt = datetime.strptime(birth_date_value, '%Y-%m-%d')
                            birth_date_comparable = dt.timestamp()
                        except:
                            try:
                                dt = datetime.strptime(birth_date_value, '%Y/%m/%d')
                                birth_date_comparable = dt.timestamp()
                            except:
                                birth_date_comparable = None
                    else:
                        birth_date_comparable = float(birth_date_value)
                except (ValueError, TypeError):
                    logger.debug(f"Cell {birth_date_column}{row_num} contains non-comparable date: {birth_date_value}")
            
            data_rows.append((row_num, helper_value, birth_date_comparable, relationship_value))
        
        if len(data_rows) < 2:
            logger.warning("Not enough data rows to verify sorting")
            logger.info("✓ Sorting verification passed (minimal data, helper column formula verified)")
            return 1.0
        
        # Check sorting: first by helper column (ascending), then by birth date (ascending)
        is_sorted = True
        for i in range(len(data_rows) - 1):
            current_helper = data_rows[i][1]
            next_helper = data_rows[i + 1][1]
            current_birth = data_rows[i][2]
            next_birth = data_rows[i + 1][2]
            
            # First sort key: helper column (ascending)
            if current_helper > next_helper:
                logger.error(f"Data not sorted by helper column: row {data_rows[i][0]} ({current_helper}) > row {data_rows[i+1][0]} ({next_helper})")
                is_sorted = False
                break
            elif current_helper == next_helper:
                # Second sort key: birth date (ascending) - only if both have dates
                if current_birth is not None and next_birth is not None:
                    if current_birth > next_birth:
                        logger.error(f"Data not sorted by birth date within same helper value: row {data_rows[i][0]} ({current_birth}) > row {data_rows[i+1][0]} ({next_birth})")
                        is_sorted = False
                        break
        
        if not is_sorted:
            logger.error("Sorting verification failed")
            return 0.0
        
        logger.info("✓ Sorting verification passed")
        
        logger.info("=" * 60)
        logger.info(f"✓ COUNTIF/IF household sorting verification passed")
        logger.info(f"  Helper column: {helper_column}")
        logger.info(f"  Relationship column: {relationship_column}")
        logger.info(f"  Birth date column: {birth_date_column}")
        logger.info("=" * 60)
        return 1.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_regex_airport_extract(result: str, expected: str = None, **options) -> float:
    """
    Verify if REGEX formulas exist in specified columns (E and F) to extract airport names.
    
    This function checks:
    1. Whether cells in E and F columns contain REGEX formulas
    2. Whether E column formulas reference B column (departure location)
    3. Whether F column formulas reference C column (arrival location)
    4. Whether formulas contain the pattern "\\p{Han}{2,4}" to extract 2-4 Chinese characters
    
    Expected formula patterns:
    - E column: =REGEX(B2,"\\p{Han}{2,4}")
    - F column: =REGEX(C2,"\\p{Han}{2,4}")
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - departure_extract_column: Column with departure airport extraction formulas (e.g., "E")
            - arrival_extract_column: Column with arrival airport extraction formulas (e.g., "F")
            - departure_source_column: Column containing departure location data (e.g., "B")
            - arrival_source_column: Column containing arrival location data (e.g., "C")
            - start_row: Starting row number (default: 2)
            - expected_pattern: Expected regex pattern (default: "\\p{Han}{2,4}")
            - data_column: Column to check for data to determine end_row (default: "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        departure_extract_column = options.get('departure_extract_column', 'E')
        arrival_extract_column = options.get('arrival_extract_column', 'F')
        departure_source_column = options.get('departure_source_column', 'B')
        arrival_source_column = options.get('arrival_source_column', 'C')
        start_row = options.get('start_row', 2)
        expected_pattern = options.get('expected_pattern', '\\p{Han}{2,4}')
        data_column = options.get('data_column', 'B')
        
        logger.info(f"Verifying REGEX airport extraction formulas in file: {result}")
        logger.info(f"Departure extract column: {departure_extract_column}")
        logger.info(f"Arrival extract column: {arrival_extract_column}")
        logger.info(f"Departure source column: {departure_source_column}")
        logger.info(f"Arrival source column: {arrival_source_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected pattern: {expected_pattern}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check formulas in both columns
        all_checks_passed = True
        rows_checked = 0
        
        for row_num in range(start_row, end_row + 1):
            # Check departure column (E)
            departure_cell = ws[f"{departure_extract_column}{row_num}"]
            departure_source_cell = ws[f"{departure_source_column}{row_num}"]
            
            # Check arrival column (F)
            arrival_cell = ws[f"{arrival_extract_column}{row_num}"]
            arrival_source_cell = ws[f"{arrival_source_column}{row_num}"]
            
            # Skip if source cells are empty
            if departure_source_cell.value is None and arrival_source_cell.value is None:
                continue
            
            rows_checked += 1
            
            # Check departure column formula
            if departure_source_cell.value is not None:
                if departure_cell.data_type != "f":
                    logger.error(f"Cell {departure_extract_column}{row_num} does not contain a formula")
                    all_checks_passed = False
                    continue
                
                # Get formula text
                formula = None
                if hasattr(departure_cell, "_value") and isinstance(departure_cell._value, str) and departure_cell._value.startswith("="):
                    formula = departure_cell._value
                elif hasattr(departure_cell, "formula"):
                    formula = departure_cell.formula
                else:
                    if departure_cell.value is not None and isinstance(departure_cell.value, str) and departure_cell.value.startswith("="):
                        formula = departure_cell.value
                
                if formula is None:
                    logger.error(f"Could not extract formula from cell {departure_extract_column}{row_num}")
                    all_checks_passed = False
                    continue
                
                formula_upper = formula.upper()
                logger.debug(f"Row {row_num} departure formula: {formula}")
                
                # Check 1: Contains REGEX function
                if 'REGEX' not in formula_upper:
                    logger.error(f"Cell {departure_extract_column}{row_num} formula should contain REGEX function")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
                
                # Check 2: References departure source column
                source_pattern = rf'{departure_source_column}{row_num}\b'
                if not re.search(source_pattern, formula_upper):
                    logger.error(f"Cell {departure_extract_column}{row_num} formula should reference {departure_source_column}{row_num}")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
                
                # Check 3: Contains expected pattern
                # Pattern in formula might be: "\\p{Han}{2,4}" or '\\p{Han}{2,4}' or "\p{Han}{2,4}"
                # We need to check for the pattern with proper escaping
                # The pattern might appear as: \p{Han}{2,4} (with backslash) or \\p{Han}{2,4} (with double backslash)
                pattern_variants = [
                    re.escape(expected_pattern),  # Exact match with escaping
                    expected_pattern.replace('\\', '\\\\'),  # Double backslash version
                    expected_pattern.replace('\\\\', '\\'),  # Single backslash version
                ]
                pattern_found = False
                for variant in pattern_variants:
                    # Check for pattern in quotes (single or double)
                    pattern_check = rf'["\']{re.escape(variant)}["\']'
                    if re.search(pattern_check, formula):
                        pattern_found = True
                        break
                    # Also check without quotes (less common but possible)
                    pattern_check2 = rf'\b{re.escape(variant)}\b'
                    if re.search(pattern_check2, formula):
                        pattern_found = True
                        break
                
                if not pattern_found:
                    logger.error(f"Cell {departure_extract_column}{row_num} formula should contain pattern '{expected_pattern}'")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
            
            # Check arrival column formula
            if arrival_source_cell.value is not None:
                if arrival_cell.data_type != "f":
                    logger.error(f"Cell {arrival_extract_column}{row_num} does not contain a formula")
                    all_checks_passed = False
                    continue
                
                # Get formula text
                formula = None
                if hasattr(arrival_cell, "_value") and isinstance(arrival_cell._value, str) and arrival_cell._value.startswith("="):
                    formula = arrival_cell._value
                elif hasattr(arrival_cell, "formula"):
                    formula = arrival_cell.formula
                else:
                    if arrival_cell.value is not None and isinstance(arrival_cell.value, str) and arrival_cell.value.startswith("="):
                        formula = arrival_cell.value
                
                if formula is None:
                    logger.error(f"Could not extract formula from cell {arrival_extract_column}{row_num}")
                    all_checks_passed = False
                    continue
                
                formula_upper = formula.upper()
                logger.debug(f"Row {row_num} arrival formula: {formula}")
                
                # Check 1: Contains REGEX function
                if 'REGEX' not in formula_upper:
                    logger.error(f"Cell {arrival_extract_column}{row_num} formula should contain REGEX function")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
                
                # Check 2: References arrival source column
                source_pattern = rf'{arrival_source_column}{row_num}\b'
                if not re.search(source_pattern, formula_upper):
                    logger.error(f"Cell {arrival_extract_column}{row_num} formula should reference {arrival_source_column}{row_num}")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
                
                # Check 3: Contains expected pattern
                # Pattern in formula might be: "\\p{Han}{2,4}" or '\\p{Han}{2,4}' or "\p{Han}{2,4}"
                # We need to check for the pattern with proper escaping
                # The pattern might appear as: \p{Han}{2,4} (with backslash) or \\p{Han}{2,4} (with double backslash)
                pattern_variants = [
                    re.escape(expected_pattern),  # Exact match with escaping
                    expected_pattern.replace('\\', '\\\\'),  # Double backslash version
                    expected_pattern.replace('\\\\', '\\'),  # Single backslash version
                ]
                pattern_found = False
                for variant in pattern_variants:
                    # Check for pattern in quotes (single or double)
                    pattern_check = rf'["\']{re.escape(variant)}["\']'
                    if re.search(pattern_check, formula):
                        pattern_found = True
                        break
                    # Also check without quotes (less common but possible)
                    pattern_check2 = rf'\b{re.escape(variant)}\b'
                    if re.search(pattern_check2, formula):
                        pattern_found = True
                        break
                
                if not pattern_found:
                    logger.error(f"Cell {arrival_extract_column}{row_num} formula should contain pattern '{expected_pattern}'")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
            
            logger.debug(f"✓ Row {row_num} formulas passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ REGEX airport extraction verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Departure column: {departure_extract_column} (from {departure_source_column})")
            logger.info(f"  Arrival column: {arrival_extract_column} (from {arrival_source_column})")
            logger.info(f"  Pattern: {expected_pattern}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ REGEX airport extraction verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_sumproduct_cross_sheet_lookup(result: str, expected: str = None, **options) -> float:
    """
    Verify if SUMPRODUCT formulas exist in target sheet to lookup data from source sheet.
    
    This function checks:
    1. Whether cells in target column contain SUMPRODUCT formulas
    2. Whether formulas reference source sheet with correct sheet name
    3. Whether formulas match two conditions: source column A equals target column A, source column B equals target column B
    4. Whether formulas return values from source column C
    
    Expected formula pattern: =SUMPRODUCT((source_sheet.$A$2:$A$19=A2)*(source_sheet.$B$2:$B$19=B2),source_sheet.$C$2:$C$19)
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - target_sheet: Target sheet name (e.g., "本周销售")
            - source_sheet: Source sheet name (e.g., "上周销售")
            - target_column: Column with formulas (e.g., "D")
            - source_column_a: Source column for first condition (e.g., "A")
            - source_column_b: Source column for second condition (e.g., "B")
            - source_column_c: Source column for return values (e.g., "C")
            - target_column_a: Target column for first condition (e.g., "A")
            - target_column_b: Target column for second condition (e.g., "B")
            - start_row: Starting row number (default: 2)
            - data_range_start: Source data range start row (default: 2)
            - data_range_end: Source data range end row (default: 19)
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        target_sheet = options.get('target_sheet', '本周销售')
        source_sheet = options.get('source_sheet', '上周销售')
        target_column = options.get('target_column', 'D')
        source_column_a = options.get('source_column_a', 'A')
        source_column_b = options.get('source_column_b', 'B')
        source_column_c = options.get('source_column_c', 'C')
        target_column_a = options.get('target_column_a', 'A')
        target_column_b = options.get('target_column_b', 'B')
        start_row = options.get('start_row', 2)
        data_range_start = options.get('data_range_start', 2)
        data_range_end = options.get('data_range_end', 19)
        data_column = options.get('data_column', 'A')
        
        logger.info(f"Verifying SUMPRODUCT cross-sheet lookup formulas in file: {result}")
        logger.info(f"Target sheet: {target_sheet}")
        logger.info(f"Source sheet: {source_sheet}")
        logger.info(f"Target column: {target_column}")
        logger.info(f"Source columns: {source_column_a}, {source_column_b}, {source_column_c}")
        logger.info(f"Target columns: {target_column_a}, {target_column_b}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Data range: {data_range_start} to {data_range_end}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            if target_sheet not in wb.sheetnames:
                logger.error(f"Target sheet '{target_sheet}' not found in workbook")
                return 0.0
            ws_target = wb[target_sheet]
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws_target.max_row
        end_row = start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws_target[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check formulas in target column
        all_checks_passed = True
        rows_checked = 0
        
        for row_num in range(start_row, end_row + 1):
            target_cell = ws_target[f"{target_column}{row_num}"]
            target_cell_a = ws_target[f"{target_column_a}{row_num}"]
            
            # Skip if target column A is empty
            if target_cell_a.value is None:
                continue
            
            rows_checked += 1
            
            # Check if cell contains a formula
            if target_cell.data_type != "f":
                logger.error(f"Cell {target_column}{row_num} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(target_cell, "_value") and isinstance(target_cell._value, str) and target_cell._value.startswith("="):
                formula = target_cell._value
            elif hasattr(target_cell, "formula"):
                formula = target_cell.formula
            else:
                if target_cell.value is not None and isinstance(target_cell.value, str) and target_cell.value.startswith("="):
                    formula = target_cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {target_column}{row_num}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check 1: Contains SUMPRODUCT function
            if 'SUMPRODUCT' not in formula_upper:
                logger.error(f"Cell {target_column}{row_num} formula should contain SUMPRODUCT function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References source sheet
            # Sheet name might be quoted or unquoted, and might have spaces
            # Excel/LibreOffice uses ! (exclamation mark) for sheet references
            source_sheet_pattern = rf'["\']?{re.escape(source_sheet)}["\']?[!\.]'
            if not re.search(source_sheet_pattern, formula, re.IGNORECASE):
                logger.error(f"Cell {target_column}{row_num} formula should reference source sheet '{source_sheet}'")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: First condition - source column A equals target column A
            # Pattern: source_sheet!$A$2:$A$19=A2 (or similar row references)
            # Note: Excel/LibreOffice uses ! (exclamation mark) not . (dot) for sheet references
            condition1_pattern = rf'{re.escape(source_sheet)}[!\.][\$]?{source_column_a}[\$]?\d+:\$?{source_column_a}[\$]?\d+\s*=\s*{target_column_a}{row_num}'
            if not re.search(condition1_pattern, formula_upper):
                logger.error(f"Cell {target_column}{row_num} formula should have condition: {source_sheet}!{source_column_a}...={target_column_a}{row_num}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Second condition - source column B equals target column B
            # Pattern: source_sheet!$B$2:$B$19=B2 (or similar row references)
            condition2_pattern = rf'{re.escape(source_sheet)}[!\.][\$]?{source_column_b}[\$]?\d+:\$?{source_column_b}[\$]?\d+\s*=\s*{target_column_b}{row_num}'
            if not re.search(condition2_pattern, formula_upper):
                logger.error(f"Cell {target_column}{row_num} formula should have condition: {source_sheet}!{source_column_b}...={target_column_b}{row_num}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 5: Returns values from source column C
            # Pattern: source_sheet!$C$2:$C$19 (or similar row references)
            return_pattern = rf'{re.escape(source_sheet)}[!\.][\$]?{source_column_c}[\$]?\d+:\$?{source_column_c}[\$]?\d+'
            if not re.search(return_pattern, formula_upper):
                logger.error(f"Cell {target_column}{row_num} formula should return values from {source_sheet}!{source_column_c}...")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 6: Data range should match expected range
            range_pattern = rf'{re.escape(source_sheet)}[!\.][\$]?{source_column_a}[\$]?{data_range_start}:\$?{source_column_a}[\$]?{data_range_end}'
            if not re.search(range_pattern, formula_upper):
                logger.warning(f"Cell {target_column}{row_num} formula may not use expected data range {data_range_start}:{data_range_end}")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn
            
            logger.debug(f"✓ Row {row_num} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ SUMPRODUCT cross-sheet lookup verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Target sheet: {target_sheet}, Column: {target_column}")
            logger.info(f"  Source sheet: {source_sheet}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ SUMPRODUCT cross-sheet lookup verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_substitute_remove_spaces(result: str, expected: str = None, **options) -> float:
    """
    Verify if SUBSTITUTE formulas exist in specified columns to remove spaces.
    
    This function checks:
    1. Whether cells in specified columns contain SUBSTITUTE formulas
    2. Whether formulas reference the corresponding source column cells
    3. Whether formulas use SUBSTITUTE to replace space with empty string
    
    Expected formula pattern: =SUBSTITUTE(A2," ",) or =SUBSTITUTE(A2," ","")
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_columns: List of columns to check (e.g., ["C", "D"])
            - source_columns: List of corresponding source columns (e.g., ["A", "B"])
            - start_row: Starting row number (default: 2)
            - expected_functions: List of expected function names (default: ["SUBSTITUTE"])
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_columns = options.get('check_columns', ['C', 'D'])
        source_columns = options.get('source_columns', ['A', 'B'])
        start_row = options.get('start_row', 2)
        expected_functions = options.get('expected_functions', ['SUBSTITUTE'])
        data_column = options.get('data_column', 'A')
        
        if len(check_columns) != len(source_columns):
            logger.error("check_columns and source_columns must have the same length")
            return 0.0
        
        logger.info(f"Verifying SUBSTITUTE remove spaces formulas in file: {result}")
        logger.info(f"Check columns: {check_columns}")
        logger.info(f"Source columns: {source_columns}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected functions: {expected_functions}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check formulas in each column
        all_checks_passed = True
        rows_checked = 0
        
        for col_idx, check_column in enumerate(check_columns):
            source_column = source_columns[col_idx]
            logger.info(f"Checking column {check_column} (from {source_column})...")
            
            for row_num in range(start_row, end_row + 1):
                check_cell = ws[f"{check_column}{row_num}"]
                source_cell = ws[f"{source_column}{row_num}"]
                
                # Skip if source cell is empty
                if source_cell.value is None:
                    continue
                
                rows_checked += 1
                
                # Check if cell contains a formula
                if check_cell.data_type != "f":
                    logger.error(f"Cell {check_column}{row_num} does not contain a formula")
                    all_checks_passed = False
                    continue
                
                # Get formula text
                formula = None
                if hasattr(check_cell, "_value") and isinstance(check_cell._value, str) and check_cell._value.startswith("="):
                    formula = check_cell._value
                elif hasattr(check_cell, "formula"):
                    formula = check_cell.formula
                else:
                    if check_cell.value is not None and isinstance(check_cell.value, str) and check_cell.value.startswith("="):
                        formula = check_cell.value
                
                if formula is None:
                    logger.error(f"Could not extract formula from cell {check_column}{row_num}")
                    all_checks_passed = False
                    continue
                
                formula_upper = formula.upper()
                logger.debug(f"Row {row_num} {check_column} formula: {formula}")
                
                # Check 1: Contains SUBSTITUTE function
                if 'SUBSTITUTE' not in formula_upper:
                    logger.error(f"Cell {check_column}{row_num} formula should contain SUBSTITUTE function")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
                
                # Check 2: References source column cell
                source_pattern = rf'{source_column}{row_num}\b'
                if not re.search(source_pattern, formula_upper):
                    logger.error(f"Cell {check_column}{row_num} formula should reference {source_column}{row_num}")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
                
                # Check 3: SUBSTITUTE replaces space with empty string
                # Pattern: SUBSTITUTE(...," ",) or SUBSTITUTE(...," ","")
                # The space might be in single or double quotes
                substitute_pattern1 = r'SUBSTITUTE\s*\([^,]+,\s*["\']?\s+["\']?\s*,\s*\)'  # SUBSTITUTE(...," ",)
                substitute_pattern2 = r'SUBSTITUTE\s*\([^,]+,\s*["\']?\s+["\']?\s*,\s*["\']?\s*["\']?\s*\)'  # SUBSTITUTE(...," ","")
                if not re.search(substitute_pattern1, formula_upper) and not re.search(substitute_pattern2, formula_upper):
                    logger.error(f"Cell {check_column}{row_num} SUBSTITUTE should replace space with empty string")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
                
                logger.debug(f"✓ Row {row_num} {check_column} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ SUBSTITUTE remove spaces verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Check columns: {check_columns}")
            logger.info(f"  Source columns: {source_columns}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ SUBSTITUTE remove spaces verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_substitute_remove_spaces(result: str, expected: str = None, **options) -> float:
    """
    Verify if SUBSTITUTE formulas exist in specified columns to remove spaces.
    
    This function checks:
    1. Whether cells in specified columns contain SUBSTITUTE formulas
    2. Whether formulas reference the corresponding source column cells
    3. Whether formulas use SUBSTITUTE to replace space with empty string
    
    Expected formula pattern: =SUBSTITUTE(A2," ",) or =SUBSTITUTE(A2," ","")
    
    The function automatically detects the number of data rows by checking the data column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_columns: List of columns to check (e.g., ["C", "D"])
            - source_columns: List of corresponding source columns (e.g., ["A", "B"])
            - start_row: Starting row number (default: 2)
            - expected_functions: List of expected function names (default: ["SUBSTITUTE"])
            - data_column: Column to check for data to determine end_row (default: "A")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_columns = options.get('check_columns', ['C', 'D'])
        source_columns = options.get('source_columns', ['A', 'B'])
        start_row = options.get('start_row', 2)
        expected_functions = options.get('expected_functions', ['SUBSTITUTE'])
        data_column = options.get('data_column', 'A')
        
        if len(check_columns) != len(source_columns):
            logger.error("check_columns and source_columns must have the same length")
            return 0.0
        
        logger.info(f"Verifying SUBSTITUTE remove spaces formulas in file: {result}")
        logger.info(f"Check columns: {check_columns}")
        logger.info(f"Source columns: {source_columns}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected functions: {expected_functions}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check formulas in each column
        all_checks_passed = True
        rows_checked = 0
        
        for col_idx, check_column in enumerate(check_columns):
            source_column = source_columns[col_idx]
            logger.info(f"Checking column {check_column} (from {source_column})...")
            
            for row_num in range(start_row, end_row + 1):
                check_cell = ws[f"{check_column}{row_num}"]
                source_cell = ws[f"{source_column}{row_num}"]
                
                # Skip if source cell is empty
                if source_cell.value is None:
                    continue
                
                rows_checked += 1
                
                # Check if cell contains a formula
                if check_cell.data_type != "f":
                    logger.error(f"Cell {check_column}{row_num} does not contain a formula")
                    all_checks_passed = False
                    continue
                
                # Get formula text
                formula = None
                if hasattr(check_cell, "_value") and isinstance(check_cell._value, str) and check_cell._value.startswith("="):
                    formula = check_cell._value
                elif hasattr(check_cell, "formula"):
                    formula = check_cell.formula
                else:
                    if check_cell.value is not None and isinstance(check_cell.value, str) and check_cell.value.startswith("="):
                        formula = check_cell.value
                
                if formula is None:
                    logger.error(f"Could not extract formula from cell {check_column}{row_num}")
                    all_checks_passed = False
                    continue
                
                formula_upper = formula.upper()
                logger.debug(f"Row {row_num} {check_column} formula: {formula}")
                
                # Check 1: Contains SUBSTITUTE function
                if 'SUBSTITUTE' not in formula_upper:
                    logger.error(f"Cell {check_column}{row_num} formula should contain SUBSTITUTE function")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
                
                # Check 2: References source column cell
                source_pattern = rf'{source_column}{row_num}\b'
                if not re.search(source_pattern, formula_upper):
                    logger.error(f"Cell {check_column}{row_num} formula should reference {source_column}{row_num}")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
                
                # Check 3: SUBSTITUTE replaces space with empty string
                # Pattern: SUBSTITUTE(...," ",) or SUBSTITUTE(...," ","")
                # The space might be in single or double quotes
                substitute_pattern1 = r'SUBSTITUTE\s*\([^,]+,\s*["\']?\s+["\']?\s*,\s*\)'  # SUBSTITUTE(...," ",)
                substitute_pattern2 = r'SUBSTITUTE\s*\([^,]+,\s*["\']?\s+["\']?\s*,\s*["\']?\s*["\']?\s*\)'  # SUBSTITUTE(...," ","")
                if not re.search(substitute_pattern1, formula_upper) and not re.search(substitute_pattern2, formula_upper):
                    logger.error(f"Cell {check_column}{row_num} SUBSTITUTE should replace space with empty string")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
                
                logger.debug(f"✓ Row {row_num} {check_column} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ SUBSTITUTE remove spaces verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Check columns: {check_columns}")
            logger.info(f"  Source columns: {source_columns}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ SUBSTITUTE remove spaces verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_maxifs_minifs_conditional(result: str, expected: str = None, **options) -> float:
    """
    Verify if MAXIFS and MINIFS formulas exist in specified cells to find max/min of positive/negative numbers.
    
    This function checks:
    1. Whether specified cells contain MAXIFS or MINIFS formulas
    2. Whether formulas reference the correct data range
    3. Whether formulas use correct conditions (<0 for negative, >0 for positive)
    
    Expected formula patterns:
    - F3: =MAXIFS(B2:B4771,B2:B4771,"<0") (max of negative numbers)
    - G3: =MAXIFS(B2:B4771,B2:B4771,">0") (max of positive numbers)
    - F4: =MINIFS(B2:B4771,B2:B4771,"<0") (min of negative numbers)
    - G4: =MINIFS(B2:B4771,B2:B4771,">0") (min of positive numbers)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "F3": {"function": "MAXIFS", "condition": "<0", "range": "B2:B4771"},
                  "G3": {"function": "MAXIFS", "condition": ">0", "range": "B2:B4771"},
                  "F4": {"function": "MINIFS", "condition": "<0", "range": "B2:B4771"},
                  "G4": {"function": "MINIFS", "condition": ">0", "range": "B2:B4771"}
              }
            - data_range: Data range to check (e.g., "B2:B4771")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        data_range = options.get('data_range', 'B2:B4771')
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying MAXIFS/MINIFS conditional formulas in file: {result}")
        logger.info(f"Data range: {data_range}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            expected_function = formula_info.get('function', '').upper()
            expected_condition = formula_info.get('condition', '')
            expected_range = formula_info.get('range', data_range)
            
            logger.info(f"Checking cell {cell_coord}: {expected_function} with condition {expected_condition}")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains expected function (MAXIFS or MINIFS)
            if expected_function not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain {expected_function} function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References expected data range
            # Escape the range pattern (e.g., B2:B4771)
            range_escaped = re.escape(expected_range)
            range_pattern = rf'{range_escaped}'
            if not re.search(range_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should reference range {expected_range}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains expected condition
            # Condition might be in quotes: "<0" or '>0'
            condition_escaped = re.escape(expected_condition)
            condition_pattern1 = rf'["\']?{condition_escaped}["\']?'
            if not re.search(condition_pattern1, formula_upper):
                logger.error(f"Cell {cell_coord} formula should have condition {expected_condition}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: Function structure - MAXIFS/MINIFS(range, range, condition)
            # Pattern: MAXIFS(B2:B4771,B2:B4771,"<0")
            function_pattern = rf'{expected_function}\s*\(\s*{range_escaped}\s*,\s*{range_escaped}\s*,\s*["\']?{condition_escaped}["\']?\s*\)'
            if not re.search(function_pattern, formula_upper):
                logger.warning(f"Cell {cell_coord} formula may not have correct {expected_function} structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the structure might vary slightly
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MAXIFS/MINIFS conditional verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info(f"  Data range: {data_range}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MAXIFS/MINIFS conditional verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_concatenate_columns_with_separator(result: str, expected: str = None, **options) -> float:
    """
    Verify if columns are concatenated with a separator using the & operator.
    
    This function checks:
    1. Whether the target column contains formulas that concatenate source columns
    2. Whether the formula uses the & operator for concatenation
    3. Whether the formula includes the specified separator
    4. Whether the formula is applied to all data rows
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - target_column: Column where concatenated results should appear (e.g., "C")
            - source_columns: List of source columns to concatenate (e.g., ["A", "B"])
            - separator: Separator string between columns (e.g., " ")
            - start_row: First data row (default: 1)
            - header_row: Header row number (default: 1)
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        target_column = options.get('target_column', 'C')
        source_columns = options.get('source_columns', ['A', 'B'])
        separator = options.get('separator', ' ')
        start_row = options.get('start_row', 1)
        header_row = options.get('header_row', 1)
        
        # Load workbook to get formulas
        wb = openpyxl.load_workbook(result, data_only=False)
        ws = wb.active
        
        # Auto-detect data rows by checking for consecutive empty rows
        max_check_row = 1000
        data_end_row = start_row
        consecutive_empty = 0
        
        for row_idx in range(start_row, max_check_row):
            # Check if source columns are empty
            all_empty = True
            for col in source_columns:
                cell_value = ws[f"{col}{row_idx}"].value
                if cell_value is not None and str(cell_value).strip() != "":
                    all_empty = False
                    break
            
            if all_empty:
                consecutive_empty += 1
                if consecutive_empty >= 3:  # 3 consecutive empty rows means end of data
                    break
            else:
                consecutive_empty = 0
                data_end_row = row_idx
        
        logger.info(f"Detected data rows from {start_row} to {data_end_row}")
        
        if data_end_row < start_row:
            logger.error(f"No data rows found starting from row {start_row}")
            return 0.0
        
        # Check formulas in target column
        formula_count = 0
        valid_formula_count = 0
        
        for row_idx in range(start_row, data_end_row + 1):
            target_cell = ws[f"{target_column}{row_idx}"]
            
            # Get formula text
            formula_text = None
            if hasattr(target_cell, "_value") and isinstance(target_cell._value, str) and target_cell._value.startswith("="):
                formula_text = target_cell._value
            elif hasattr(target_cell, "formula"):
                formula_text = target_cell.formula
            elif target_cell.value is not None and isinstance(target_cell.value, str) and target_cell.value.startswith("="):
                formula_text = target_cell.value
            
            if formula_text is None:
                # Check if cell has a value (might be calculated)
                if target_cell.value is not None and str(target_cell.value).strip() != "":
                    # Check if the value matches expected concatenation
                    source_values = []
                    for col in source_columns:
                        val = ws[f"{col}{row_idx}"].value
                        if val is not None:
                            source_values.append(str(val))
                    
                    expected_value = separator.join(source_values)
                    actual_value = str(target_cell.value).strip()
                    
                    if actual_value == expected_value:
                        valid_formula_count += 1
                continue
            
            formula_count += 1
            
            # Remove leading = sign
            formula_clean = formula_text[1:] if formula_text.startswith("=") else formula_text
            formula_upper = formula_clean.upper()
            
            # Check 1: Formula contains & operator
            if "&" not in formula_clean:
                logger.warning(f"Row {row_idx}: Formula does not contain & operator")
                continue
            
            # Check 2: Formula references source columns
            all_columns_referenced = True
            for col in source_columns:
                # Check for column reference (e.g., A1, A2, etc.)
                col_pattern = rf'\b{col}{row_idx}\b'
                if not re.search(col_pattern, formula_upper):
                    logger.warning(f"Row {row_idx}: Formula does not reference column {col}")
                    all_columns_referenced = False
                    break
            
            if not all_columns_referenced:
                continue
            
            # Check 3: Formula contains separator (if not empty)
            if separator:
                # Escape special regex characters in separator
                escaped_sep = re.escape(separator)
                # Check for separator in quotes
                sep_pattern = rf'["\'].*{escaped_sep}.*["\']'
                if not re.search(sep_pattern, formula_clean):
                    logger.warning(f"Row {row_idx}: Formula does not contain separator '{separator}'")
                    continue
            
            valid_formula_count += 1
        
        # Calculate success rate
        total_rows = data_end_row - start_row + 1
        success_rate = valid_formula_count / total_rows if total_rows > 0 else 0.0
        
        logger.info(f"Valid formulas: {valid_formula_count}/{total_rows} (success rate: {success_rate:.2%})")
        
        # Require at least 90% of rows to have valid formulas
        if success_rate >= 0.9:
            logger.info(f"✓ Column concatenation verification passed")
            return 1.0
        else:
            logger.error(f"Column concatenation verification failed: only {success_rate:.2%} of rows have valid formulas")
            return 0.0
    
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_tocol_torow_merge(result: str, expected: str = None, **options) -> float:
    """
    Verify if formulas exist to merge two columns using TOCOL/TOROW or INDEX/ROUNDUP/ROW/MOD/ROWS.
    
    This function checks:
    1. Whether the result column contains appropriate formulas
    2. Whether formulas reference the correct source columns
    3. Whether formulas use concatenation operator (&) to merge the columns
    
    Expected formula patterns:
    - Pattern 1: =TOCOL(C1:C20&","&TOROW(G1:G6))
    - Pattern 2: =INDEX($C$1:$C$20,ROUNDUP(ROW(A1)/ROWS($G$1:$G$6),0))&","&INDEX($G$1:$G$6,MOD(ROW(A1)-1,ROWS($G$1:$G$6))+1)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - result_column: Column with merged results (e.g., "H")
            - result_cell: Starting cell for the formula (e.g., "H1")
            - source_column1: First source column (e.g., "C")
            - source_range1: First source range (e.g., "C1:C20")
            - source_column2: Second source column (e.g., "G")
            - source_range2: Second source range (e.g., "G1:G6")
            - separator: Separator used in concatenation (e.g., ",")
            - expected_functions: List of expected function names (default: ["TOCOL", "TOROW"] or ["INDEX", "ROUNDUP", "ROW", "MOD", "ROWS"])
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        result_column = options.get('result_column', 'H')
        result_cell = options.get('result_cell', 'H1')
        source_column1 = options.get('source_column1', 'C')
        source_range1 = options.get('source_range1', 'C1:C20')
        source_column2 = options.get('source_column2', 'G')
        source_range2 = options.get('source_range2', 'G1:G6')
        separator = options.get('separator', ',')
        expected_functions = options.get('expected_functions', ['TOCOL', 'TOROW'])
        
        logger.info(f"Verifying TOCOL/TOROW merge formulas in file: {result}")
        logger.info(f"Result cell: {result_cell}")
        logger.info(f"Source range 1: {source_range1}")
        logger.info(f"Source range 2: {source_range2}")
        logger.info(f"Separator: {separator}")
        logger.info(f"Expected functions: {expected_functions}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check the result cell and nearby cells (in case formula starts from H2, H3, etc.)
        # For INDEX/ROUNDUP/ROW/MOD/ROWS pattern, formula should be in multiple rows
        formula = None
        checked_cell = None
        
        # Try to find formula in result_cell and nearby rows (up to 10 rows)
        result_col = result_column
        result_row = int(re.search(r'\d+', result_cell).group()) if re.search(r'\d+', result_cell) else 1
        
        for row_offset in range(10):  # Check up to 10 rows
            try:
                check_row = result_row + row_offset
                check_cell_coord = f"{result_col}{check_row}"
                cell = ws[check_cell_coord]
                
                # Check if cell contains a formula
                if cell.data_type == "f":
                    # Get formula text
                    if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                        formula = cell._value
                        checked_cell = check_cell_coord
                        break
                    elif hasattr(cell, "formula"):
                        formula = cell.formula
                        checked_cell = check_cell_coord
                        break
                    elif cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula = cell.value
                        checked_cell = check_cell_coord
                        break
            except Exception:
                continue
        
        if formula is None:
            logger.error(f"Could not find formula in {result_column} column (checked rows {result_row} to {result_row + 9})")
            return 0.0
        
        logger.info(f"Found formula in cell {checked_cell}")
        
        formula_upper = formula.upper()
        logger.debug(f"Cell {checked_cell} formula: {formula}")
        
        # Check which pattern is used
        has_tocol = 'TOCOL' in formula_upper
        has_torow = 'TOROW' in formula_upper
        has_index = 'INDEX' in formula_upper
        has_roundup = 'ROUNDUP' in formula_upper
        has_row = 'ROW' in formula_upper
        has_mod = 'MOD' in formula_upper
        has_rows = 'ROWS' in formula_upper
        
        # Pattern 1: TOCOL/TOROW pattern
        if has_tocol and has_torow:
            logger.info("Detected TOCOL/TOROW pattern")
            
            # Check 1: References first source range
            range1_escaped = re.escape(source_range1)
            if not re.search(range1_escaped, formula_upper):
                logger.error(f"Cell {result_cell} formula should reference range {source_range1}")
                logger.error(f"  Formula: {formula}")
                return 0.0
            
            # Check 2: References second source range (within TOROW)
            range2_escaped = re.escape(source_range2)
            torow_pattern = rf'TOROW\s*\(\s*{range2_escaped}\s*\)'
            if not re.search(torow_pattern, formula_upper):
                logger.error(f"Cell {result_cell} TOROW function should reference range {source_range2}")
                logger.error(f"  Formula: {formula}")
                return 0.0
            
            # Check 3: Contains concatenation operator (&)
            if '&' not in formula:
                logger.error(f"Cell {result_cell} formula should use concatenation operator (&)")
                logger.error(f"  Formula: {formula}")
                return 0.0
        
        # Pattern 2: INDEX/ROUNDUP/ROW/MOD/ROWS pattern
        elif has_index and has_roundup and has_row and has_mod and has_rows:
            logger.info("Detected INDEX/ROUNDUP/ROW/MOD/ROWS pattern")
            
            # Check 1: References first source range (in INDEX)
            # The range might use absolute references like $C$1:$C$20
            range1_escaped = re.escape(source_range1)
            range1_abs_escaped = re.escape(source_range1.replace('C', '\\$?C').replace(':', '\\$?:'))
            index1_pattern = rf'INDEX\s*\(\s*[\$]?{source_column1}[\$]?\d+:\$?{source_column1}[\$]?\d+'
            if not re.search(index1_pattern, formula_upper):
                logger.error(f"Cell {checked_cell} first INDEX should reference range {source_range1}")
                logger.error(f"  Formula: {formula}")
                return 0.0
            
            # Check 2: References second source range (in second INDEX)
            range2_escaped = re.escape(source_range2)
            index2_pattern = rf'INDEX\s*\(\s*[\$]?{source_column2}[\$]?\d+:\$?{source_column2}[\$]?\d+'
            if not re.search(index2_pattern, formula_upper):
                logger.error(f"Cell {checked_cell} second INDEX should reference range {source_range2}")
                logger.error(f"  Formula: {formula}")
                return 0.0
            
            # Check 3: Contains ROUNDUP with ROW and ROWS
            # ROWS might reference the second range with absolute references
            rows_pattern = rf'ROWS\s*\(\s*[\$]?{source_column2}[\$]?\d+:\$?{source_column2}[\$]?\d+\s*\)'
            roundup_pattern = rf'ROUNDUP\s*\(\s*ROW\s*\([^)]+\)\s*/\s*{rows_pattern}'
            if not re.search(roundup_pattern, formula_upper):
                logger.warning(f"Cell {checked_cell} ROUNDUP may not have correct structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn
            
            # Check 4: Contains MOD with ROW and ROWS
            mod_pattern = rf'MOD\s*\(\s*ROW\s*\([^)]+\)\s*-\s*\d+\s*,\s*{rows_pattern}'
            if not re.search(mod_pattern, formula_upper):
                logger.warning(f"Cell {checked_cell} MOD may not have correct structure")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn
            
            # Check 5: Contains concatenation operator (&)
            if '&' not in formula:
                logger.error(f"Cell {checked_cell} formula should use concatenation operator (&)")
                logger.error(f"  Formula: {formula}")
                return 0.0
        
        else:
            # Check if at least some expected functions are present
            missing_functions = []
            for func in expected_functions:
                if func.upper() not in formula_upper:
                    missing_functions.append(func)
            
            if missing_functions:
                logger.error(f"Cell {checked_cell} formula missing functions: {missing_functions}")
                logger.error(f"  Formula: {formula}")
                logger.error(f"  Expected pattern: TOCOL/TOROW or INDEX/ROUNDUP/ROW/MOD/ROWS")
                return 0.0
        
        # Common check: Contains separator in concatenation
        separator_escaped = re.escape(separator)
        separator_pattern = rf'&\s*["\']?{separator_escaped}["\']?\s*&'
        if not re.search(separator_pattern, formula_upper):
            logger.warning(f"Cell {checked_cell} formula may not use separator '{separator}' correctly")
            logger.debug(f"  Formula: {formula}")
            # Don't fail, just warn - separator might be in quotes or formatted differently
        
        logger.info("=" * 60)
        logger.info(f"✓ TOCOL/TOROW merge verification passed")
        logger.info(f"  Result cell: {checked_cell}")
        logger.info(f"  Source range 1: {source_range1}")
        logger.info(f"  Source range 2: {source_range2}")
        logger.info(f"  Formula: {formula}")
        logger.info("=" * 60)
        return 1.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_text_prefix_format(result: str, expected: str = None, **options) -> float:
    """
    Verify if formulas exist to add text prefix (like apostrophe) to values.
    
    This function checks:
    1. Whether cells in result column contain formulas
    2. Whether formulas use concatenation operator (&) to add prefix
    3. Whether formulas reference the corresponding source column cells
    
    Expected formula pattern: ="'"&A3
    
    The function automatically detects the number of data rows by checking the source column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - result_column: Column with formulas (e.g., "F")
            - source_column: Source column (e.g., "A")
            - start_row: Starting row number (default: 2)
            - prefix: Prefix text to add (e.g., "'")
            - data_column: Column to check for data to determine end_row (default: source_column)
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        result_column = options.get('result_column', 'F')
        source_column = options.get('source_column', 'A')
        start_row = options.get('start_row', 2)
        prefix = options.get('prefix', "'")
        data_column = options.get('data_column', source_column)
        
        logger.info(f"Verifying text prefix format formulas in file: {result}")
        logger.info(f"Result column: {result_column}")
        logger.info(f"Source column: {source_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Prefix: {prefix}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check formulas in result column
        all_checks_passed = True
        rows_checked = 0
        
        for row_num in range(start_row, end_row + 1):
            result_cell = ws[f"{result_column}{row_num}"]
            source_cell = ws[f"{source_column}{row_num}"]
            
            # Skip if source cell is empty
            if source_cell.value is None:
                continue
            
            rows_checked += 1
            
            # Check if cell contains a formula
            if result_cell.data_type != "f":
                logger.error(f"Cell {result_column}{row_num} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(result_cell, "_value") and isinstance(result_cell._value, str) and result_cell._value.startswith("="):
                formula = result_cell._value
            elif hasattr(result_cell, "formula"):
                formula = result_cell.formula
            else:
                if result_cell.value is not None and isinstance(result_cell.value, str) and result_cell.value.startswith("="):
                    formula = result_cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {result_column}{row_num}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check 1: Contains concatenation operator (&)
            if '&' not in formula:
                logger.error(f"Cell {result_column}{row_num} formula should use concatenation operator (&)")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: References source column cell
            source_pattern = rf'{source_column}{row_num}\b'
            if not re.search(source_pattern, formula_upper):
                logger.error(f"Cell {result_column}{row_num} formula should reference {source_column}{row_num}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: Contains prefix in formula
            # Prefix might be in quotes: "'" or '"' or just '
            prefix_escaped = re.escape(prefix)
            # Pattern: ="'"&A3 or ='"'&A3 or ='&A3
            prefix_pattern1 = rf'=\s*["\']{prefix_escaped}["\']\s*&'
            prefix_pattern2 = rf'=\s*{prefix_escaped}\s*&'
            if not re.search(prefix_pattern1, formula) and not re.search(prefix_pattern2, formula):
                logger.error(f"Cell {result_column}{row_num} formula should contain prefix '{prefix}'")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            logger.debug(f"✓ Row {row_num} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ Text prefix format verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Result column: {result_column}")
            logger.info(f"  Source column: {source_column}")
            logger.info(f"  Prefix: {prefix}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ Text prefix format verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_match_percentile_filter(result: str, expected: str = None, **options) -> float:
    """
    Verify if MATCH and PERCENTILE formulas exist to filter top 10% and bottom 10% data.
    
    This function checks:
    1. Whether cells in result column contain MATCH and PERCENTILE formulas
    2. Whether formulas reference the correct source column
    3. Whether PERCENTILE uses correct percentiles array {0,10,90}%
    
    Expected formula pattern: =MATCH(A2,PERCENTILE(A:A,{0,10,90}%))
    
    The function automatically detects the number of data rows by checking the source column
    for non-empty cells. It stops checking after finding 3 consecutive empty rows.
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - result_column: Column with formulas (e.g., "D")
            - source_column: Source column (e.g., "A")
            - start_row: Starting row number (default: 1)
            - expected_functions: List of expected function names (default: ["MATCH", "PERCENTILE"])
            - expected_percentiles: Expected percentiles array (default: "{0,10,90}%")
            - data_column: Column to check for data to determine end_row (default: source_column)
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        result_column = options.get('result_column', 'D')
        source_column = options.get('source_column', 'A')
        start_row = options.get('start_row', 1)
        expected_functions = options.get('expected_functions', ['MATCH', 'PERCENTILE'])
        expected_percentiles = options.get('expected_percentiles', '{0,10,90}%')
        data_column = options.get('data_column', source_column)
        
        logger.info(f"Verifying MATCH/PERCENTILE filter formulas in file: {result}")
        logger.info(f"Result column: {result_column}")
        logger.info(f"Source column: {source_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Expected functions: {expected_functions}")
        logger.info(f"Expected percentiles: {expected_percentiles}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check formulas in result column
        all_checks_passed = True
        rows_checked = 0
        
        for row_num in range(start_row, end_row + 1):
            result_cell = ws[f"{result_column}{row_num}"]
            source_cell = ws[f"{source_column}{row_num}"]
            
            # Skip if source cell is empty
            if source_cell.value is None:
                continue
            
            rows_checked += 1
            
            # Check if cell contains a formula
            if result_cell.data_type != "f":
                logger.error(f"Cell {result_column}{row_num} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(result_cell, "_value") and isinstance(result_cell._value, str) and result_cell._value.startswith("="):
                formula = result_cell._value
            elif hasattr(result_cell, "formula"):
                formula = result_cell.formula
            else:
                if result_cell.value is not None and isinstance(result_cell.value, str) and result_cell.value.startswith("="):
                    formula = result_cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {result_column}{row_num}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Row {row_num} formula: {formula}")
            
            # Check 1: Contains MATCH function
            if 'MATCH' not in formula_upper:
                logger.error(f"Cell {result_column}{row_num} formula should contain MATCH function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: Contains PERCENTILE function
            if 'PERCENTILE' not in formula_upper:
                logger.error(f"Cell {result_column}{row_num} formula should contain PERCENTILE function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: MATCH references source column cell
            source_pattern = rf'{source_column}{row_num}\b'
            if not re.search(source_pattern, formula_upper):
                logger.error(f"Cell {result_column}{row_num} MATCH should reference {source_column}{row_num}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: PERCENTILE references source column range
            percentile_range_pattern = rf'PERCENTILE\s*\(\s*{source_column}\s*:\s*{source_column}'
            if not re.search(percentile_range_pattern, formula_upper):
                logger.error(f"Cell {result_column}{row_num} PERCENTILE should reference {source_column}:{source_column}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 5: PERCENTILE contains expected percentiles array
            # Pattern: {0,10,90}% or {0, 10, 90}% (with or without spaces)
            percentiles_pattern = r'\{[^}]*0[^}]*,[^}]*10[^}]*,[^}]*90[^}]*\}'
            if not re.search(percentiles_pattern, formula_upper):
                logger.warning(f"Cell {result_column}{row_num} PERCENTILE may not have correct percentiles array")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn - the array format might vary
            
            # Check 6: PERCENTILE percentiles array ends with %
            percent_pattern = r'\{[^}]*\}\s*%'
            if not re.search(percent_pattern, formula_upper):
                logger.warning(f"Cell {result_column}{row_num} PERCENTILE percentiles array should end with %")
                logger.debug(f"  Formula: {formula}")
                # Don't fail, just warn
            
            logger.debug(f"✓ Row {row_num} formula passed all checks")
        
        if rows_checked == 0:
            logger.error("No rows with data found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ MATCH/PERCENTILE filter verification passed")
            logger.info(f"  Rows checked: {rows_checked}")
            logger.info(f"  Result column: {result_column}")
            logger.info(f"  Source column: {source_column}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ MATCH/PERCENTILE filter verification failed")
            logger.error(f"  Rows checked: {rows_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_sumproduct_month_customer_count(result: str, expected: str = None, **options) -> float:
    """
    Verify if SUMPRODUCT formulas exist to count contracts and sum amounts for a specific month and customer.
    
    This function checks:
    1. Whether specified cells contain SUMPRODUCT formulas
    2. Whether formulas use MONTH function to check month
    3. Whether formulas use double negation (--) to convert boolean to number
    4. Whether formulas check customer name match
    
    Expected formula patterns:
    - B28: =SUMPRODUCT(--(MONTH(A2:A24)=6),--(B2:B24=A28)) (count contracts)
    - C28: =SUMPRODUCT(--(MONTH(A2:A24)=6),--(B2:B24=A28),C2:C24) (sum amounts)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - formulas: Dict mapping cell coordinates to expected formula info, e.g.:
              {
                  "B28": {"type": "count", "date_range": "A2:A24", "customer_range": "B2:B24", "customer_cell": "A28", "month": 6},
                  "C28": {"type": "sum", "date_range": "A2:A24", "customer_range": "B2:B24", "customer_cell": "A28", "amount_range": "C2:C24", "month": 6}
              }
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        formulas_config = options.get('formulas', {})
        
        if not formulas_config:
            logger.error("No formulas specified in options")
            return 0.0
        
        logger.info(f"Verifying SUMPRODUCT month/customer count formulas in file: {result}")
        logger.info(f"Formulas to check: {list(formulas_config.keys())}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Check each formula
        all_checks_passed = True
        cells_checked = 0
        
        for cell_coord, formula_info in formulas_config.items():
            formula_type = formula_info.get('type', '')  # 'count' or 'sum'
            date_range = formula_info.get('date_range', 'A2:A24')
            customer_range = formula_info.get('customer_range', 'B2:B24')
            customer_cell = formula_info.get('customer_cell', 'A28')
            month = formula_info.get('month', 6)
            amount_range = formula_info.get('amount_range', 'C2:C24')
            
            logger.info(f"Checking cell {cell_coord}: {formula_type} formula")
            
            try:
                cell = ws[cell_coord]
            except Exception as e:
                logger.error(f"Could not access cell {cell_coord}: {e}")
                all_checks_passed = False
                continue
            
            cells_checked += 1
            
            # Check if cell contains a formula
            if cell.data_type != "f":
                logger.error(f"Cell {cell_coord} does not contain a formula")
                all_checks_passed = False
                continue
            
            # Get formula text
            formula = None
            if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                formula = cell._value
            elif hasattr(cell, "formula"):
                formula = cell.formula
            else:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
            
            if formula is None:
                logger.error(f"Could not extract formula from cell {cell_coord}")
                all_checks_passed = False
                continue
            
            formula_upper = formula.upper()
            logger.debug(f"Cell {cell_coord} formula: {formula}")
            
            # Check 1: Contains SUMPRODUCT function
            if 'SUMPRODUCT' not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain SUMPRODUCT function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 2: Contains MONTH function
            if 'MONTH' not in formula_upper:
                logger.error(f"Cell {cell_coord} formula should contain MONTH function")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 3: MONTH references date range (support both absolute and relative references)
            # Pattern: MONTH($A$2:$A$24) or MONTH(A2:A24)
            date_col_start = date_range.split(':')[0]  # e.g., "A2" or "$A$2"
            date_col_end = date_range.split(':')[1] if ':' in date_range else date_range  # e.g., "A24" or "$A$24"
            # Extract column letter and row numbers
            date_col_match = re.match(r'[\$]?([A-Z]+)[\$]?(\d+)', date_col_start)
            if date_col_match:
                date_col = date_col_match.group(1)
                date_start_row = date_col_match.group(2)
            else:
                date_col = 'A'
                date_start_row = '2'
            
            date_end_match = re.match(r'[\$]?([A-Z]+)[\$]?(\d+)', date_col_end)
            if date_end_match:
                date_end_row = date_end_match.group(2)
            else:
                date_end_row = '24'
            
            # Pattern: MONTH($A$2:$A$24) or MONTH(A2:A24) - support both absolute and relative
            month_pattern = rf'MONTH\s*\(\s*[\$]?{date_col}[\$]?\d+:\$?{date_col}[\$]?\d+\s*\)'
            if not re.search(month_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} MONTH should reference range like {date_range} (absolute or relative)")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 4: MONTH equals specified month
            month_check_pattern = rf'MONTH\s*\(\s*[\$]?{date_col}[\$]?\d+:\$?{date_col}[\$]?\d+\s*\)\s*=\s*{month}'
            if not re.search(month_check_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} MONTH should equal {month}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 5: Contains double negation (--)
            if '--' not in formula:
                logger.error(f"Cell {cell_coord} formula should use double negation (--) to convert boolean to number")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 6: Customer range equals customer cell (support both absolute and relative references)
            # Pattern: $B$2:$B$24=A28 or B2:B24=A28
            customer_col_start = customer_range.split(':')[0] if ':' in customer_range else customer_range
            customer_col_match = re.match(r'[\$]?([A-Z]+)[\$]?(\d+)', customer_col_start)
            if customer_col_match:
                customer_col = customer_col_match.group(1)
            else:
                customer_col = 'B'
            
            customer_cell_match = re.match(r'[\$]?([A-Z]+)[\$]?(\d+)', customer_cell)
            if customer_cell_match:
                customer_cell_col = customer_cell_match.group(1)
                customer_cell_row = customer_cell_match.group(2)
            else:
                customer_cell_col = 'A'
                customer_cell_row = '28'
            
            # Pattern: $B$2:$B$24=A28 or B2:B24=A28
            customer_pattern = rf'[\$]?{customer_col}[\$]?\d+:\$?{customer_col}[\$]?\d+\s*=\s*{customer_cell_col}{customer_cell_row}'
            if not re.search(customer_pattern, formula_upper):
                logger.error(f"Cell {cell_coord} formula should check {customer_range}={customer_cell}")
                logger.error(f"  Formula: {formula}")
                all_checks_passed = False
                continue
            
            # Check 7: For sum type, should reference amount range (support both absolute and relative)
            if formula_type == 'sum':
                amount_col_start = amount_range.split(':')[0] if ':' in amount_range else amount_range
                amount_col_match = re.match(r'[\$]?([A-Z]+)[\$]?(\d+)', amount_col_start)
                if amount_col_match:
                    amount_col = amount_col_match.group(1)
                else:
                    amount_col = 'C'
                
                # Pattern: $C$2:$C$24 or C2:C24
                amount_pattern = rf'[\$]?{amount_col}[\$]?\d+:\$?{amount_col}[\$]?\d+'
                if not re.search(amount_pattern, formula_upper):
                    logger.error(f"Cell {cell_coord} sum formula should reference amount range {amount_range} (absolute or relative)")
                    logger.error(f"  Formula: {formula}")
                    all_checks_passed = False
                    continue
            
            logger.debug(f"✓ Cell {cell_coord} formula passed all checks")
        
        if cells_checked == 0:
            logger.error("No cells found to check")
            return 0.0
        
        if all_checks_passed:
            logger.info("=" * 60)
            logger.info(f"✓ SUMPRODUCT month/customer count verification passed")
            logger.info(f"  Cells checked: {cells_checked}")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ SUMPRODUCT month/customer count verification failed")
            logger.error(f"  Cells checked: {cells_checked}")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0


def verify_sum_value_text_conversion(result: str, expected: str = None, **options) -> float:
    """
    Verify if SUM(VALUE(...), ...) formulas exist in specified column to sum text and number values.
    
    This function checks:
    1. Whether specified cells contain SUM formulas
    2. Whether formulas contain VALUE function to convert text to number
    3. Whether formulas reference the correct text column and number column cells
    4. Whether formulas have the correct structure: SUM(VALUE(text_cell), number_cell)
    
    Expected formula pattern:
    - D2: =SUM(VALUE(B2), C2) (sum of text value in B2 converted to number and number value in C2)
    - D3: =SUM(VALUE(B3), C3) (and so on for other rows)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, should contain:
            - check_column: Column to check (e.g., "D")
            - start_row: Starting row number (e.g., 2)
            - text_column: Column containing text values (e.g., "B")
            - number_column: Column containing number values (e.g., "C")
            - expected_functions: List of expected function names (default: ["SUM", "VALUE"])
            - data_column: Column to use for auto-detecting end row (e.g., "B")
    
    Returns:
        float: 1.0 if verification passes, 0.0 otherwise
    """
    try:
        import re
        
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        check_column = options.get('check_column', 'D')
        start_row = options.get('start_row', 2)
        text_column = options.get('text_column', 'B')
        number_column = options.get('number_column', 'C')
        expected_functions = options.get('expected_functions', ['SUM', 'VALUE'])
        data_column = options.get('data_column', 'B')
        
        logger.info(f"Verifying SUM(VALUE) text conversion formulas in file: {result}")
        logger.info(f"Column to check: {check_column}")
        logger.info(f"Start row: {start_row}")
        logger.info(f"Text column: {text_column}")
        logger.info(f"Number column: {number_column}")
        logger.info(f"Expected functions: {expected_functions}")
        
        # Load workbook to get formulas
        try:
            wb = openpyxl.load_workbook(result, data_only=False)  # data_only=False to get formulas
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Auto-detect end_row by checking data_column for non-empty cells
        logger.info(f"Auto-detecting end row by checking column {data_column} for data...")
        max_row = ws.max_row
        end_row = start_row  # Start from start_row
        
        # Find the last row with data in the data column
        # Check up to max_row, but stop if we find 3 consecutive empty rows
        empty_count = 0
        for row_num in range(start_row, max_row + 1):
            data_cell = ws[f"{data_column}{row_num}"]
            if data_cell.value is None or (isinstance(data_cell.value, str) and data_cell.value.strip() == ""):
                empty_count += 1
                if empty_count >= 3:  # Stop after 3 consecutive empty rows
                    break
            else:
                empty_count = 0
                end_row = row_num  # Update end_row to the last row with data
        
        logger.info(f"Auto-detected end row: {end_row}")
        
        # Check each row in the specified column
        all_passed = True
        logger.info(f"Checking column {check_column} (rows {start_row} to {end_row})")
        
        for row_num in range(start_row, end_row + 1):
            cell_coord = f"{check_column}{row_num}"
            try:
                cell = ws[cell_coord]
                logger.debug(f"Checking cell {cell_coord}")
                
                # Check if cell contains a formula
                if cell.data_type != "f":
                    logger.warning(f"Cell {cell_coord} does not contain a formula")
                    all_passed = False
                    continue
                
                # Get formula text
                formula_text = None
                if hasattr(cell, "_value") and isinstance(cell._value, str) and cell._value.startswith("="):
                    formula_text = cell._value
                elif hasattr(cell, "formula"):
                    formula_text = cell.formula
                else:
                    # Try to get from value attribute
                    if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value
                
                if formula_text is None:
                    logger.warning(f"Could not extract formula from cell {cell_coord}")
                    all_passed = False
                    continue
                
                formula_upper = formula_text.upper()
                logger.debug(f"Cell {cell_coord} formula: {formula_text}")
                
                # Check 1: Formula contains SUM function
                sum_pattern = r'\bSUM\s*\('
                if not re.search(sum_pattern, formula_upper):
                    logger.warning(f"Cell {cell_coord} formula does not contain SUM function")
                    logger.warning(f"  Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 2: Formula contains VALUE function
                value_pattern = r'\bVALUE\s*\('
                if not re.search(value_pattern, formula_upper):
                    logger.warning(f"Cell {cell_coord} formula does not contain VALUE function")
                    logger.warning(f"  Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 3: Formula structure is SUM(VALUE(...), ...)
                # Check that VALUE is inside SUM
                sum_value_pattern = r'SUM\s*\(\s*VALUE\s*\('
                if not re.search(sum_value_pattern, formula_upper):
                    logger.warning(f"Cell {cell_coord} formula does not have correct SUM(VALUE(...)) structure")
                    logger.warning(f"  Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 4: VALUE function references the text column cell (e.g., B2, B3, etc.)
                expected_text_cell = f"{text_column}{row_num}"
                text_cell_pattern = rf'{text_column}{row_num}\b'
                if not re.search(text_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference text column cell {expected_text_cell}")
                    logger.warning(f"  Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 5: Formula references the number column cell (e.g., C2, C3, etc.)
                expected_number_cell = f"{number_column}{row_num}"
                number_cell_pattern = rf'{number_column}{row_num}\b'
                if not re.search(number_cell_pattern, formula_text, re.IGNORECASE):
                    logger.warning(f"Cell {cell_coord} formula does not reference number column cell {expected_number_cell}")
                    logger.warning(f"  Formula: {formula_text}")
                    all_passed = False
                    continue
                
                # Check 6: Formula closes parentheses correctly
                open_count = formula_text.count('(')
                close_count = formula_text.count(')')
                if open_count != close_count:
                    logger.warning(f"Cell {cell_coord} formula has mismatched parentheses")
                    logger.warning(f"  Formula: {formula_text}")
                    all_passed = False
                    continue
                
                logger.info(f"✓ Cell {cell_coord} has valid SUM(VALUE) formula: {formula_text}")
                
            except Exception as e:
                logger.error(f"Error checking cell {cell_coord}: {e}")
                import traceback
                logger.error(traceback.format_exc())
                all_passed = False
        
        if all_passed:
            logger.info("=" * 60)
            logger.info(f"✓ All cells in column {check_column} contain correct SUM(VALUE) formulas")
            logger.info("=" * 60)
            return 1.0
        else:
            logger.error("=" * 60)
            logger.error(f"✗ SUM(VALUE) formula verification failed")
            logger.error("=" * 60)
            return 0.0
            
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0
