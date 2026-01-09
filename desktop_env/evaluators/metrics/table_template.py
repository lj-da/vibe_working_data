import logging
import os
import re
from typing import Set, List, Tuple, Dict

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.borders import Border, Side

logger = logging.getLogger("desktopenv.metric.table_template")


def normalize_for_matching(text: str, case_sensitive: bool = False) -> str:
    """
    Normalize text for matching by:
    1. Removing newlines and multiple spaces
    2. Removing common punctuation (colons, periods, etc.)
    3. Removing all spaces for flexible matching
    4. Converting to lowercase if case insensitive
    
    Args:
        text: Text to normalize
        case_sensitive: Whether to preserve case
    
    Returns:
        Normalized text
    """
    # Remove newlines and normalize spaces
    normalized = ' '.join(text.split())
    # Remove common punctuation (colons, periods, commas, etc.)
    normalized = re.sub(r'[：:，,。.、；;！!？?]', '', normalized)
    # Remove all spaces for more flexible matching
    normalized = normalized.replace(' ', '')
    if not case_sensitive:
        normalized = normalized.lower()
    return normalized


def find_field(field: str, text_values: Set[str], case_sensitive: bool = False) -> Tuple[bool, str]:
    """
    Check if a field exists in text values.
    Returns (found, matched_text) tuple.
    Normalizes text by removing newlines, multiple spaces, and punctuation.
    Prioritizes more precise matches (exact match > field in text > text in field).
    
    Args:
        field: Field name to search for
        text_values: Set of text values to search in
        case_sensitive: Whether matching is case sensitive
    
    Returns:
        Tuple of (found: bool, matched_text: str)
    """
    # Normalize field text for matching
    field_normalized = normalize_for_matching(field, case_sensitive)
    
    # Collect all matches with their match quality scores
    matches: List[Tuple[int, str]] = []  # (score, text_value)
    
    for text_value in text_values:
        text_normalized = normalize_for_matching(text_value, case_sensitive)
        
        # Exact match after normalization (highest priority: score 3)
        if field_normalized == text_normalized:
            matches.append((3, text_value))
        # Field is contained in text (medium priority: score 2)
        # This means the text is longer/more specific, e.g., "入职 日期" contains "入职日期"
        elif field_normalized in text_normalized:
            matches.append((2, text_value))
        # Text is contained in field (lowest priority: score 1)
        # This means the text is shorter, e.g., "日期" is in "入职日期"
        elif text_normalized in field_normalized:
            matches.append((1, text_value))
    
    # Return the best match (highest score, and if tie, prefer longer text)
    if matches:
        # Sort by score (descending), then by text length (descending)
        matches.sort(key=lambda x: (x[0], len(x[1])), reverse=True)
        return True, matches[0][1]
    
    return False, ""


def collect_text_and_field_cells(wb: Workbook, core_fields: List[str], optional_fields: List[str], 
                                  case_sensitive: bool = False) -> Tuple[Set[str], Dict[str, List[Tuple[Worksheet, int, int]]]]:
    """
    Collect all text values and field cell positions from the workbook.
    
    Args:
        wb: OpenPyXL workbook
        core_fields: List of core field names
        optional_fields: List of optional field names
        case_sensitive: Whether matching is case sensitive
    
    Returns:
        Tuple of (all_text_values: Set[str], field_cells: Dict[str, List[Tuple[Worksheet, int, int]]])
    """
    all_text_values: Set[str] = set()
    field_cells: Dict[str, List[Tuple[Worksheet, int, int]]] = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.info(f"Scanning sheet: {sheet_name}")
        
        # Iterate through all cells in the worksheet
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value is not None:
                    # Convert to string and strip whitespace
                    cell_text = str(cell.value).strip()
                    if cell_text:
                        # Normalize text: replace newlines and multiple spaces with single space
                        cell_text_normalized = ' '.join(cell_text.split())
                        if not case_sensitive:
                            cell_text_normalized = cell_text_normalized.lower()
                        all_text_values.add(cell_text_normalized)
                        
                        # Store cell position for field matching (using normalized matching)
                        for field in core_fields + optional_fields:
                            field_normalized = normalize_for_matching(field, case_sensitive)
                            cell_normalized = normalize_for_matching(cell_text, case_sensitive)
                            
                            # Check if this cell matches the field
                            if field_normalized == cell_normalized or \
                               field_normalized in cell_normalized or \
                               cell_normalized in field_normalized:
                                if field not in field_cells:
                                    field_cells[field] = []
                                field_cells[field].append((ws, row_idx, col_idx))
    
    logger.info(f"Found {len(all_text_values)} unique text values across all sheets")
    return all_text_values, field_cells


def verify_field_existence(core_fields: List[str], optional_fields: List[str], 
                           all_text_values: Set[str], case_sensitive: bool = False) -> Tuple[List[str], List[str], List[str], List[str]]:
    """
    Verify if core and optional fields exist in the template.
    
    Args:
        core_fields: List of core field names
        optional_fields: List of optional field names
        all_text_values: Set of all text values in the template
        case_sensitive: Whether matching is case sensitive
    
    Returns:
        Tuple of (found_core_fields, missing_core_fields, found_optional_fields, missing_optional_fields)
    """
    found_core_fields: List[str] = []
    missing_core_fields: List[str] = []
    
    for field in core_fields:
        found, matched_text = find_field(field, all_text_values, case_sensitive)
        if found:
            found_core_fields.append(field)
            logger.info(f"✓ Found core field: {field} (matched: {matched_text})")
        else:
            missing_core_fields.append(field)
            logger.warning(f"✗ Missing core field: {field}")
    
    # Check for optional fields (for logging only)
    found_optional_fields: List[str] = []
    missing_optional_fields: List[str] = []
    
    for field in optional_fields:
        found, matched_text = find_field(field, all_text_values, case_sensitive)
        if found:
            found_optional_fields.append(field)
            logger.info(f"✓ Found optional field: {field} (matched: {matched_text})")
        else:
            missing_optional_fields.append(field)
    
    return found_core_fields, missing_core_fields, found_optional_fields, missing_optional_fields


def verify_borders(found_core_fields: List[str], field_cells: Dict[str, List[Tuple[Worksheet, int, int]]], 
                   min_border_ratio: float) -> Tuple[bool, int, int]:
    """
    Verify borders on field cells.
    
    Args:
        found_core_fields: List of found core fields
        field_cells: Dictionary mapping fields to their cell positions
        min_border_ratio: Minimum ratio of fields with borders
    
    Returns:
        Tuple of (passed: bool, fields_with_borders: int, total_field_cells: int)
    """
    logger.info("Checking borders on field cells...")
    fields_with_borders = 0
    total_field_cells = 0
    
    for field in found_core_fields:
        if field in field_cells:
            for ws, row, col in field_cells[field]:
                total_field_cells += 1
                cell = ws.cell(row, col)
                
                # Check if cell has any border
                has_border = False
                if cell.border:
                    border = cell.border
                    # Check if any side has a border style
                    if (border.left and border.left.style) or \
                       (border.right and border.right.style) or \
                       (border.top and border.top.style) or \
                       (border.bottom and border.bottom.style):
                        has_border = True
                
                if has_border:
                    fields_with_borders += 1
    
    if total_field_cells > 0:
        border_ratio = fields_with_borders / total_field_cells
        logger.info(f"Border check: {fields_with_borders}/{total_field_cells} field cells have borders ({border_ratio:.2%})")
        
        if border_ratio < min_border_ratio:
            logger.error(f"Border ratio {border_ratio:.2%} is below minimum {min_border_ratio:.2%}")
            logger.error("Template should have borders on field cells for better formatting")
            return False, fields_with_borders, total_field_cells
        else:
            logger.info(f"✓ Border check passed: {border_ratio:.2%} >= {min_border_ratio:.2%}")
            return True, fields_with_borders, total_field_cells
    else:
        logger.warning("No field cells found for border checking")
        return True, 0, 0


def verify_layout_and_formatting(found_core_fields: List[str], field_cells: Dict[str, List[Tuple[Worksheet, int, int]]],
                                 min_rows: int, min_columns: int, min_formatting_ratio: float) -> Tuple[bool, int, int, float]:
    """
    Verify layout distribution and formatting quality.
    
    Args:
        found_core_fields: List of found core fields
        field_cells: Dictionary mapping fields to their cell positions
        min_rows: Minimum number of rows
        min_columns: Minimum number of columns
        min_formatting_ratio: Minimum ratio of cells with formatting
    
    Returns:
        Tuple of (passed: bool, unique_rows: int, unique_cols: int, formatting_ratio: float)
    """
    logger.info("Checking layout distribution and formatting quality...")
    all_field_rows: Set[int] = set()
    all_field_cols: Set[int] = set()
    cells_with_formatting = 0
    total_field_cells_checked = 0
    
    for field in found_core_fields:
        if field in field_cells:
            for ws, row, col in field_cells[field]:
                all_field_rows.add(row)
                all_field_cols.add(col)
                total_field_cells_checked += 1
                
                cell = ws.cell(row, col)
                
                # Check for various formatting indicators
                has_formatting = False
                
                # Check border
                if cell.border:
                    border = cell.border
                    if (border.left and border.left.style) or \
                       (border.right and border.right.style) or \
                       (border.top and border.top.style) or \
                       (border.bottom and border.bottom.style):
                        has_formatting = True
                
                # Check if cell is merged (merged cells indicate template structure)
                if hasattr(ws, 'merged_cells'):
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            has_formatting = True
                            break
                
                # Check fill/background color (non-default background indicates formatting)
                if cell.fill and cell.fill.patternType:
                    if cell.fill.patternType != 'none':
                        has_formatting = True
                
                # Check alignment (non-default alignment indicates formatting)
                if cell.alignment:
                    if cell.alignment.horizontal and cell.alignment.horizontal != 'general':
                        has_formatting = True
                    if cell.alignment.vertical and cell.alignment.vertical != 'bottom':
                        has_formatting = True
                
                if has_formatting:
                    cells_with_formatting += 1
    
    unique_rows = len(all_field_rows)
    unique_cols = len(all_field_cols)
    formatting_ratio = cells_with_formatting / total_field_cells_checked if total_field_cells_checked > 0 else 0
    
    logger.info(f"Layout check: Fields distributed across {unique_rows} rows and {unique_cols} columns")
    logger.info(f"Formatting check: {cells_with_formatting}/{total_field_cells_checked} cells have formatting ({formatting_ratio:.2%})")
    
    # Check 1: Minimum layout distribution (prevent all in one row/column)
    if unique_rows < min_rows:
        logger.error(f"Fields are in only {unique_rows} row(s), but template should have at least {min_rows} rows")
        logger.error("Template layout is too compact - fields should be distributed across multiple rows for better readability")
        return False, unique_rows, unique_cols, formatting_ratio
    
    if unique_cols < min_columns:
        logger.error(f"Fields are in only {unique_cols} column(s), but template should have at least {min_columns} columns")
        logger.error("Template layout is too narrow - fields should be distributed across multiple columns for better structure")
        return False, unique_rows, unique_cols, formatting_ratio
    
    # Check 2: Formatting quality (at least some cells should have formatting)
    if formatting_ratio < min_formatting_ratio:
        logger.error(f"Formatting ratio {formatting_ratio:.2%} is below minimum {min_formatting_ratio:.2%}")
        logger.error("Template lacks proper formatting - cells should have borders, alignment, fill, or other styling for better appearance")
        return False, unique_rows, unique_cols, formatting_ratio
    
    logger.info(f"✓ Layout check passed:")
    logger.info(f"  - Distribution: {unique_rows} rows >= {min_rows}, {unique_cols} columns >= {min_columns}")
    logger.info(f"  - Formatting: {formatting_ratio:.2%} of cells have formatting")
    
    return True, unique_rows, unique_cols, formatting_ratio


def verify_template_fields(result: str, expected: str = None, **options) -> float:
    """
    Verify if an Excel template file contains the required core fields and optional fields.
    Also verifies template formatting (borders, layout) to ensure it's a proper template, not just plain data.
    
    This is a generic function that can be used for any template type (resume, invoice, contract, etc.)
    by configuring the field lists in the options.
    
    This function checks:
    1. Whether the Excel file exists and can be loaded
    2. Whether the template contains ALL core fields (required, specified in options)
    3. Whether the template contains optional fields (optional, for logging only, specified in options)
    4. Flexible field matching (allows variations and partial matches)
    5. Whether fields have borders (template formatting requirement)
    6. Whether fields are distributed across multiple rows/columns (not all in one row)
    7. Whether cells have proper formatting (borders, alignment, fill, merged cells, etc.)
    
    Args:
        result (str): Path to result Excel file
        expected (str): Not used (for compatibility with framework interface)
        options (dict): Configuration options, must contain:
            - core_fields: List of core field names that MUST be found (required)
            - optional_fields: List of optional field names (optional, default: empty list)
            - case_sensitive: Whether field matching is case sensitive (optional, default: False)
            - check_borders: Whether to check for borders on field cells (optional, default: True)
            - min_border_ratio: Minimum ratio of fields with borders (optional, default: 0.7, i.e., 70%)
            - check_layout: Whether to check layout distribution and formatting quality (optional, default: True)
            - min_rows: Minimum number of rows that should contain fields (optional, default: 3)
            - min_columns: Minimum number of columns that should contain fields (optional, default: 2)
            - min_formatting_ratio: Minimum ratio of cells with formatting (borders, alignment, fill, etc.) (optional, default: 0.5, i.e., 50%)
            - core_field_ratio: Minimum ratio of core fields that must be found (optional, default: 1.0, i.e., 100%)
    
    Returns:
        float: 1.0 if all checks pass, 0.0 otherwise
    """
    try:
        if result is None or not os.path.exists(result):
            logger.error(f"Result file not found: {result}")
            return 0.0
        
        # Get options from JSON configuration
        core_fields = options.get('core_fields', [])
        optional_fields = options.get('optional_fields', [])
        case_sensitive = options.get('case_sensitive', False)
        check_borders = options.get('check_borders', True)
        min_border_ratio = options.get('min_border_ratio', 0.7)
        check_layout = options.get('check_layout', True)
        min_rows = options.get('min_rows', 3)
        min_columns = options.get('min_columns', 2)
        min_formatting_ratio = options.get('min_formatting_ratio', 0.5)
        core_field_ratio = options.get('core_field_ratio', 1.0)
        
        # Validate that core_fields is provided
        if not core_fields or len(core_fields) == 0:
            logger.error("core_fields must be provided in options")
            return 0.0
        
        logger.info(f"Verifying template fields in file: {result}")
        logger.info(f"Core fields (required): {core_fields}")
        logger.info(f"Optional fields: {optional_fields}")
        logger.info(f"Case sensitive: {case_sensitive}")
        logger.info(f"Check borders: {check_borders}, Min border ratio: {min_border_ratio}")
        logger.info(f"Check layout: {check_layout}, Min rows: {min_rows}, Min columns: {min_columns}, Min formatting ratio: {min_formatting_ratio}")
        logger.info(f"Core field ratio: {core_field_ratio:.2%} (must find at least {core_field_ratio:.2%} of core fields)")
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(result)
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            return 0.0
        
        # Step 1: Collect all text and field cell positions
        all_text_values, field_cells = collect_text_and_field_cells(
            wb, core_fields, optional_fields, case_sensitive
        )
        
        # Step 2: Verify field existence
        found_core_fields, missing_core_fields, found_optional_fields, missing_optional_fields = \
            verify_field_existence(core_fields, optional_fields, all_text_values, case_sensitive)
        
        # Verify core field ratio
        if len(core_fields) > 0:
            found_ratio = len(found_core_fields) / len(core_fields)
            logger.info(f"Core field ratio: {len(found_core_fields)}/{len(core_fields)} = {found_ratio:.2%}")
            
            if found_ratio < core_field_ratio:
                logger.error(f"Core field ratio {found_ratio:.2%} is below minimum {core_field_ratio:.2%}")
                logger.error(f"Missing {len(missing_core_fields)} core field(s): {missing_core_fields}")
                logger.error(f"Found {len(found_core_fields)} out of {len(core_fields)} core fields")
                return 0.0
            else:
                logger.info(f"✓ Core field ratio check passed: {found_ratio:.2%} >= {core_field_ratio:.2%}")
        else:
            logger.error("No core fields specified")
            return 0.0
        
        # Step 3: Verify borders (if enabled)
        border_check_passed = True
        fields_with_borders = 0
        total_field_cells = 0
        if check_borders:
            border_check_passed, fields_with_borders, total_field_cells = \
                verify_borders(found_core_fields, field_cells, min_border_ratio)
            if not border_check_passed:
                return 0.0
        
        # Step 4: Verify layout and formatting (if enabled)
        layout_check_passed = True
        unique_rows = 0
        unique_cols = 0
        formatting_ratio = 0.0
        if check_layout:
            layout_check_passed, unique_rows, unique_cols, formatting_ratio = \
                verify_layout_and_formatting(found_core_fields, field_cells, 
                                            min_rows, min_columns, min_formatting_ratio)
            if not layout_check_passed:
                return 0.0
        
        # Verification passed
        logger.info("=" * 60)
        logger.info(f"✓ Template verification passed")
        logger.info(f"  - Core fields: {len(found_core_fields)}/{len(core_fields)} found")
        logger.info(f"    Found: {', '.join(found_core_fields)}")
        logger.info(f"  - Optional fields: {len(found_optional_fields)}/{len(optional_fields)} found")
        if found_optional_fields:
            logger.info(f"    Found: {', '.join(found_optional_fields)}")
        if missing_optional_fields:
            logger.info(f"    Missing: {', '.join(missing_optional_fields)}")
        if check_borders and total_field_cells > 0:
            logger.info(f"  - Borders: {fields_with_borders}/{total_field_cells} cells have borders")
        if check_layout:
            logger.info(f"  - Layout: Fields in {unique_rows} rows and {unique_cols} columns")
        logger.info("=" * 60)
        return 1.0
        
    except Exception as e:
        import traceback
        logger.error(f"Verification failed: {e}")
        logger.error(traceback.format_exc())
        return 0.0
